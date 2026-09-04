"""
ENPRIZON LINDI PROJECT — Flask 主入口
"""
import json, os, sys, socket, io, time, secrets, re
from datetime import datetime, timezone, timedelta
from flask import Flask, jsonify, request, send_from_directory, render_template, send_file, session, redirect, url_for, g
from werkzeug.middleware.proxy_fix import ProxyFix
from werkzeug.utils import secure_filename
from functools import wraps

# 业务时区：坦桑尼亚 UTC+3（服务器可能为其他时区，全系统统一以此为准）
EAT = timezone(timedelta(hours=3))

# ── 会话月份解析（P2 session view_month） ──────────────────
# 优先级：?month= 查询参数 > session['view_month'] > MONTH_CACHE 当前 > EAT.now 默认
# 依赖 login_required：session['view_month'] 仅在登录后由 POST /set-month 写入，匿名请求回退到默认
MONTH_RE = re.compile(r'^\d{4}-(0[1-9]|1[0-2])$')

app = Flask(__name__)
app.config['PREFERRED_URL_SCHEME'] = 'http'
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_prefix=1)

def _resolve_secret_key():
    """Resolve Flask secret key with production fail-fast and dev pinning.

    - If KILWA_SECRET_KEY env is set, use it directly.
    - In production (ENV=production|prod or KILWA_SECRET_KEY_REQUIRED=1) without
      KILWA_SECRET_KEY, fail-fast so misconfiguration is visible immediately
      instead of silently invalidating all sessions on restart.
    - In development, pin a random key to data/.kilwa_secret (gitignored) so
      the second restart reuses the same key and keeps login sessions alive.
    """
    env_key = os.environ.get('KILWA_SECRET_KEY')
    if env_key:
        return env_key
    is_prod = os.environ.get('ENV', '').lower() in ('production', 'prod') or os.environ.get('KILWA_SECRET_KEY_REQUIRED') == '1'
    if is_prod:
        raise RuntimeError("KILWA_SECRET_KEY must be set in production (ENV=production or KILWA_SECRET_KEY_REQUIRED=1)")
    # dev fallback: pin to file so restarts keep the same key
    secret_path = os.path.join(os.path.dirname(__file__), 'data', '.kilwa_secret')
    try:
        os.makedirs(os.path.dirname(secret_path), exist_ok=True)
        if os.path.exists(secret_path):
            with open(secret_path, 'r') as f:
                cached = f.read().strip()
            if cached:
                return cached
        # generate and persist
        new_key = secrets.token_hex(32)
        with open(secret_path, 'w') as f:
            f.write(new_key)
        try:
            os.chmod(secret_path, 0o600)
        except Exception:
            pass
        print(f"[WARN] KILWA_SECRET_KEY not set — generated and pinned to {secret_path} (dev only)", file=sys.stderr)
        return new_key
    except RuntimeError:
        raise
    except Exception as e:
        # If file pinning fails, warn and fall back to ephemeral key (session will reset)
        print(f"[WARN] Failed to pin secret to {secret_path}: {e} — using ephemeral key", file=sys.stderr)
        return secrets.token_hex(32)

app.secret_key = _resolve_secret_key()
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB 上传上限
# 会话 Cookie 安全加固（防 CSRF/劫持：SameSite 严格 + HttpOnly；Secure 仅 HTTPS 时开启）
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['SESSION_COOKIE_SECURE'] = os.environ.get('KILWA_HTTPS', '') == '1'
app.config['SESSION_COOKIE_HTTPONLY'] = True

@app.context_processor
def inject_static_url():
    prefix = os.environ.get('KILWA_SCRIPT_NAME', '')
    def _static(filename):
        return f'{prefix}/static/{filename}'
    return dict(static_url=_static)

def _current_version():
    """P33: 版本号 = git 短哈希（跨 gunicorn worker 稳定）；无 git 时回退启动时间戳"""
    try:
        import subprocess
        _base = os.path.dirname(os.path.abspath(__file__))
        _h = subprocess.check_output(['git', '-C', _base, 'rev-parse', '--short', 'HEAD'],
                                     stderr=subprocess.DEVNULL, timeout=3).decode().strip()
        if _h:
            return _h
    except Exception:
        pass
    return str(int(time.time()))

APP_VERSION = _current_version()

# 禁用浏览器缓存，确保每次加载最新数据
@app.after_request
def disable_cache(response):
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response

BASE_DIR = os.path.dirname(__file__)
app.config['UPLOAD_FOLDER'] = os.path.join(BASE_DIR, 'uploads')
app.config['DATA_FOLDER'] = os.path.join(BASE_DIR, 'data')
# DEPRECATED: 纯采集模式已移除 Excel 源，仅首次种子回退保留空 data/source 目录（P3 审计 #1）
SOURCE_DIR = os.path.join(BASE_DIR, 'data', 'source')

# ── 硬排除名单（这5人全局不显示、不计薪） ─────────────
HARD_EXCLUDE_IDS = set()
for raw_name in ['Eric Wang QM', 'JIMMY', 'Set sail', '宋家成（Daria）', '宋科举KEJU', '宋科举']:
    from core.namematch import make_employee_id
    eid = make_employee_id(raw_name)
    if eid:
        HARD_EXCLUDE_IDS.add(eid)

# ── 登录认证 ────────────────────────────────────────────

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('logged_in'):
            return jsonify({'ok': False, 'error': 'unauthorized', 'need_login': True}), 401
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    """仅允许 admin 及以上角色执行的操作（editor/viewer 不可用）"""
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('logged_in'):
            return jsonify({'ok': False, 'error': 'unauthorized', 'need_login': True}), 401
        from core.database import ROLE_LEVELS
        lvl = ROLE_LEVELS.get(session.get('role', ''), 0)
        if lvl < ROLE_LEVELS['admin']:
            return jsonify({'ok': False, 'error': 'forbidden', 'need_admin': True}), 403
        return f(*args, **kwargs)
    return decorated

def editor_required(f):
    """仅允许 editor 及以上角色执行的操作（viewer 不可用）"""
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('logged_in'):
            return jsonify({'ok': False, 'error': 'unauthorized', 'need_login': True}), 401
        from core.database import ROLE_LEVELS
        lvl = ROLE_LEVELS.get(session.get('role', ''), 0)
        if lvl < ROLE_LEVELS['editor']:
            return jsonify({'ok': False, 'error': 'forbidden', 'need_admin': True}), 403
        return f(*args, **kwargs)
    return decorated

def super_admin_required(f):
    """仅允许 super_admin 执行的操作"""
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('logged_in'):
            return jsonify({'ok': False, 'error': 'unauthorized', 'need_login': True}), 401
        from core.database import ROLE_LEVELS
        lvl = ROLE_LEVELS.get(session.get('role', ''), 0)
        if lvl < ROLE_LEVELS['super_admin']:
            return jsonify({'ok': False, 'error': 'forbidden', 'need_admin': True}), 403
        return f(*args, **kwargs)
    return decorated

def require_permission(module, action):
    """细粒度权限检查：角色继承 + 单独授权"""
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            if not session.get('logged_in'):
                return jsonify({'ok': False, 'error': 'unauthorized', 'need_login': True}), 401
            from core.database import check_permission
            username = session.get('username', '')
            if not check_permission(app.config['DATA_FOLDER'], username, module, action):
                _audit('perm_denied', '', json.dumps({'user': username, 'module': module, 'action': action}))
                return jsonify({'ok': False, 'error': 'forbidden', 'need_permission': module}), 403
            return f(*args, **kwargs)
        return decorated
    return decorator

@app.route('/api/login', methods=['POST'])
def api_login():
    from core.database import verify_admin, has_admin, set_admin_password
    data = request.get_json(silent=True) or {}
    username = data.get('username', '').strip()
    password = data.get('password', '')

    if not has_admin(app.config['DATA_FOLDER']):
        return jsonify({'ok': False, 'error': 'no_admin', 'need_setup': True})

    if verify_admin(app.config['DATA_FOLDER'], username, password):
        # 登录前重建会话，防 session fixation（攻击者预置的 cookie 在登录后作废）
        session.clear()
        session['logged_in'] = True
        session['username'] = username
        from core.database import get_user_role
        # 安全修复：角色查询失败时回退到最低权限 viewer，而非 admin
        # （原为 or 'admin'，一旦 get_user_role 返回 None 即越权为管理员）
        session['role'] = get_user_role(app.config['DATA_FOLDER'], username) or 'viewer'
        _audit('login', '', json.dumps({'user': username}))
        return jsonify({'ok': True})
    _audit('login_fail', '', json.dumps({'user': username}))
    return jsonify({'ok': False, 'error': 'invalid_credentials'})

@app.route('/api/logout', methods=['POST'])
def api_logout():
    _audit('logout', '', json.dumps({'user': session.get('username', '')}))
    session.clear()
    return jsonify({'ok': True})

@app.route('/api/auth/status', methods=['GET'])
def auth_status():
    from core.database import has_admin, get_user_permissions
    return jsonify({
        'logged_in': session.get('logged_in', False),
        'username': session.get('username', ''),
        'role': session.get('role', ''),
        'has_admin': has_admin(app.config['DATA_FOLDER']),
        # P18 阶段2: 用户有效权限摘要(如 ['dashboard:view','salary:view'])
        'permissions': get_user_permissions(app.config['DATA_FOLDER'], session.get('username', ''))
            if session.get('logged_in') else [],
    })

@app.route('/api/version', methods=['GET'])
def api_version():
    """轻量版本心跳 — 前端轮询/回前台检测部署更新（无载荷、免登录，仅返回进程版本号）"""
    return jsonify({'ok': True, 'version': APP_VERSION})

@app.route('/api/admin/setup', methods=['POST'])
def admin_setup():
    from core.database import has_admin, set_admin_password
    if has_admin(app.config['DATA_FOLDER']):
        return jsonify({'ok': False, 'error': 'admin_exists'})
    data = request.get_json(silent=True) or {}
    username = data.get('username', '').strip()
    password = data.get('password', '')
    if not username or not password:
        return jsonify({'ok': False, 'error': 'missing_fields'})
    if len(password) < 4:
        return jsonify({'ok': False, 'error': 'password_too_short'})
    set_admin_password(app.config['DATA_FOLDER'], username, password)
    session['logged_in'] = True
    session['username'] = username
    session['role'] = 'super_admin'
    _audit('admin_setup', '', json.dumps({'user': username}))
    return jsonify({'ok': True})

@app.route('/api/admin/change-password', methods=['POST'])
@login_required
def admin_change_password():
    from core.database import verify_admin, set_admin_password
    data = request.get_json(silent=True) or {}
    old_password = data.get('old_password', '')
    new_password = data.get('new_password', '')
    if not new_password or len(new_password) < 6:
        return jsonify({'ok': False, 'error': 'password_too_short'})
    if not verify_admin(app.config['DATA_FOLDER'], session['username'], old_password):
        return jsonify({'ok': False, 'error': 'invalid_old_password'})
    set_admin_password(app.config['DATA_FOLDER'], session['username'], new_password)
    _audit('password_change', '', json.dumps({'user': session['username']}))
    return jsonify({'ok': True})

# ═══════════════════════════════════════════════════════════
#  API: 用户角色管理（仅 super_admin）
# ═══════════════════════════════════════════════════════════

@app.route('/admin/users', methods=['GET'])
@super_admin_required
def list_users():
    from core.database import list_all_users
    users = list_all_users(app.config['DATA_FOLDER'])
    return jsonify({'ok': True, 'users': users})

@app.route('/admin/users/role', methods=['POST'])
@super_admin_required
def update_user_role():
    from core.database import set_user_role, ROLE_LEVELS
    data = request.get_json(silent=True) or {}
    username = data.get('username', '').strip()
    role = data.get('role', '').strip()
    if not username or not role:
        return jsonify({'ok': False, 'error': '无效参数'}), 400
    if role == 'editor':
        return jsonify({'ok': False, 'error': 'editor 角色已弃用，请选择 super_admin/admin/collector/applicant/viewer'}), 400
    try:
        set_user_role(app.config['DATA_FOLDER'], username, role)
        _audit('role_change', username, json.dumps({'new_role': role}))
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 400

@app.route('/admin/users/create', methods=['POST'])
@super_admin_required
def create_user():
    """超级管理员新增登录用户（含角色）"""
    from core.database import create_admin_user
    data = request.get_json(silent=True) or {}
    username = (data.get('username') or '').strip()
    password = data.get('password') or ''
    role = (data.get('role') or 'viewer').strip()
    if role == 'editor':
        return jsonify({'ok': False, 'error': 'editor 角色已弃用，请选择 super_admin/admin/collector/applicant/viewer'}), 400
    result = create_admin_user(app.config['DATA_FOLDER'], username, password, role)
    if result == 'ok':
        _audit('user_create', username, json.dumps({'role': role}))
        return jsonify({'ok': True})
    if result == 'exists':
        return jsonify({'ok': False, 'error': '用户名已存在'}), 400
    if result == 'invalid_role':
        return jsonify({'ok': False, 'error': '无效角色'}), 400
    return jsonify({'ok': False, 'error': '用户名或密码不合法（密码至少6位）'}), 400

@app.route('/admin/users/change-password', methods=['POST'])
@super_admin_required
def change_user_password():
    """超级管理员修改指定用户密码（不要求旧密码）"""
    from core.database import reset_admin_password
    data = request.get_json(silent=True) or {}
    username = (data.get('username') or '').strip()
    new_password = data.get('new_password') or ''
    result = reset_admin_password(app.config['DATA_FOLDER'], username, new_password)
    if result == 'ok':
        _audit('user_password_reset', username, json.dumps({}))
        return jsonify({'ok': True})
    if result == 'not_found':
        return jsonify({'ok': False, 'error': '用户不存在'}), 404
    return jsonify({'ok': False, 'error': '密码至少6位'}), 400

@app.route('/admin/users/rename', methods=['POST'])
@super_admin_required
def rename_user():
    """超级管理员重命名用户（同步 user_grants 授权 + approval_routes 审批人）"""
    from core.database import rename_admin_user
    data = request.get_json(silent=True) or {}
    old_username = (data.get('username') or '').strip()
    new_username = (data.get('new_username') or '').strip()
    result = rename_admin_user(app.config['DATA_FOLDER'], old_username, new_username)
    if result == 'ok':
        _audit('user_rename', old_username, json.dumps({'new_username': new_username}))
        return jsonify({'ok': True})
    if result == 'not_found':
        return jsonify({'ok': False, 'error': '用户不存在'}), 404
    if result == 'exists':
        return jsonify({'ok': False, 'error': '用户名已存在'}), 400
    return jsonify({'ok': False, 'error': '用户名不合法'}), 400

@app.route('/admin/users/delete', methods=['POST'])
@super_admin_required
def delete_user():
    """超级管理员删除用户（自动清理 user_grants 授权 + approval_routes 审批人）"""
    from core.database import delete_admin_user
    data = request.get_json(silent=True) or {}
    username = (data.get('username') or '').strip()
    result = delete_admin_user(app.config['DATA_FOLDER'], username)
    if result == 'ok':
        _audit('user_delete', username, json.dumps({}))
        return jsonify({'ok': True})
    if result == 'not_found':
        return jsonify({'ok': False, 'error': '用户不存在'}), 404
    if result == 'last_super_admin':
        return jsonify({'ok': False, 'error': '不能删除最后一个超级管理员'}), 400
    return jsonify({'ok': False, 'error': '删除失败'}), 400

# ═══════════════════════════════════════════════════════════
#  API: 细粒度权限管理（super_admin 管理授权）
# ═══════════════════════════════════════════════════════════

@app.route('/api/permissions/users', methods=['GET'])
@admin_required
def api_permissions_users():
    """P18C: 用户列表 + 角色 + permissions 摘要 + 单用户 grants 明细"""
    from core.database import (list_all_users, get_user_permissions_summary,
                               get_conn, PERMISSION_CATALOG)
    conn = get_conn(app.config['DATA_FOLDER'])
    users = list_all_users(app.config['DATA_FOLDER'])
    try:
        for u in users:
            u['permissions'] = get_user_permissions_summary(app.config['DATA_FOLDER'], u['username'])
            rows = conn.execute(
                """SELECT p.module, p.action, g.grant_type
                   FROM user_grants g JOIN permissions p ON g.permission_id = p.id
                   WHERE g.username=? ORDER BY p.module, p.action""",
                (u['username'],)).fetchall()
            u['grants'] = [{
                'module': r['module'], 'action': r['action'], 'grant_type': r['grant_type'],
                'name': (PERMISSION_CATALOG.get(r['module'] + ':' + r['action']) or {})
                        .get('name', r['module'] + ':' + r['action']),
            } for r in rows]
    finally:
        conn.close()
    return jsonify({'ok': True, 'users': users})

@app.route('/api/permissions/roles', methods=['GET'])
@super_admin_required
def api_permissions_roles():
    """P18b+P18D: 角色权限列表(含继承展开 + 来源标记 + 元数据 + 角色 CRUD 元信息)

    query: ?role=X 只返回该角色(含继承展开 effective + source_role)
    """
    from core.database import (get_conn, ROLE_LEVELS, ROLE_HIERARCHY,
                               ROLE_DEFAULT_PERMISSIONS, PERMISSION_CATALOG,
                               get_role_permissions, BUILTIN_ROLES)
    conn = get_conn(app.config['DATA_FOLDER'])
    try:
        rows = conn.execute(
            "SELECT role, module, action, allow FROM role_permissions ORDER BY role, module, action"
        ).fetchall()
        # 权限项计数(allow=1)
        perm_counts = {}
        for r in rows:
            if r['allow'] == 1:
                perm_counts[r['role']] = perm_counts.get(r['role'], 0) + 1
        # 用户分配计数
        user_counts = {}
        for r in conn.execute("SELECT role, COUNT(*) c FROM admin_users GROUP BY role").fetchall():
            user_counts[r['role']] = r['c']
    finally:
        conn.close()

    table_roles = set(perm_counts.keys())
    all_roles = sorted(set(table_roles) | set(ROLE_LEVELS.keys()))

    def _inherit_chain(r):
        chain = [r]
        q = list(ROLE_HIERARCHY.get(r, []))
        while q:
            x = q.pop(0)
            chain.append(x)
            q.extend(ROLE_HIERARCHY.get(x, []))
        return chain

    def _source_of(r, module, action):
        # 该权限项来自继承链上哪一级(自身或下级角色);表里 allow=1 的第一级
        for rr in _inherit_chain(r):
            for row in rows:
                if row['role'] == rr and row['module'] == module and row['action'] == action and row['allow'] == 1:
                    return rr
        return r

    role_filter = request.args.get('role', '').strip()
    inherited = {}
    for r in all_roles:
        perms = get_role_permissions(app.config['DATA_FOLDER'], r)
        inherited[r] = [
            {'module': m, 'action': a,
             'source_role': _source_of(r, m, a) if r != 'super_admin' else 'super_admin'}
            for m, acts in perms.items() for a in acts if m != '*'
        ]
    resp = {
        'ok': True,
        'roles': [{'role': r, 'builtin': (r in BUILTIN_ROLES or r == 'super_admin'),
                   'permission_count': perm_counts.get(r, 0), 'user_count': user_counts.get(r, 0)}
                  for r in all_roles],
        'hierarchy': ROLE_HIERARCHY,
        'catalog': PERMISSION_CATALOG,
        'defaults': ROLE_DEFAULT_PERMISSIONS,
        'permissions': [dict(r) for r in rows],
        'effective': inherited,
    }
    if role_filter:
        if role_filter not in all_roles:
            return jsonify({'ok': False, 'error': 'unknown_role'}), 400
        resp['effective'] = {role_filter: inherited[role_filter]}
        resp['roles'] = [x for x in resp['roles'] if x['role'] == role_filter]
    return jsonify(resp)

@app.route('/api/permissions/roles', methods=['POST'])
@super_admin_required
def api_permissions_roles_create():
    """P18D: 新增角色 — body {name, permissions:[{module,action,allow}]}"""
    from core.database import get_conn, ROLE_LEVELS, PERMISSION_CATALOG
    data = request.get_json(silent=True) or {}
    name = (data.get('name') or '').strip()
    if not name:
        return jsonify({'ok': False, 'error': 'missing_fields'}), 400
    conn = get_conn(app.config['DATA_FOLDER'])
    try:
        # 冲突: 内置角色名 / 已存在角色名
        if name in ROLE_LEVELS:
            return jsonify({'ok': False, 'error': 'role_name_exists'}), 400
        exists = conn.execute("SELECT 1 FROM role_permissions WHERE role=? LIMIT 1", (name,)).fetchone()
        if exists:
            return jsonify({'ok': False, 'error': 'role_name_exists'}), 400
        permissions = data.get('permissions') or []
        for p in permissions:
            module = (p.get('module') or '').strip()
            action = (p.get('action') or '').strip()
            if module and action:
                conn.execute(
                    "INSERT OR REPLACE INTO role_permissions (role, module, action, allow) VALUES (?,?,?,?)",
                    (name, module, action, 1 if int(p.get('allow', 1)) else 0))
        conn.commit()
    finally:
        conn.close()
    return jsonify({'ok': True})

@app.route('/api/permissions/roles/rename', methods=['PUT'])
@super_admin_required
def api_permissions_roles_rename():
    """P18D: 重命名角色 — body {old, new},同步 role_permissions + admin_users"""
    from core.database import get_conn, ROLE_LEVELS, BUILTIN_ROLES
    data = request.get_json(silent=True) or {}
    old = (data.get('old') or '').strip()
    new = (data.get('new') or '').strip()
    if not old or not new:
        return jsonify({'ok': False, 'error': 'missing_fields'}), 400
    conn = get_conn(app.config['DATA_FOLDER'])
    try:
        # old 必须存在(内置或自定义)
        exists_old = conn.execute("SELECT 1 FROM role_permissions WHERE role=? LIMIT 1", (old,)).fetchone()
        if not exists_old and old not in ROLE_LEVELS:
            return jsonify({'ok': False, 'error': 'unknown_role'}), 400
        # new 不能与内置或其他角色冲突
        if new in ROLE_LEVELS:
            return jsonify({'ok': False, 'error': 'role_name_exists'}), 400
        exists_new = conn.execute("SELECT 1 FROM role_permissions WHERE role=? LIMIT 1", (new,)).fetchone()
        if exists_new:
            return jsonify({'ok': False, 'error': 'role_name_exists'}), 400
        conn.execute("UPDATE role_permissions SET role=? WHERE role=?", (new, old))
        conn.execute("UPDATE admin_users SET role=? WHERE role=?", (new, old))
        conn.commit()
    finally:
        conn.close()
    return jsonify({'ok': True})

@app.route('/api/permissions/roles/<role>', methods=['DELETE'])
@super_admin_required
def api_permissions_roles_delete(role):
    """P18D: 删除角色 — 内置拒绝,有用户拒绝"""
    from core.database import get_conn, BUILTIN_ROLES
    conn = get_conn(app.config['DATA_FOLDER'])
    try:
        if role in BUILTIN_ROLES or role == 'super_admin':
            return jsonify({'ok': False, 'error': 'role_builtin_nodel'}), 400
        cnt = conn.execute("SELECT COUNT(*) c FROM admin_users WHERE role=?", (role,)).fetchone()
        if cnt and cnt['c'] > 0:
            return jsonify({'ok': False, 'error': 'role_has_users'}), 400
        conn.execute("DELETE FROM role_permissions WHERE role=?", (role,))
        conn.commit()
    finally:
        conn.close()
    return jsonify({'ok': True})

@app.route('/api/permissions/roles/reset', methods=['POST'])
@super_admin_required
def api_permissions_roles_reset():
    """P18b: 将指定角色 role_permissions 重置为 ROLE_DEFAULT_PERMISSIONS 默认"""
    from core.database import get_conn, ROLE_DEFAULT_PERMISSIONS
    data = request.get_json(silent=True) or {}
    role = (data.get('role') or '').strip()
    if role not in ROLE_DEFAULT_PERMISSIONS or role == 'super_admin':
        return jsonify({'ok': False, 'error': 'invalid_role'}), 400
    conn = get_conn(app.config['DATA_FOLDER'])
    try:
        conn.execute("DELETE FROM role_permissions WHERE role=?", (role,))
        grants = ROLE_DEFAULT_PERMISSIONS[role]
        for module, actions in grants.items():
            if module == '*':
                continue
            for action in actions:
                conn.execute(
                    "INSERT OR REPLACE INTO role_permissions (role, module, action, allow) VALUES (?,?,?,1)",
                    (role, module, action))
    finally:
        conn.commit()
        conn.close()
    return jsonify({'ok': True})

@app.route('/api/permissions/roles', methods=['PUT'])
@super_admin_required
def api_permissions_roles_put():
    """P18: 角色权限编辑 — 写 role_permissions(REPLACE)

    body: {role, module, action, allow} 或 {updates:[{role,module,action,allow}, ...]}
    """
    from core.database import get_conn
    data = request.get_json(silent=True) or {}
    updates = data.get('updates') if isinstance(data.get('updates'), list) else [data]
    if not updates:
        return jsonify({'ok': False, 'error': 'missing_fields'}), 400
    conn = get_conn(app.config['DATA_FOLDER'])
    try:
        for u in updates:
            role = (u.get('role') or '').strip()
            module = (u.get('module') or '').strip()
            action = (u.get('action') or '').strip()
            if not role or not module or not action:
                return jsonify({'ok': False, 'error': 'missing_fields'}), 400
            if role == 'super_admin':
                return jsonify({'ok': False, 'error': 'super_admin_immutable'}), 400
            allow = 1 if int(u.get('allow', 1)) else 0
            conn.execute(
                "INSERT OR REPLACE INTO role_permissions (role, module, action, allow) VALUES (?,?,?,?)",
                (role, module, action, allow))
    finally:
        conn.commit()
        conn.close()
    return jsonify({'ok': True})

@app.route('/api/permissions/grant', methods=['POST'])
@super_admin_required
def api_permissions_grant():
    from core.database import grant_user_permission
    data = request.get_json(silent=True) or {}
    username = data.get('username', '').strip()
    module = data.get('module', '').strip()
    action = data.get('action', '').strip()
    grant_type = data.get('grant_type', 'allow')
    if not username or not module or not action:
        return jsonify({'ok': False, 'error': 'missing_fields'}), 400
    try:
        grant_user_permission(app.config['DATA_FOLDER'], username, module, action, grant_type)
        _audit('perm_grant', username, json.dumps({'module': module, 'action': action, 'type': grant_type}))
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 400

@app.route('/api/permissions/grant', methods=['DELETE'])
@super_admin_required
def api_permissions_revoke():
    from core.database import revoke_user_grant_by_action
    data = request.get_json(silent=True) or {}
    username = data.get('username', '').strip()
    module = data.get('module', '').strip()
    action = data.get('action', '').strip()
    if not username or not module or not action:
        return jsonify({'ok': False, 'error': 'missing_fields'}), 400
    try:
        revoke_user_grant_by_action(app.config['DATA_FOLDER'], username, module, action)
        _audit('perm_revoke', username, json.dumps({'module': module, 'action': action}))
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 400

@app.route('/api/permissions/grants/clear', methods=['POST'])
@super_admin_required
def api_permissions_grants_clear():
    """P29-c2: 一键清空某用户全部单独授权, 恢复纯角色默认权限"""
    from core.database import clear_user_grants
    data = request.get_json(silent=True) or {}
    username = (data.get('username') or '').strip()
    if not username:
        return jsonify({'ok': False, 'error': 'missing_username'}), 400
    try:
        cleared = clear_user_grants(app.config['DATA_FOLDER'], username)
        _audit('grants_clear', username, json.dumps({'cleared': cleared}))
        return jsonify({'ok': True, 'cleared': cleared})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 400

@app.route('/api/permissions/init-defaults', methods=['POST'])
@super_admin_required
def api_permissions_init_defaults():
    from core.database import init_default_permissions
    init_default_permissions(app.config['DATA_FOLDER'])
    _audit('perm_init_defaults', '', '{}')
    return jsonify({'ok': True})

# ═══════════════════════════════════════════════════════════
#  API: 全局搜索
# ═══════════════════════════════════════════════════════════

@app.route('/api/search', methods=['GET'])
@login_required
def global_search():
    q = request.args.get('q', '').strip()
    scope = request.args.get('scope', 'all')
    if len(q) < 2:
        return jsonify({'ok': False, 'error': 'query_too_short', 'results': []})
    from core.database import search_all
    results = search_all(app.config['DATA_FOLDER'], q, scope)
    return jsonify({'ok': True, 'results': results, 'q': q, 'scope': scope})

# ═══════════════════════════════════════════════════════════
#  API: 表单自定义管理
# ═══════════════════════════════════════════════════════════

@app.route('/api/forms/schemas', methods=['GET'])
@login_required
def api_form_schemas():
    from core.database import list_form_schemas
    schemas = list_form_schemas(app.config['DATA_FOLDER'])
    return jsonify({'ok': True, 'schemas': schemas})

@app.route('/api/forms/schema/<int:schema_id>', methods=['GET'])
@login_required
def api_form_schema(schema_id):
    from core.database import get_form_schema
    schema = get_form_schema(app.config['DATA_FOLDER'], schema_id)
    if not schema:
        return jsonify({'ok': False, 'error': 'not_found'}), 404
    return jsonify({'ok': True, 'schema': schema})

@app.route('/api/forms/schema', methods=['POST'])
@super_admin_required
def api_create_form_schema():
    from core.database import create_form_schema
    data = request.get_json(silent=True) or {}
    if not data.get('name'):
        return jsonify({'ok': False, 'error': 'missing_name'}), 400
    try:
        sid = create_form_schema(app.config['DATA_FOLDER'], data)
        _audit('form_schema_create', '', json.dumps({'name': data['name']}))
        return jsonify({'ok': True, 'id': sid})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 400

@app.route('/api/forms/schema/<int:schema_id>', methods=['PUT'])
@super_admin_required
def api_update_form_schema(schema_id):
    from core.database import update_form_schema
    data = request.get_json(silent=True) or {}
    update_form_schema(app.config['DATA_FOLDER'], schema_id, data)
    _audit('form_schema_update', '', json.dumps({'id': schema_id}))
    return jsonify({'ok': True})

@app.route('/api/forms/schema/<int:schema_id>', methods=['DELETE'])
@super_admin_required
def api_delete_form_schema(schema_id):
    from core.database import delete_form_schema
    delete_form_schema(app.config['DATA_FOLDER'], schema_id)
    _audit('form_schema_delete', '', json.dumps({'id': schema_id}))
    return jsonify({'ok': True})

@app.route('/api/forms/schema/<int:schema_id>/fields', methods=['POST'])
@super_admin_required
def api_add_form_field(schema_id):
    from core.database import add_form_field
    data = request.get_json(silent=True) or {}
    if not data.get('field_key'):
        return jsonify({'ok': False, 'error': 'missing_field_key'}), 400
    add_form_field(app.config['DATA_FOLDER'], schema_id, data)
    _audit('form_field_add', '', json.dumps({'schema_id': schema_id, 'key': data['field_key']}))
    return jsonify({'ok': True})

@app.route('/api/forms/schema/<int:schema_id>/fields/<int:field_id>', methods=['PUT'])
@super_admin_required
def api_update_form_field(schema_id, field_id):
    from core.database import update_form_field
    data = request.get_json(silent=True) or {}
    update_form_field(app.config['DATA_FOLDER'], field_id, data)
    return jsonify({'ok': True})

@app.route('/api/forms/schema/<int:schema_id>/fields/<int:field_id>', methods=['DELETE'])
@super_admin_required
def api_delete_form_field(schema_id, field_id):
    from core.database import delete_form_field
    delete_form_field(app.config['DATA_FOLDER'], field_id)
    return jsonify({'ok': True})

@app.route('/api/forms/seed-defaults', methods=['POST'])
@super_admin_required
def api_seed_default_forms():
    from core.database import seed_default_forms
    seed_default_forms(app.config['DATA_FOLDER'])
    _audit('form_seed_defaults', '', '{}')
    return jsonify({'ok': True})

# ═══════════════════════════════════════════════════════════
#  API: 旧数据归档
# ═══════════════════════════════════════════════════════════

@app.route('/api/archive/months', methods=['GET'])
@login_required
@require_permission('salary', 'view')  # P29 T4 A11
def archive_months():
    from core.database import list_archive_months
    months = list_archive_months(app.config['DATA_FOLDER'])
    # 检查归档是否存在
    import os
    exists = os.path.exists(os.path.join(app.config['DATA_FOLDER'], 'archived_kilwa.db'))
    return jsonify({'ok': True, 'months': months, 'archived': exists})

@app.route('/api/archive/salary', methods=['GET'])
@login_required
@require_permission('salary', 'view')  # P29 T4 A11
def archive_salary():
    from core.database import get_archive_salary
    month = resolve_month(request)
    requested = (request.args.get('month') or '').strip()
    if requested and MONTH_RE.match(requested[:7]):
        month = requested[:7]
    md = _get_month_data(month)
    if not month:
        return jsonify({'ok': False, 'error': 'missing_month'}), 400
    data = get_archive_salary(app.config['DATA_FOLDER'], month)
    if data is None:
        if md is not None and md.get('salary_result') is not None:
            data = md.get('salary_result')
            return jsonify({'ok': True, 'data': data, 'month': month, 'source': 'live'})
        return jsonify({'ok': False, 'error': 'archive_unavailable'}), 404
    return jsonify({'ok': True, 'data': data, 'month': month})

def strip_dept(dept):
    """去掉 ENPRIZON LINDI PROJECT 前缀，保留子部门；纯顶层部门保留原名"""
    if not dept:
        return ''
    if dept == 'ENPRIZON LINDI PROJECT':
        return 'ENPRIZON LINDI PROJECT'
    return dept.replace('ENPRIZON LINDI PROJECT/', '')

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['DATA_FOLDER'], exist_ok=True)
os.makedirs(SOURCE_DIR, exist_ok=True)

# ── 数据状态 ─────────────────────────────────────────────
APP_STATE = {
    'main_file': None,
    'advance_file': None,
    'addressbook_file': None,
    'basic_info_file': None,
    'parsed': False,
    'calculated': False,
    'employees': [],
    'salary_result': None,
    'address_book': {},
    'source_info': {},          # {type: filename} 记录实际加载的源文件
    'month': None,              # 当前筛选的月份 "2026-05"
    '_month_stamp': '',         # P27: month_filter[:7] 原子戳，与 month 同步
    'headless': False,          # 无源数据月份模式
}

# ── P2 月份缓存（MonthData = {main_data, employees, salary_result, config_snapshot, headless, built_at}) ──
import threading as _threading
import copy as _copy
MONTH_CACHE: dict[str, dict] = {}
MONTH_CACHE_LOCK = _threading.Lock()
MONTH_CACHE_LOCKS: dict[str, _threading.Lock] = {}
MONTH_CACHE_MAX = 3

def _get_month_lock(month: str) -> _threading.Lock:
    """Per-month lock (double-checked creation under global lock)."""
    with MONTH_CACHE_LOCK:
        lk = MONTH_CACHE_LOCKS.get(month)
        if lk is None:
            lk = _threading.Lock()
            MONTH_CACHE_LOCKS[month] = lk
        return lk

def _month_cache_evict_if_needed():
    """LRU cap 3: evict oldest built_at (caller must hold MONTH_CACHE_LOCK)."""
    while len(MONTH_CACHE) > MONTH_CACHE_MAX:
        oldest = min(MONTH_CACHE.items(), key=lambda kv: kv[1].get('built_at', 0))[0]
        MONTH_CACHE.pop(oldest, None)
        MONTH_CACHE_LOCKS.pop(oldest, None)


# ── P2 Cache invalidation (per-month) ─────────────────────
# Write endpoint -> invalidated MONTH_CACHE key (per-month only):
#   POST /api/collection/submit               -> month = submission_date[:7] + __all__（2026-09-01 补：__all__ 全量快照失效，防旧快照回写）
#   POST /api/collection/edit/<id>            -> old_date[:7] + new_date[:7] + __all__
#   POST /attendance/toggle                   -> month = body date[:7] (fallback g.view_month) + __all__
#   POST /employees/override                  -> month = effective_from or start_date[:7] or g.view_month（_refresh_employees_cache 已全量重建刷新 APP_STATE）
#   POST /employees/remove-override           -> month = g.view_month（同上）
#   POST /employees/remove-temp-override      -> month = g.view_month（同上）
#   POST /employees/remove-override-by-id     -> month = override row effective_from/start_date or g.view_month（同上）
#   POST /employees/bonus-penalty             -> month = body month[:7]
#   POST /api/salary/inline-edit              -> month = body month[:7]
#   POST /recalculate                         -> month = g.view_month (per-month recompute)
#   POST /reload                              -> full clear (all keys) + recompute current month
# 不变量：MONTH_CACHE miss 后必须走完整 _run_pipeline（含出勤重建+事件推导覆盖），
#        禁止用 __all__/APP_STATE 旧快照 deepcopy-filter 回写（2026-08-31 数据缺失 bug 根因）。

def _invalidate_all_cache_base():
    """失效 __all__ 全量快照（数据事实变更点：采集提交/编辑/出勤标记）。
    __all__ 条目由 _run_pipeline(month_filter=None) 构建，失效后由下一次
    view_month='all' 读自动重建，保证全量视图不携带陈旧数据。"""
    with MONTH_CACHE_LOCK:
        MONTH_CACHE.pop('__all__', None)

def _invalidate_month_cache(month: str | None):
    """Evict single MONTH_CACHE entry for `month` (per-month only). No-op if month invalid."""
    if not month:
        return
    month = (month or '')[:7]
    if not month or not MONTH_RE.match(month):
        return
    with MONTH_CACHE_LOCK:
        MONTH_CACHE.pop(month, None)
        MONTH_CACHE_LOCKS.pop(month, None)


def _clear_all_month_cache():
    """Full clear MONTH_CACHE (for /reload)."""
    with MONTH_CACHE_LOCK:
        MONTH_CACHE.clear()
        MONTH_CACHE_LOCKS.clear()


# ── P27 导出守卫：严格模式开关（默认 lenient，不阻断；置 True 则错月 409）──
EXPORT_STRICT_MONTH = False


def _get_requested_month():
    """DEPRECATED (P3): 薄包装 — 保留供旧测试调用，逻辑归一至 resolve_month。"""
    try:
        return resolve_month(request)
    except Exception:
        return ''


def _month_stamp():
    """当前全局戳（优先 _month_stamp，回退 month[:7]）"""
    s = APP_STATE.get('_month_stamp', '')
    if s:
        return s
    mv = APP_STATE.get('month')
    if mv:
        return (mv or '')[:7]
    return ''


def resolve_month(req=None):
    """会话月份解析器 — 优先级 ?month= > JSON body month(POST) > session['view_month'] > MONTH_CACHE 当前 > EAT.now

    依赖 login_required：session['view_month'] 仅在登录后写入；未登录时跳过该级回退。
    所有读端点应通过此函数或 g.view_month 获取月份，禁止直接读 APP_STATE['month']。
    POST 额外支持 body JSON 中的 month 字段，便于导出类 POST 携带月份。
    """
    if req is None:
        req = request
    try:
        q = (req.args.get('month') or '').strip()[:7]
    except Exception:
        q = ''
    if q and MONTH_RE.match(q):
        return q
    # POST body month support (export etc.)
    try:
        if getattr(req, 'method', '') == 'POST':
            _j = req.get_json(silent=True) or {}
            if isinstance(_j, dict):
                b = (_j.get('month') or '').strip()[:7]
                if b and MONTH_RE.match(b):
                    return b
    except Exception:
        pass
    try:
        sess_m = (session.get('view_month') or '').strip()[:7]
    except Exception:
        sess_m = ''
    if sess_m and MONTH_RE.match(sess_m):
        return sess_m
    with MONTH_CACHE_LOCK:
        if MONTH_CACHE:
            try:
                filtered = {k: v for k, v in MONTH_CACHE.items() if k != '__all__'}
                pool = filtered if filtered else MONTH_CACHE
                cur = max(pool.items(), key=lambda kv: kv[1].get('built_at', 0))[0]
                if cur and MONTH_RE.match(cur):
                    return cur
            except Exception:
                pass
    stamp = _month_stamp()
    if stamp and MONTH_RE.match(stamp):
        return stamp
    return datetime.now(EAT).strftime('%Y-%m')


def _resolve_export_month_data(requested_month):
    """DEPRECATED (P3): 历史 lenient 路径保留供旧测试 — deep-copy 过滤不污染全局。"""
    import copy
    from core.calculator import calculate_all
    from core.exceptions import load_overrides, load_daily_exclusions
    from core.database import load_bonus_penalties as _load_bp
    md_src = APP_STATE.get('main_data') or {}
    emps = APP_STATE.get('employees') or []
    if not md_src and not emps:
        return requested_month, None, {}
    md = copy.deepcopy(md_src)
    for key in ('dates', 'shift_production', 'driller_production', 'attendance', 'crush_production'):
        if md.get(key):
            if key == 'dates':
                md[key] = [d for d in md[key] if d.startswith(requested_month)]
            else:
                md[key] = [d for d in md[key] if d.get('date', '').startswith(requested_month)]
    overrides = load_overrides(app.config['DATA_FOLDER'], month=requested_month)
    exclusions = load_daily_exclusions(app.config['DATA_FOLDER'])
    bonus_penalties = _load_bp(app.config['DATA_FOLDER'], requested_month)
    ug_team_members = _build_ug_team_members(app.config['DATA_FOLDER'])
    try:
        result = calculate_all(md, emps, overrides=overrides, exclusions=exclusions,
                               pricing=APP_STATE.get('config', {}), data_folder=app.config['DATA_FOLDER'],
                               bonus_penalties=bonus_penalties, ug_team_members=ug_team_members)
    except TypeError:
        result = calculate_all(md, emps, overrides=overrides, exclusions=exclusions,
                               pricing=APP_STATE.get('config', {}), data_folder=app.config['DATA_FOLDER'],
                               bonus_penalties=bonus_penalties)
    return requested_month, result, md


def _build_attendance_grid_for_month(requested_month):
    """P3 thin wrapper — single source of truth is _build_attendance_grid.
    Delegates via _get_month_data(requested_month) to avoid duplicated day_status logic."""
    if not requested_month or requested_month == _month_stamp():
        return _build_attendance_grid()
    md_wrap = _get_month_data(requested_month)
    if md_wrap is not None:
        return _build_attendance_grid(md_wrap.get('main_data'), md_wrap.get('employees'))
    return _build_attendance_grid()

def _audit(action, employee_id='', detail='{}', operator=None):
    """写审计日志（快捷包装）"""
    from core.database import log_audit
    if operator is None:
        operator = session.get('username', '')
    log_audit(app.config['DATA_FOLDER'], action, employee_id, detail, operator=operator)

# ═══════════════════════════════════════════════════════════
#  P9: Web 采集数据回填（纯采集模式唯一数据源，见 collection_submit）
# ═══════════════════════════════════════════════════════════
#  核心解析+计算引擎（被 auto_load 和 /reload 复用）
# ═══════════════════════════════════════════════════════════

def _derive_overrides_from_events(data_folder, month):
    """从已批准的 employee_events 推导当月 overrides（事件→覆盖桥接）"""
    from core.database import get_approved_events_for_month
    from calendar import monthrange

    events = get_approved_events_for_month(data_folder, month)
    derived = {}

    if month and len(month) == 7:
        y, m = int(month[:4]), int(month[5:7])
        _, last_day = monthrange(y, m)
        month_end = f'{month}-{last_day:02d}'
    else:
        month_end = ''

    for ev in events:
        eid = ev['employee_id']
        if eid not in derived:
            derived[eid] = []

        try:
            payload = json.loads(ev.get('payload', '{}'))
        except:
            payload = {}

        eff = ev.get('effective_date', '')

        if ev['event_type'] == 'transfer':
            # 调岗 → 临时覆盖（effective_date 到月底）
            new_type = payload.get('new_type', payload.get('salary_type', ''))
            if new_type in ('day_rate', 'monthly', 'piece_underground', 'piece_driller', 'piece_crush'):
                derived[eid].append({
                    'id': f'event_{ev["id"]}',
                    'salary_type': new_type,
                    'day_rate': payload.get('day_rate', 0),
                    'monthly_salary': payload.get('monthly_salary', 0),
                    'start_date': eff,
                    'end_date': month_end,
                    'note': f'OA调岗 #{ev["id"]}',
                    'type': '',
                    'shift': '', 'captain': '',
                    'effective_from': month,
                })
        elif ev['event_type'] == 'salary_change':
            # 薪资变更 → 临时覆盖
            derived[eid].append({
                'id': f'event_{ev["id"]}',
                'salary_type': payload.get('salary_type', ''),
                'day_rate': payload.get('day_rate', 0),
                'monthly_salary': payload.get('monthly_salary', 0),
                'start_date': eff,
                'end_date': month_end,
                'note': f'OA薪资变 #{ev["id"]}',
                'type': '',
                'shift': '', 'captain': '',
                'effective_from': month,
            })
        elif ev['event_type'] in ('resign', 'dismiss'):
            # 离职 → 排除（从 effective_date 开始不计薪）
            derived[eid].append({
                'id': f'event_{ev["id"]}',
                'salary_type': '',
                'day_rate': 0, 'monthly_salary': 0,
                'start_date': eff,
                'end_date': month_end,
                'note': f'OA离职 #{ev["id"]}',
                'type': 'exclusion',
                'shift': '', 'captain': '',
                'effective_from': month,
            })

    return derived

def load_employees_from_db(data_folder):
    """纯采集模式：从 employees 表读取员工列表（含薪资覆盖、NSSF、离职过滤、硬排除）"""
    from core.database import get_conn, load_dismissed
    from core.namematch import make_employee_id

    conn = get_conn(data_folder)
    rows = conn.execute("""
        SELECT id, name, department, default_type, day_rate, monthly_salary,
               nssf_enrolled, nssf_number, nida_number, phone, team_id, custom_number, tin_number, status
        FROM employees ORDER BY CAST(id AS INTEGER)
    """).fetchall()
    conn.close()

    employees = []
    for r in rows:
        eid = r['id']
        if not eid or eid in HARD_EXCLUDE_IDS:
            continue
        # status 优先于 dismissed_employees 旧表（rehire 后旧 dismissed 残留不应隐藏）
        try:
            _st = (r['status'] or '').strip().lower()
        except Exception:
            _st = ''
        if _st == 'dismissed':
            continue
        employees.append({
            'id': eid,
            'name': r['name'] or eid,
            'department': r['department'] or '',
            'default_type': r['default_type'] or 'day_rate',
            'source': 'db',
            'override_type': None,
            'overrides': [],
            'day_rate': r['day_rate'] or 0,
            'monthly_salary': r['monthly_salary'] or 0,
            'advance_total': 0,
            'phone': r['phone'] or '',
            'nssf_enrolled': bool(r['nssf_enrolled']),
            'nssf_number': r['nssf_number'] or '',
            'nida_number': r['nida_number'] or '',
            'team_id': r['team_id'] or 0,
            'custom_number': r['custom_number'] or '',
            'tin_number': r['tin_number'] or '',
            'status': r['status'] or 'active',
        })

    # 离职过滤（仅对 status 非 active 的旧数据兜底；active 员工即便在 dismissed_employees 有残留也显示）
    dismissed = load_dismissed(data_folder)
    employees = [e for e in employees if e.get('status') == 'active' or e['id'] not in dismissed]
    return employees


def _resolve_base_from_history(employees, month):
    """月度隔离：按基线台账解析员工当月部门/薪资基线。
    命中台账则覆盖 department/default_type/day_rate/monthly_salary/team_id，
    未命中保持当前 employees 主档（回退基线）。
    必须在 _build_db_ab_index / overrides 叠加之前调用。"""
    if not month or not employees:
        return employees
    from core.database import resolve_base_for_month, ensure_base_history_table
    try:
        ensure_base_history_table(app.config['DATA_FOLDER'])
    except Exception:
        return employees
    for emp in employees:
        eid = emp.get('id')
        try:
            base = resolve_base_for_month(app.config['DATA_FOLDER'], eid, month)
        except Exception:
            base = None
        if not base:
            continue
        if base.get('department') is not None:
            emp['department'] = base['department'] or ''
        if base.get('default_type'):
            emp['default_type'] = base['default_type']
        if base.get('day_rate') is not None:
            emp['day_rate'] = base['day_rate'] or 0
        if base.get('monthly_salary') is not None:
            emp['monthly_salary'] = base['monthly_salary'] or 0
        if base.get('team_id') is not None:
            emp['team_id'] = int(base['team_id'] or 0)
    return employees


def _build_db_ab_index(data_folder):
    """纯采集模式：从 employees 表构建 namematch 索引（_AB_INDEX），替代通讯录 Excel
    使 make_employee_id('EMA BUKWIMBA') 反查到新ID（如 '34'）。
    注意：若 _AB_INDEX 已有通讯录加载的数据（如启动时通讯录种子），本函数用 DB 覆盖/补充，
    DB 优先保证新ID体系一致。
    """
    from core.namematch import _AB_INDEX, _EXTRA_AB_ENTRIES, strip_alias
    from core.database import get_conn
    import re as _re
    _conn = get_conn(data_folder)
    _rows = _conn.execute("SELECT id, name, alias FROM employees").fetchall()
    _conn.close()
    # 保留已有索引（通讯录加载的变体），DB 精确姓名覆盖
    for _eid, _name, _alias in _rows:
        if not _name:
            continue
        _sa = strip_alias(str(_name))
        _key = _re.sub(r'\s+', '', _sa).upper()
        if _key:
            _AB_INDEX[_key] = (str(_eid), _sa)
        # 去最后一个词的短名
        _words = _sa.split()
        if len(_words) > 1:
            _short = _re.sub(r'\s+', '', ' '.join(_words[:-1])).upper()
            _AB_INDEX.setdefault(_short, (str(_eid), _sa))
        # P19: 非空别名也建立索引(setdefault 不覆盖姓名主键)
        if _alias:
            _sa_alias = strip_alias(str(_alias))
            _ak = _re.sub(r'\s+', '', _sa_alias).upper()
            if _ak:
                _AB_INDEX.setdefault(_ak, (str(_eid), _sa_alias))
    # 补充通讯录外人员（离职等）
    for _k, _v in _EXTRA_AB_ENTRIES.items():
        _AB_INDEX.setdefault(_k, _v)


def build_attendance_from_overrides(main_data, data_folder):
    """纯采集模式：从 attendance_overrides 构建 main_data['attendance']
    将 P（出勤）写入 normal 列表，供出勤网格与日薪/月薪计算使用
    """
    from core.database import get_conn
    conn = get_conn(data_folder)
    rows = conn.execute(
        "SELECT employee_id, date FROM attendance_overrides WHERE status='P' ORDER BY date"
    ).fetchall()
    conn.close()
    by_date = {}
    for r in rows:
        by_date.setdefault(r['date'], []).append(r['employee_id'])
    attendance = []
    for dt in sorted(by_date):
        attendance.append({'date': dt, 'normal': by_date[dt], 'leave': [], 'absent': []})
    main_data['attendance'] = attendance


def _run_pipeline(month_filter=None):
    """
    纯采集模式：从数据库重建并计算，落 MONTH_CACHE 并同步 APP_STATE 别名。
    month_filter: "2026-05" 或 None（全部）
    Returns: MonthData dict on success, None on empty employees.
    MonthData = {main_data, employees, salary_result, config_snapshot, headless, built_at, month}
    """
    from core.calculator import calculate_all

    main_data = {
        'shift_production': [], 'driller_production': [],
        'crush_production': [], 'attendance': [], 'dates': [],
        'piece_rate_people': {'driller': set(), 'underground': set()},
        'daily_salary_people': set(),
    }
    rebuild_main_data_from_collections(main_data)
    build_attendance_from_overrides(main_data, app.config.get('DATA_FOLDER'))
    _dates = set(main_data.get('dates', []))
    for _k in ('shift_production', 'driller_production', 'crush_production', 'attendance'):
        for _d in main_data.get(_k, []):
            if _d.get('date'):
                _dates.add(_d['date'])
    main_data['dates'] = sorted(_dates)

    employees = load_employees_from_db(app.config.get('DATA_FOLDER'))
    if not employees:
        return None

    # 月度隔离：按基线台账解析当月部门/薪资基线（须早于 _build_db_ab_index / overrides 叠加）
    if month_filter:
        _resolve_base_from_history(employees, month_filter)

    _build_db_ab_index(app.config.get('DATA_FOLDER'))

    for emp in employees:
        emp['nssf_enrolled'] = bool((emp.get('nssf_number') or '').strip())

    from core.database import load_overrides as _load_ov
    saved_overrides = _load_ov(app.config.get('DATA_FOLDER'), month=month_filter)

    events_overrides = _derive_overrides_from_events(app.config.get('DATA_FOLDER'), month_filter) if month_filter else {}
    for eid, eovs in events_overrides.items():
        if eid not in saved_overrides:
            saved_overrides[eid] = []
        saved_overrides[eid] = eovs + saved_overrides[eid]

    for emp in employees:
        eid = emp.get('id')
        if eid in saved_overrides:
            for o in saved_overrides[eid]:
                has_range = bool(o.get('start_date', '') or o.get('end_date', ''))
                st = o.get('salary_type', '')
                if not has_range and st in ('day_rate', 'monthly', 'piece_underground', 'piece_driller', 'piece_crush'):
                    emp['override_type'] = st
                if not has_range:
                    if st == 'day_rate' and o.get('day_rate', 0) > 0:
                        emp['day_rate'] = o['day_rate']
                    if st == 'monthly' and o.get('monthly_salary', 0) > 0:
                        emp['monthly_salary'] = o['monthly_salary']
        ot = emp.get('override_type')
        if emp.get('default_type') == 'piece_underground':
            pass
        elif ot == 'day_rate':
            emp['monthly_salary'] = 0
        elif ot == 'monthly':
            emp['day_rate'] = 0
        elif ot in ('piece_underground', 'piece_driller', 'piece_crush'):
            emp['day_rate'] = 0
            emp['monthly_salary'] = 0

    advance_data = None

    if month_filter:
        for key in ('dates', 'shift_production', 'driller_production', 'attendance', 'crush_production'):
            if main_data.get(key):
                if key == 'dates':
                    main_data[key] = [d for d in main_data[key] if d.startswith(month_filter)]
                else:
                    main_data[key] = [d for d in main_data[key] if d.get('date', '').startswith(month_filter)]

    cfg = APP_STATE.get('config')
    if not cfg:
        from core.pricing import load_config
        cfg = load_config(app.config.get('DATA_FOLDER'))
        APP_STATE.update({'config': cfg})

    from core.exceptions import load_overrides as _load_override_ov, load_daily_exclusions as _load_excl
    from core.database import load_bonus_penalties as _load_bp
    overrides = _load_override_ov(app.config.get('DATA_FOLDER'), month=month_filter)
    exclusions = _load_excl(app.config.get('DATA_FOLDER'))
    bonus_penalties = _load_bp(app.config.get('DATA_FOLDER'), month_filter) if month_filter else {}
    ug_team_members = _build_ug_team_members(app.config.get('DATA_FOLDER'), month=month_filter)
    try:
        result = calculate_all(main_data, employees, overrides=overrides, exclusions=exclusions,
                               pricing=cfg, data_folder=app.config.get('DATA_FOLDER'),
                               bonus_penalties=bonus_penalties, ug_team_members=ug_team_members)
    except TypeError:
        result = calculate_all(main_data, employees, overrides=overrides, exclusions=exclusions,
                               pricing=cfg, data_folder=app.config.get('DATA_FOLDER'),
                               bonus_penalties=bonus_penalties)

    headless = not bool(main_data.get('dates'))
    if month_filter and headless and employees:
        import calendar as _cal
        try:
            y, m = int(month_filter[:4]), int(month_filter[5:7])
            _, last_day = _cal.monthrange(y, m)
            generated = [f'{month_filter}-{d:02d}' for d in range(1, last_day + 1)]
            main_data['dates'] = generated
            main_data['shift_production'] = []
            main_data['driller_production'] = []
            main_data['attendance'] = []
            headless = True
        except Exception:
            pass
    else:
        headless = not bool(main_data.get('dates')) if month_filter else False

    built_at = time.time()
    month_key = (month_filter or '')[:7] if month_filter else '__all__'
    month_data = {
        'main_data': main_data,
        'employees': employees,
        'salary_result': result,
        'config_snapshot': _copy.deepcopy(cfg),
        'headless': headless,
        'built_at': built_at,
        'month': month_filter,
        'advance_data': advance_data,
        'parsed': True,
        'calculated': True,
    }

    per_lock = _get_month_lock(month_key)
    with per_lock:
        with MONTH_CACHE_LOCK:
            MONTH_CACHE[month_key] = month_data
            _month_cache_evict_if_needed()

    APP_STATE.update({
        'parsed': True,
        'calculated': True,
        'employees': employees,
        'main_data': main_data,
        'advance_data': advance_data,
        'salary_result': result,
        'month': month_filter,
        '_month_stamp': (month_filter or '')[:7],
        'source_info': {},
        'headless': headless,
        'address_book': {},
        'nssf_sdl_members': {},
    })

    if result and month_filter and main_data.get('dates'):
        from core.database import save_monthly_result
        try:
            save_monthly_result(app.config.get('DATA_FOLDER'), month_filter, result)
        except Exception:
            pass

    return month_data

def _get_month_data(month: str | None) -> dict | None:
    """
    Cache-aware accessor.
    - hit: return MONTH_CACHE[month] directly
    - miss: full pipeline build via _run_pipeline (deepcopy-filter fast path removed 2026-09-01,
      it rebuilt from stale __all__/APP_STATE base and poisoned the month cache)
    Headless derived inside _run_pipeline per month.
    """
    if not month or month == 'all':
        try:
            month = g.view_month
        except Exception:
            try:
                month = resolve_month(request)
            except Exception:
                month = (APP_STATE.get('month') or '')[:7] or datetime.now(EAT).strftime('%Y-%m')
    month = month[:7]

    with MONTH_CACHE_LOCK:
        cached = MONTH_CACHE.get(month)
        if cached is not None:
            return cached

    # ── 缓存 miss → 一律走完整管线 ──
    # 快路径(deepcopy-filter: __all__ 快照或 APP_STATE['main_data'] 作 base 过滤重算)已移除。
    # 该路径用陈旧 base 回写单月缓存，绕过出勤重建与事件推导覆盖，
    # 导致跨月补交后数据台/薪资缺最新提交日（2026-09-01: 8/31 井下钻工 9/1 补交后
    # 数据台只显示到 8/30 的 bug 根因）。miss 即全量重建，成本可接受（LRU 命中时无重建）。
    md = _run_pipeline(month_filter=month)
    if md is not None:
        return md
    # if pipeline returned None (empty employees), try to return whatever cache has
    with MONTH_CACHE_LOCK:
        return MONTH_CACHE.get(month)


def _refresh_employees_cache(month: str | None = None):
    """Per-month cache refresh (not global): rebuild MONTH_CACHE[month] via _run_pipeline."""
    if not APP_STATE.get('parsed'):
        return None
    _month = month if month is not None else APP_STATE.get('month')
    return _run_pipeline(month_filter=_month if _month != 'all' else None)


def _sync_employee_cache(eid, fields):
    """P23 R4: 轻量同步单个员工字段（不触发全量重建），用于纯展示字段。"""
    for _emp in APP_STATE.get('employees', []):
        if _emp.get('id') == eid:
            _emp.update(fields)
            break

# ═══════════════════════════════════════════════════════════
#  静态页面
# ═══════════════════════════════════════════════════════════

@app.route('/')
def index():
    return render_template('index.html', version=APP_VERSION)

@app.route('/static/<path:path>')
def static_files(path):
    return send_from_directory('static', path)

# ── 移动端专属路由（P16）──
@app.route('/m')
def mobile_index():
    return render_template('mobile.html', version=APP_VERSION)


_MOBILE_UA_TOKENS = ('iphone', 'ipad', 'ipod', 'android', 'mobile', 'windows phone', 'blackberry')


@app.before_request
def mobile_redirect():
    """移动端 UA 访问 / 时重定向到 /m；桌面端访问 /m 不反向跳转（便于调试）。
    使用 url_for 以尊重 KILWA_SCRIPT_NAME 子路径（如 /salary），避免重定向到根 /m 导致 404。"""
    if request.path == '/' and any(t in request.headers.get('User-Agent', '').lower() for t in _MOBILE_UA_TOKENS):
        return redirect(url_for('mobile_index'))


@app.before_request
def inject_view_month():
    try:
        g.view_month = resolve_month(request)
    except Exception:
        try:
            g.view_month = datetime.now(EAT).strftime('%Y-%m')
        except Exception:
            g.view_month = ''

# ═══════════════════════════════════════════════════════════
#  API: 数据源信息
# ═══════════════════════════════════════════════════════════

@app.route('/source-info', methods=['GET'])
@login_required
def get_source_info():
    """纯采集模式：无源文件信息，返回空"""
    return jsonify({})

@app.route('/available-months', methods=['GET'])
@login_required
def get_available_months():
    """返回可选的月份列表（纯采集模式：从 collection_submissions + 数据库补充）"""
    from datetime import datetime
    from core.database import get_conn
    months = set()
    # 从已加载的 main_data（采集回填后的数据）提取
    md = APP_STATE.get('main_data', {})
    for d in md.get('dates', []):
        months.add(d[:7])
    # 从 collection_submissions 提取（采集数据的权威月份来源）
    try:
        conn = get_conn(app.config['DATA_FOLDER'])
        for r in conn.execute(
            "SELECT DISTINCT substr(submission_date,1,7) FROM collection_submissions "
            "WHERE submission_date LIKE '____-__-__'").fetchall():
            if r[0]: months.add(r[0])
        conn.close()
    except Exception:
        pass
    # 从数据库补充历史月份
    from core.database import list_monthly_months
    months.update(list_monthly_months(app.config['DATA_FOLDER']))
    # 从出勤覆盖表和奖金罚款表补充月份
    try:
        conn = get_conn(app.config['DATA_FOLDER'])
        for r in conn.execute("SELECT DISTINCT substr(date,1,7) FROM attendance_overrides WHERE date LIKE '____-__-__'").fetchall():
            if r[0]: months.add(r[0])
        for r in conn.execute("SELECT DISTINCT month FROM bonus_penalties").fetchall():
            if r[0]: months.add(r[0])
        conn.close()
    except: pass
    # 包含当前月及未来2个月（支持提前记录出勤）
    now = datetime.now(EAT)
    for i in range(3):
        y = now.year + (now.month + i - 1) // 12
        m = (now.month + i - 1) % 12 + 1
        months.add(f'{y}-{m:02d}')
    return jsonify(sorted(months, reverse=True))

# ═══════════════════════════════════════════════════════════
#  API: 重新加载
# ═══════════════════════════════════════════════════════════

@app.route('/reload', methods=['POST'])
@admin_required
def reload_source():
    # map: POST /reload -> full clear MONTH_CACHE (all keys) + recompute current month
    _clear_all_month_cache()
    try:
        cur_month = g.view_month
    except Exception:
        cur_month = resolve_month(request) if request else APP_STATE.get('month')
    md = _run_pipeline(month_filter=cur_month)
    if md is None:
        return jsonify({'ok': False, 'error': '员工表为空，请先导入通讯录或员工数据'})
    msg = f'已加载 {len(md.get("employees", []))} 名员工，应发 {md.get("salary_result", {}).get("total_gross", 0):,} TZS'
    _audit('reload_source', '', json.dumps({'employees': len(md.get('employees', []))}))
    emps = md.get('employees', [])
    return jsonify({
        'ok': True, 'message': msg,
        'summary': {
            'total_employees': len(emps),
            'piece_underground': sum(1 for e in emps if e.get('default_type') == 'piece_underground'),
            'piece_driller': sum(1 for e in emps if e.get('default_type') == 'piece_driller'),
            'piece_crush': sum(1 for e in emps if e.get('default_type') == 'piece_crush'),
            'day_rate': sum(1 for e in emps if e.get('default_type') == 'day_rate'),
            'advance_only': sum(1 for e in emps if e.get('default_type') == 'advance_only'),
            'overlap_need_decision': sum(1 for e in emps if e.get('source') in ('both',)),
        },
        'employees': emps,
        'dates': md.get('main_data', {}).get('dates', []),
        'salary': md.get('salary_result'),
    })

# ═══════════════════════════════════════════════════════════
#  API: 月份切换
# ═══════════════════════════════════════════════════════════

@app.route('/set-month', methods=['POST'])
@login_required  # P29 T4 A1: 月份上下文是 UX 状态非数据暴露，降为登录即可（spec §7）
def set_month():
    data = request.get_json() or {}
    month = (data.get('month') or '').strip()
    if month == 'all':
        session.pop('view_month', None)
        md = _run_pipeline(month_filter=None)
        if md is None:
            return jsonify({'ok': False, 'error': '员工表为空，请先导入通讯录或员工数据'})
        msg = f'已加载 {len(md.get("employees", []))} 名员工，应发 {md.get("salary_result", {}).get("total_gross", 0):,} TZS'
        return jsonify({'ok': True, 'message': msg, 'salary': md.get('salary_result'), 'headless': bool(md.get('headless', False))})
    if not MONTH_RE.match(month):
        return jsonify({'ok': False, 'error': 'invalid_month_format', 'hint': 'expected YYYY-MM'}), 400
    session['view_month'] = month
    md = _run_pipeline(month_filter=month)
    if md is None:
        return jsonify({'ok': False, 'error': '员工表为空，请先导入通讯录或员工数据'})
    msg = f'已加载 {len(md.get("employees", []))} 名员工，应发 {md.get("salary_result", {}).get("total_gross", 0):,} TZS'
    if bool(md.get('headless')):
        gen_len = len(md.get('main_data', {}).get('dates', []))
        msg = f'预览模式 — {month} 暂无源数据，已生成 {gen_len} 个日期列，仅支持出勤记录'
    return jsonify({'ok': True, 'message': msg, 'salary': md.get('salary_result'), 'headless': bool(md.get('headless', False))})

# ═══════════════════════════════════════════════════════════
#  API: 员工管理 (旧端点 - deprecated, 已迁移到 /api/employees)
# ═══════════════════════════════════════════════════════════

@app.route('/employees', methods=['GET'])
@login_required
def get_employees():
    """[DEPRECATED] 旧版员工列表 — 已迁移到 /api/employees"""
    from core.exceptions import load_overrides
    from core.database import load_bonus_penalties as _load_bp_emp
    import copy as _copy
    month = resolve_month(request)
    overrides = load_overrides(app.config['DATA_FOLDER'], month=month)
    bonus_penalties = _load_bp_emp(app.config['DATA_FOLDER'], month) if month else {}
    # 深拷贝副本再附加展示字段，绝不原地篡改 APP_STATE 缓存
    # （原实现直接改 emp，污染 override_type/day_rate/monthly_salary 导致后续计薪错误）
    employees_copy = _copy.deepcopy(APP_STATE.get('employees', []))
    for emp in employees_copy:
        eid = emp['id']
        emp['overrides'] = overrides.get(eid, [])
        # 根据 overrides 同步覆盖字段（仅永久覆盖影响 override_type）
        emp['override_type'] = None
        for o in emp['overrides']:
            has_range = bool(o.get('start_date', '') or o.get('end_date', ''))
            st = o.get('salary_type')
            if not has_range and st in ('day_rate', 'monthly', 'piece_underground', 'piece_driller', 'piece_crush'):
                emp['override_type'] = st
            if st == 'day_rate' and o.get('day_rate', 0) > 0:
                emp['day_rate'] = o['day_rate']
            if st == 'monthly' and o.get('monthly_salary', 0) > 0:
                emp['monthly_salary'] = o['monthly_salary']
        # 清零不匹配最终类型的基数（P14.4: 井下工人保留 monthly_salary，供 scoring 模式使用）
        ot = emp.get('override_type')
        if emp.get('default_type') == 'piece_underground':
            pass
        elif ot == 'day_rate':
            emp['monthly_salary'] = 0
        elif ot == 'monthly':
            emp['day_rate'] = 0
        elif ot in ('piece_underground', 'piece_driller', 'piece_crush'):
            emp['day_rate'] = 0
            emp['monthly_salary'] = 0
        # 附加当月奖金/罚款
        bp = bonus_penalties.get(eid, {})
        emp['bonus'] = bp.get('bonus', 0)
        emp['penalty'] = bp.get('penalty', 0)
    return jsonify({'employees': employees_copy, 'headless': APP_STATE.get('headless', False)})

@app.route('/employees/override', methods=['POST'])
@editor_required
def save_override():
    # map: POST /employees/override -> MONTH_CACHE[effective_from or start_date[:7] or g.view_month] per-month recompute
    data = request.get_json()
    eid = data.get('employee_id', '')
    # team_id: 班组绑定（1=LAMBA LAMBA, 2=SAKA SAKA, 3=MIZOZO），与 captain 同为 data 透传字段
    # 钻工用 captain 不加 team_id；井下例外绑定班组。显式取 int 默认 0 以防前端缺省
    data['team_id'] = int(data.get('team_id', 0) or 0)
    try:
        vm = g.view_month
    except Exception:
        vm = resolve_month(request) if request else ''
    affected_month = (data.get('effective_from') or (data.get('start_date') or '')[:7] or vm)
    if data.get('type') == 'exclusion':
        from core.exceptions import save_exclusion
        save_exclusion(app.config['DATA_FOLDER'], data)
        _invalidate_month_cache((data.get('start_date') or data.get('date') or affected_month)[:7])
    else:
        from core.exceptions import save_override as _save
        if not data.get('effective_from') and not data.get('start_date') and not data.get('end_date'):
            data['effective_from'] = vm
            affected_month = vm
        _save(app.config['DATA_FOLDER'], data)
        _refresh_employees_cache(affected_month)
        # evict to satisfy cache-miss verification (next GET rebuilds correctly)
        _invalidate_month_cache(affected_month)
    _audit('override_save', eid, json.dumps({
        'name': next((e['name'] for e in APP_STATE.get('employees',[]) if e['id']==eid), eid),
        'salary_type': data.get('salary_type'),
        'day_rate': data.get('day_rate',0),
        'monthly_salary': data.get('monthly_salary',0),
        'month': affected_month,
    }))
    return jsonify({'ok': True})

@app.route('/employees/remove-override', methods=['POST'])
@editor_required
def remove_override():
    # map: POST /employees/remove-override -> MONTH_CACHE[g.view_month] per-month evict
    data = request.get_json()
    from core.exceptions import remove_override
    try:
        vm = g.view_month
    except Exception:
        vm = resolve_month(request) if request else APP_STATE.get('month')
    remove_override(app.config['DATA_FOLDER'], data.get('employee_id'), data.get('index'))
    _refresh_employees_cache(vm)
    _invalidate_month_cache(vm)
    _audit('remove_override', data.get('employee_id',''), json.dumps({'month': vm}))
    return jsonify({'ok': True})

@app.route('/employees/remove-temp-override', methods=['POST'])
@editor_required
def remove_temp_override():
    # map: POST /employees/remove-temp-override -> MONTH_CACHE[g.view_month] per-month evict
    """删除指定员工的所有临时例外（有日期区间的 override），由薪资页面备注管理触发"""
    data = request.get_json()
    eid = data.get('employee_id', '')
    if not eid:
        return jsonify({'ok': False, 'error': '缺少 employee_id'}), 400
    try:
        vm = g.view_month
    except Exception:
        vm = resolve_month(request) if request else APP_STATE.get('month')
    import sqlite3, os
    db_path = os.path.join(app.config['DATA_FOLDER'], 'kilwa.db')
    conn = sqlite3.connect(db_path)
    conn.execute("DELETE FROM overrides WHERE employee_id=? AND (start_date!='' OR end_date!='') AND (type IS NULL OR type != 'exclusion')", (eid,))
    conn.commit()
    conn.close()
    _refresh_employees_cache(vm)
    _invalidate_month_cache(vm)
    _audit('remove_temp_override', eid, json.dumps({'month': vm}))
    return jsonify({'ok': True})

@app.route('/employees/remove-override-by-id', methods=['POST'])
@editor_required
def remove_override_by_id():
    # map: POST /employees/remove-override-by-id -> MONTH_CACHE[override_month or g.view_month] per-month evict
    """按数据库 ID 删除单条覆盖记录"""
    data = request.get_json()
    oid = data.get('override_id')
    if not oid:
        return jsonify({'ok': False, 'error': '缺少 override_id'}), 400
    import sqlite3, os
    db_path = os.path.join(app.config['DATA_FOLDER'], 'kilwa.db')
    conn = sqlite3.connect(db_path)
    # capture affected month before delete for per-month invalidation
    row = conn.execute("SELECT effective_from, start_date FROM overrides WHERE id=?", (oid,)).fetchone()
    affected = (row[0] if row and row[0] else (row[1][:7] if row and row[1] else '')) if row else ''
    try:
        vm = g.view_month
    except Exception:
        vm = resolve_month(request) if request else APP_STATE.get('month')
    if not affected:
        affected = vm
    conn.execute("DELETE FROM overrides WHERE id=?", (oid,))
    conn.commit()
    conn.close()
    _refresh_employees_cache(affected)
    _invalidate_month_cache(affected)
    _audit('remove_override_by_id', str(oid), json.dumps({'month': affected}))
    return jsonify({'ok': True})

@app.route('/api/employees/<employee_id>/temp-overrides', methods=['GET'])
@login_required
@require_permission('employees', 'view')
def api_employee_temp_overrides(employee_id):
    """P14.6: 返回该员工的所有临时例外（有日期区间的覆盖记录）"""
    from core.database import get_conn
    conn = get_conn(app.config['DATA_FOLDER'])
    rows = conn.execute(
        "SELECT id, salary_type, day_rate, monthly_salary, start_date, end_date, note, team_id "
        "FROM overrides WHERE employee_id=? AND (start_date!='' OR end_date!='') "
        "AND (type IS NULL OR type != 'exclusion') ORDER BY start_date",
        (employee_id,)).fetchall()
    conn.close()
    return jsonify({'overrides': [dict(r) for r in rows]})

@app.route('/employees/bonus-penalty', methods=['POST'])
@editor_required
@require_permission('employees', 'edit')  # P29 T4 A2
def save_bonus_penalty():
    # map: POST /employees/bonus-penalty -> MONTH_CACHE[body month[:7]] per-month evict
    """保存单个员工的奖金/罚款（当月独立）"""
    import json as _json
    data = request.get_json()
    eid = data.get('employee_id', '')
    month = data.get('month', '')
    bonus = data.get('bonus', 0)
    penalty = data.get('penalty', 0)
    if not eid or not month:
        return jsonify({'ok': False, 'error': '缺少 employee_id 或 month'}), 400
    from core.database import save_bonus_penalty as _save_bp
    _save_bp(app.config['DATA_FOLDER'], eid, month, bonus, penalty)
    _invalidate_month_cache(month[:7])
    _audit('bonus_penalty_update', eid, _json.dumps({'month': month, 'bonus': bonus, 'penalty': penalty, 'invalidated': month[:7]}))
    return jsonify({'ok': True})

# ── 离职员工管理 ──

@app.route('/employees/dismissed', methods=['GET'])
@login_required
def get_dismissed_employees():
    """获取已离职员工列表（P29-F: load_dismissed_with_info 已 LEFT JOIN employees 带姓名/部门/薪资）"""
    from core.database import load_dismissed_with_info
    dismissed = load_dismissed_with_info(app.config['DATA_FOLDER'])
    for d in dismissed:
        d.setdefault('name', d['employee_id'])
    return jsonify(dismissed)

@app.route('/employees/dismiss', methods=['POST'])
@editor_required
def dismiss_employee_api():
    """标记员工为离职（从列表中隐藏，可恢复）"""
    import json as _json
    data = request.get_json()
    eid = data.get('employee_id', '')
    note = data.get('note', '')
    if not eid:
        return jsonify({'ok': False, 'error': '缺少 employee_id'}), 400
    from core.database import dismiss_employee as _dismiss
    _dismiss(app.config['DATA_FOLDER'], eid, note)
    # P29-F 补丁: 手动离职同步 status, 与 OA 批准路径及复职还原保持三向一致
    from core.database import get_conn
    conn = get_conn(app.config['DATA_FOLDER'])
    conn.execute("UPDATE employees SET status='dismissed' WHERE id=?", (eid,))
    conn.commit()
    conn.close()
    _audit('dismiss_employee', eid, _json.dumps({'note': note}))
    # 按月失效，后续读取重建（避免 salary_result 与 employees 错位）
    with MONTH_CACHE_LOCK:
        affected = [k for k, v in list(MONTH_CACHE.items()) if any(x.get('id') == eid for x in v.get('employees', []))]
    for k in affected:
        if k != '__all__':
            _invalidate_month_cache(k)
    # 若无缓存命中，回退失效当前视图月
    try:
        _invalidate_month_cache(resolve_month(request)[:7])
    except Exception:
        pass
    return jsonify({'ok': True})

@app.route('/employees/restore', methods=['POST'])
@editor_required
def restore_employee_api():
    """P29-F: 复职=重新入职式恢复——必填复职日期；部门/薪资与原值不同时更新员工行并写入事件时间线"""
    import json as _json
    data = request.get_json() or {}
    eid = (data.get('employee_id') or '').strip()
    rehire_date = (data.get('rehire_date') or '').strip()
    if not eid:
        return jsonify({'ok': False, 'error': '缺少 employee_id'}), 400
    if not rehire_date:
        return jsonify({'ok': False, 'error': '缺少 rehire_date'}), 400
    from core.database import get_conn, restore_employee as _restore
    conn = get_conn(app.config['DATA_FOLDER'])
    row = conn.execute(
        "SELECT department, default_type, monthly_salary, day_rate FROM employees WHERE id=?",
        (eid,)).fetchone()
    if not row:
        conn.close()
        return jsonify({'ok': False, 'error': '员工不存在'}), 404
    old_dept = row['department'] or ''
    old_type = row['default_type'] or ''
    old_monthly = int(row['monthly_salary'] or 0)
    old_daily = int(row['day_rate'] or 0)

    new_dept = (data.get('department') or old_dept).strip()
    new_type = data.get('salary_type') or old_type
    if new_type not in ('day_rate', 'monthly', 'piece_underground', 'piece_driller', 'piece_crush'):
        conn.close()
        return jsonify({'ok': False, 'error': '非法 salary_type'}), 400
    try:
        new_monthly = int(data.get('monthly_salary', old_monthly) or 0)
        new_daily = int(data.get('day_rate', old_daily) or 0)
    except (TypeError, ValueError):
        conn.close()
        return jsonify({'ok': False, 'error': '薪资基数必须为数字'}), 400

    operator = session.get('username', '')
    dept_changed = bool(new_dept) and new_dept != old_dept
    sal_changed = (new_type, new_monthly, new_daily) != (old_type, old_monthly, old_daily)

    def _insert_event(conn, etype, payload):
        conn.execute(
            "INSERT INTO employee_events (employee_id, event_type, effective_date, snapshot, payload, operator_id, status) VALUES (?,?,?,?,?,?,?)",
            (eid, etype, rehire_date, '{}', _json.dumps(payload, ensure_ascii=False), operator, 'approved'))

    # 1) 移出离职名单
    _restore(app.config['DATA_FOLDER'], eid)
    # 2) 部门变化 → 更新 + transfer 事件（时间线显示 from → to）
    if dept_changed:
        conn.execute("UPDATE employees SET department=? WHERE id=?", (new_dept, eid))
        _insert_event(conn, 'transfer', {'from_department': old_dept, 'new_department': new_dept})
    # 3) 薪资变化 → 更新 + salary_change 事件（含旧值快照，时间线显示 前 → 后）
    if sal_changed:
        conn.execute(
            "UPDATE employees SET default_type=?, day_rate=?, monthly_salary=? WHERE id=?",
            (new_type, new_daily, new_monthly, eid))
        payload = {'salary_type': new_type, 'day_rate': new_daily, 'monthly_salary': new_monthly}
        if old_type:
            payload['old_type'] = old_type
            payload['old_salary'] = old_monthly if old_type == 'monthly' else old_daily
        _insert_event(conn, 'salary_change', payload)
    # 4) 复职事件（结构化键，前端 i18n 渲染）
    _insert_event(conn, 'reinstatement', {'department': new_dept, 'salary_type': new_type})
    conn.commit()
    conn.close()

    _audit('restore_employee', eid, _json.dumps({
        'rehire_date': rehire_date, 'department': new_dept,
        'dept_changed': dept_changed, 'salary_changed': sal_changed}, ensure_ascii=False))
    # P23 R4: 恢复员工重新进缓存（补 dismiss 过滤的对称逻辑），并重算薪资
    _refresh_employees_cache()
    return jsonify({'ok': True})


# ═══════════════════════════════════════════════════════════
#  P1: 种子导入 — 将现有员工初始化到新模型
# ═══════════════════════════════════════════════════════════

@app.route('/api/seed/employees', methods=['POST'])
@editor_required
def seed_employees():
    """从 APP_STATE + 旧 employees 表导入在职员工到新扩展列 + hire 事件"""
    import json as _json
    from core.database import get_conn, log_audit
    conn = get_conn(app.config['DATA_FOLDER'])
    count = 0
    # 来源1: APP_STATE 中的员工（来自 Excel 解析，136 人）
    for emp in APP_STATE.get('employees', []):
        eid = emp.get('id', '')
        name = emp.get('name', '')
        dept = emp.get('department', '')
        if not eid:
            continue
        # 跳过已离职
        dismissed = conn.execute(
            "SELECT 1 FROM dismissed_employees WHERE employee_id=?", (eid,)).fetchone()
        if dismissed:
            continue
        # 确保 employees 表中有记录
        conn.execute("""
            INSERT OR REPLACE INTO employees (id, name, department, status, hire_date, phone)
            VALUES (?,?,?,'active',COALESCE((SELECT hire_date FROM employees WHERE id=?),'2024-01-01'),?)
        """, (eid, name, dept, eid, emp.get('phone', '')))
        # 补 hire 事件
        existing = conn.execute(
            "SELECT 1 FROM employee_events WHERE employee_id=? AND event_type='hire'", (eid,)).fetchone()
        if not existing:
            conn.execute("""
                INSERT INTO employee_events (employee_id, event_type, effective_date,
                    payload, operator_id, status)
                VALUES (?, 'hire', COALESCE((SELECT hire_date FROM employees WHERE id=?), '2024-01-01'),
                    ?, 'system', 'approved')
            """, (eid, eid, _json.dumps({'name': name, 'department': dept})))
            count += 1
    conn.commit()
    conn.close()
    log_audit(app.config['DATA_FOLDER'], 'seed_employees', 'system',
              _json.dumps({'total': len(APP_STATE.get('employees', [])), 'imported': count}),
              operator=session.get('username',''))
    return jsonify({'ok': True, 'imported': count,
                    'total': len(APP_STATE.get('employees', [])),
                    'message': f'已导入 {count} 名员工的入职事件'})


# ═══════════════════════════════════════════════════════════
#  P1 API: 员工档案
# ═══════════════════════════════════════════════════════════

@app.route('/api/employees', methods=['GET'])
@login_required
@require_permission('employees', 'view')
def api_employees():
    """员工列表（扩展版，含新字段 + overrides + bonus/penalties）"""
    from core.database import list_employees_extended, load_bonus_penalties as _load_bp
    from core.exceptions import load_overrides as _load_ov
    month = resolve_month(request)
    status = request.args.get('status', 'active')
    dept = request.args.get('department')
    employees = list_employees_extended(app.config['DATA_FOLDER'],
                                        status_filter=status, department=dept)
    # P5-c: 补 overrides + bonus_penalties（与旧 /employees 端点对齐）
    bonus_penalties = _load_bp(app.config['DATA_FOLDER'], month) if month else {}
    overrides_data = _load_ov(app.config['DATA_FOLDER'], month=month) if month else {}
    for emp in employees:
        eid = emp['id']
        emp['overrides'] = overrides_data.get(eid, [])
        bp = bonus_penalties.get(eid, {})
        emp['bonus'] = bp.get('bonus', 0)
        emp['penalty'] = bp.get('penalty', 0)
    return jsonify({'employees': employees})

@app.route('/api/employees/<employee_id>', methods=['GET'])
@login_required
@require_permission('employees', 'view')
def api_employee_profile(employee_id):
    """员工档案详情"""
    from core.database import get_employee_profile, load_overrides
    profile = get_employee_profile(app.config['DATA_FOLDER'], employee_id)
    if not profile:
        return jsonify({'error': '员工不存在'}), 404
    # R1b: 合并覆盖基数（对齐计算侧 _run_pipeline 逻辑：永久覆盖 > 默认）
    # employees.day_rate/monthly_salary 可能为 0（真实基数在 overrides 表），
    # 档案页需展示计算侧实际生效的类型与基数，否则"日薪 · 0 TZS/天"。
    month = resolve_month(request)
    db_ovs = load_overrides(app.config['DATA_FOLDER'], month=month)
    ev_ovs = _derive_overrides_from_events(app.config['DATA_FOLDER'], month) if month else {}
    all_ovs = (ev_ovs.get(employee_id, []) or []) + (db_ovs.get(employee_id, []) or [])
    override_type = profile.get('override_type') or None
    day_rate = profile.get('day_rate') or 0
    monthly_salary = profile.get('monthly_salary') or 0
    for o in all_ovs:
        st = o.get('salary_type', '')
        has_range = bool(o.get('start_date') or o.get('end_date'))
        # R3b: 仅永久覆盖（无日期区间）更新类型与基础基数；
        # 临时例外由 get_day_rate_for_date 按天读取，不污染档案基础展示
        if not has_range and st in ('day_rate', 'monthly', 'piece_underground', 'piece_driller', 'piece_crush'):
            override_type = st
        if not has_range:
            if st == 'day_rate' and o.get('day_rate', 0) > 0:
                day_rate = o['day_rate']
            if st == 'monthly' and o.get('monthly_salary', 0) > 0:
                monthly_salary = o['monthly_salary']
    # 清零对齐计算侧（按最终 override_type）
    if override_type == 'day_rate':
        monthly_salary = 0
    elif override_type == 'monthly':
        day_rate = 0
    elif override_type in ('piece_underground', 'piece_driller', 'piece_crush'):
        day_rate = 0
        monthly_salary = 0
    profile['override_type'] = override_type
    profile['day_rate'] = day_rate
    profile['monthly_salary'] = monthly_salary
    return jsonify({'employee': profile})

@app.route('/api/employees/<employee_id>/events', methods=['GET'])
@login_required
@require_permission('employees', 'view')
def api_employee_events(employee_id):
    from core.database import get_employee_events
    events = get_employee_events(app.config['DATA_FOLDER'], employee_id)
    return jsonify({'events': events})

@app.route('/api/employees/<employee_id>/events', methods=['POST'])
@login_required
@require_permission('employees', 'edit')
def api_create_employee_event(employee_id):
    from core.database import get_conn, log_audit
    import json
    data = request.get_json() or {}
    event_type = data.get('event_type', 'note')
    effective_date = data.get('effective_date', '')
    payload = data.get('payload', {})
    operator_id = session.get('username', '')
    
    if not effective_date:
        return jsonify({'ok': False, 'error': 'Date required'}), 400
    
    conn = get_conn(app.config['DATA_FOLDER'])
    conn.execute(
        "INSERT INTO employee_events (employee_id, event_type, effective_date, snapshot, payload, operator_id, status) VALUES (?,?,?,?,?,?,?)",
        (employee_id, event_type, effective_date, '{}', json.dumps(payload), operator_id, 'approved')
    )
    conn.commit()
    log_audit(app.config['DATA_FOLDER'], 'create_event', employee_id, json.dumps({'type': event_type, 'date': effective_date}),
              operator=session.get('username',''))
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/employees/<employee_id>/annual-leave-override', methods=['POST'])
@editor_required
@require_permission('oa', 'approve')
def api_employee_annual_leave_override(employee_id):
    """P20: 切换员工年假资格豁免（仅 OA 审批人）——开启后跳过 NSSF + NIDA 检查"""
    from core.database import get_conn, log_audit
    data = request.get_json() or {}
    v = data.get('override')
    if v in (True, 'true', '1', 1):
        new_val = 1
    elif v in (False, 'false', '0', 0):
        new_val = 0
    else:
        return jsonify({'ok': False, 'error': 'override 必须是 0 或 1'}), 400
    conn = get_conn(app.config['DATA_FOLDER'])
    row = conn.execute("SELECT annual_leave_override FROM employees WHERE id=?",
                       (employee_id,)).fetchone()
    if not row:
        conn.close()
        return jsonify({'ok': False, 'error': '员工不存在'}), 404
    conn.execute("UPDATE employees SET annual_leave_override=? WHERE id=?",
                 (new_val, employee_id))
    conn.commit()
    conn.close()
    _sync_employee_cache(employee_id, {'annual_leave_override': new_val})  # P23 R4 轻量同步
    log_audit(app.config['DATA_FOLDER'], 'annual_leave_override_toggle',
              employee_id, json.dumps({'override': new_val,
                  'operator': session.get('username', '')}),
              operator=session.get('username',''))
    return jsonify({'ok': True, 'override': new_val})

@app.route('/api/employees/<employee_id>', methods=['POST'])
@editor_required
def api_employee_update(employee_id):
    """编辑员工基本信息（P21 M5/R6: 支持改名，旧名自动入 alias 并重建索引）"""
    from core.database import update_employee_fields, log_audit, get_conn
    data = request.get_json() or {}
    # 部门仅超级管理员可直改；非超管改不同部门值 → 拒绝（须走 OA 调岗审批）
    if 'department' in data and session.get('role') != 'super_admin':
        conn = get_conn(app.config['DATA_FOLDER'])
        row = conn.execute("SELECT department FROM employees WHERE id=?", (employee_id,)).fetchone()
        conn.close()
        if row and (row['department'] or '') != (data.get('department') or ''):
            return jsonify({'ok': False, 'error': '部门仅超级管理员可修改，或通过OA调岗审批'}), 403
    # P21 M5/R6: 改名联动——旧名合并进 alias（逗号分隔去重，跳过空），保证旧名反查/搜索仍命中
    renamed = False
    if 'name' in data:
        conn = get_conn(app.config['DATA_FOLDER'])
        row = conn.execute("SELECT name, alias FROM employees WHERE id=?", (employee_id,)).fetchone()
        conn.close()
        if row:
            old_name = (row['name'] or '').strip()
            new_name = (data.get('name') or '').strip()
            if new_name and old_name and old_name != new_name:
                alias_set = []
                if row['alias']:
                    alias_set = [a.strip() for a in str(row['alias']).split(',') if a.strip()]
                if old_name not in alias_set:
                    alias_set.append(old_name)
                data['alias'] = ', '.join(alias_set)
                renamed = True
    # 月度隔离：直改部门/班组 → 按生效月份记录基线（须在写主档前，取其旧值作基线）
    if 'department' in data or 'team_id' in data:
        conn = get_conn(app.config['DATA_FOLDER'])
        _old_row = conn.execute(
            "SELECT department, team_id, default_type, day_rate, monthly_salary FROM employees WHERE id=?",
            (employee_id,)).fetchone()
        conn.close()
        if _old_row:
            _record_direct_base_change(data, employee_id, dict(_old_row), '直改部门/班组')
    ok = update_employee_fields(app.config['DATA_FOLDER'], employee_id, data)
    if ok:
        if renamed or 'alias' in data:
            _build_db_ab_index(app.config['DATA_FOLDER'])  # 改名/别名即时生效（采集/出勤反查）
        log_audit(app.config['DATA_FOLDER'], 'employee_update', employee_id,
                  json.dumps(data),
                  operator=session.get('username',''))
    _refresh_employees_cache()  # P23 R4: 基本字段/月薪基数变更后全量重建
    return jsonify({'ok': ok})

def _record_direct_base_change(data, employee_id, old_row, note):
    """台账：当直改部门/班组时，按生效月份记录基线（供月度隔离，须在写主档前调用）。"""
    from core.database import record_base_change
    has_structural = ('department' in data and (data.get('department') or '') != (old_row['department'] or '')) \
        or ('team_id' in data and int(data.get('team_id') or 0) != int(old_row['team_id'] or 0))
    if not has_structural:
        return
    from_month = (data.get('effective_month') or '')[:7] or datetime.now(EAT).strftime('%Y-%m')
    new = {}
    if 'department' in data:
        new['department'] = data.get('department') or ''
    if 'team_id' in data:
        new['team_id'] = int(data.get('team_id') or 0)
    record_base_change(app.config['DATA_FOLDER'], employee_id, from_month,
                       new=new, operator_id=session.get('username', 'unknown'), note=note)

@app.route('/api/employees/<employee_id>/salary-type', methods=['POST'])
@editor_required
def api_employee_salary_type(employee_id):
    """P7: 修改员工薪资类别+基数 — 同步 employees 主档 + 写 salary_change 事件"""
    from core.database import update_employee_salary_type, create_event, log_audit, record_base_change
    from datetime import datetime
    data = request.get_json() or {}
    st = data.get('salary_type', '')
    if st not in ('day_rate', 'monthly', 'piece_underground', 'piece_driller', 'piece_crush'):
        return jsonify({'ok': False, 'error': '无效的薪资类别'}), 400
    # 基数必须为数字，否则 400（避免 int() 抛 ValueError → 500）
    def _to_int(v):
        try:
            return int(v)
        except (TypeError, ValueError):
            return None
    day_rate = _to_int(data.get('day_rate', 0))
    monthly_salary = _to_int(data.get('monthly_salary', 0))
    if day_rate is None or monthly_salary is None:
        return jsonify({'ok': False, 'error': '薪资基数必须是数字'}), 400
    # P29-c: 变更前快照取管线合并产物(APP_STATE.employees, 含 override_type/清零后基数),
    # 与档案页显示及计薪完全同口径; 冷启动无该员工时回退空
    _pre_emp = next((e for e in (APP_STATE.get('employees') or [])
                     if e.get('id') == employee_id), None)
    old_type = ((_pre_emp or {}).get('override_type')
                or (_pre_emp or {}).get('default_type') or '')
    if old_type == 'day_rate':
        old_salary = int((_pre_emp or {}).get('day_rate', 0) or 0)
    elif old_type == 'monthly':
        old_salary = int((_pre_emp or {}).get('monthly_salary', 0) or 0)
    else:
        old_salary = 0
    # 月度隔离：直改薪资 → 按生效月份记录基线（须在写主档前，旧值作基线）
    record_base_change(app.config['DATA_FOLDER'], employee_id,
                       (data.get('effective_month') or '')[:7] or datetime.now(EAT).strftime('%Y-%m'),
                       new={'default_type': st, 'day_rate': day_rate, 'monthly_salary': monthly_salary},
                       operator_id=session.get('username', 'unknown'), note='直改薪资')
    ok = update_employee_salary_type(app.config['DATA_FOLDER'], employee_id,
                                     st, day_rate, monthly_salary)
    if not ok:
        return jsonify({'ok': False, 'error': '员工不存在'}), 404
    # P26: 档案页改薪资类别 = 权威设置——清除该员工遗留的永久覆盖，避免
    # override_type 优先掩盖新 default_type（HALIMA 案例：day_rate 永久覆盖残留导致改月薪不生效）
    from core.database import clear_permanent_overrides
    clear_permanent_overrides(app.config['DATA_FOLDER'], employee_id)
    # 写 salary_change 事件（approved，本月 1 号生效）→ 时间线记录 + 下月起覆盖推导
    username = session.get('username', 'unknown')
    create_event(app.config['DATA_FOLDER'], {
        'employee_id': employee_id,
        'event_type': 'salary_change',
        'effective_date': datetime.now(EAT).strftime('%Y-%m-01'),
        'snapshot': '{}',
        'payload': json.dumps({'salary_type': st, 'day_rate': day_rate,
                               'monthly_salary': monthly_salary,
                               'old_type': old_type, 'old_salary': old_salary},
                              ensure_ascii=False),
        'operator_id': username,
        'status': 'approved',
    })
    log_audit(app.config['DATA_FOLDER'], 'employee_salary_type', employee_id,
              json.dumps({'salary_type': st, 'day_rate': day_rate,
                          'monthly_salary': monthly_salary}),
              operator=session.get('username',''))
    _refresh_employees_cache()  # P23 R4: 核心修复点——薪资类型/基数变更后立即重算
    return jsonify({'ok': True})

# ── P7: 员工头像 ─────────────────────────────
ALLOWED_AVATAR_EXTS = ('.png', '.jpg', '.jpeg')

@app.route('/api/employees/avatar', methods=['POST'])
@admin_required
def api_employee_avatar_upload():
    """P7: 上传员工头像 → static/avatars/<employee_id>.<ext>（PNG/JPG ≤2MB）"""
    from core.database import update_employee_fields, log_audit
    eid = request.form.get('employee_id', '').strip()
    file = request.files.get('file')
    if not eid or not file or not file.filename:
        return jsonify({'ok': False, 'error': '缺少员工ID或文件'}), 400
    # 路径穿越防护：eid 仅允许字母数字/下划线/连字符
    if not re.fullmatch(r'[A-Za-z0-9_\-]+', eid):
        return jsonify({'ok': False, 'error': '无效的员工ID'}), 400
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in ALLOWED_AVATAR_EXTS:
        return jsonify({'ok': False, 'error': '仅支持 PNG/JPG 图片'}), 400
    # 大小校验（≤2MB）
    file.seek(0, os.SEEK_END)
    size = file.tell()
    file.seek(0)
    if size > 2 * 1024 * 1024:
        return jsonify({'ok': False, 'error': '图片超过 2MB 限制'}), 400
    # 内容魔数校验（PNG / JPEG）
    head = file.read(8)
    file.seek(0)
    is_png = head[:8] == b'\x89PNG\r\n\x1a\n'
    is_jpg = head[:3] == b'\xff\xd8\xff'
    if not (is_png or is_jpg):
        return jsonify({'ok': False, 'error': '文件内容不是有效的 PNG/JPG 图片'}), 400
    avatar_dir = os.path.join(BASE_DIR, 'static', 'avatars')
    os.makedirs(avatar_dir, exist_ok=True)
    if not os.path.exists(os.path.join(avatar_dir, '.gitkeep')):
        open(os.path.join(avatar_dir, '.gitkeep'), 'w').close()
    # 删除同 id 的旧头像（可能是不同扩展名）
    for old in os.listdir(avatar_dir):
        if old.startswith(eid + '.'):
            os.remove(os.path.join(avatar_dir, old))
    new_name = f'{eid}{ext}'
    file.save(os.path.join(avatar_dir, new_name))
    avatar_path = f'static/avatars/{new_name}'
    update_employee_fields(app.config['DATA_FOLDER'], eid, {'avatar_path': avatar_path})
    _sync_employee_cache(eid, {'avatar_path': avatar_path})  # P23 R4 纯展示字段轻量同步
    log_audit(app.config['DATA_FOLDER'], 'employee_avatar', eid,
              json.dumps({'avatar_path': avatar_path}),
              operator=session.get('username',''))
    return jsonify({'ok': True, 'avatar_path': avatar_path})

@app.route('/api/employees/avatar/delete', methods=['POST'])
@admin_required
def api_employee_avatar_delete():
    """P7: 删除员工头像（删文件 + 清空 avatar_path）"""
    from core.database import update_employee_fields, log_audit
    data = request.get_json() or {}
    eid = data.get('employee_id', '')
    if not re.fullmatch(r'[A-Za-z0-9_\-]+', eid):
        return jsonify({'ok': False, 'error': '无效的员工ID'}), 400
    # 读取当前 avatar_path 删除文件
    from core.database import get_employee_profile
    profile = get_employee_profile(app.config['DATA_FOLDER'], eid)
    if profile and profile.get('avatar_path'):
        rel = profile['avatar_path']
        if rel.startswith('static/avatars/'):
            fpath = os.path.join(BASE_DIR, rel.replace('/', os.sep))
            if os.path.exists(fpath):
                os.remove(fpath)
    update_employee_fields(app.config['DATA_FOLDER'], eid, {'avatar_path': ''})
    _sync_employee_cache(eid, {'avatar_path': ''})  # P23 R4 纯展示字段轻量同步
    log_audit(app.config['DATA_FOLDER'], 'employee_avatar_delete', eid,
              json.dumps({'avatar_path': ''}),
              operator=session.get('username',''))
    return jsonify({'ok': True})


# ═══════════════════════════════════════════════════════════
#  P1 API: OA 审批
# ═══════════════════════════════════════════════════════════

@app.route('/api/oa/events', methods=['POST'])
@login_required
@require_permission('oa', 'apply')  # P29 T4 A3（editor 角色地板移除：applicant/collector 为 level 0，凭 oa:apply 键准入）
def oa_create_event():
    """发起 OA 事件（入职/调岗/离职/薪资变更等）"""
    from core.database import create_event, log_audit
    data = request.get_json()
    if not data or not data.get('event_type'):
        return jsonify({'ok': False, 'error': '缺少必填字段'}), 400
    # P8: 入职申请无 employee_id 时，用姓名经 namematch 生成
    if not data.get('employee_id') and data.get('event_type') == 'hire':
        from core.namematch import make_employee_id
        name = (data.get('payload') or {}).get('name', '') or ''
        data['employee_id'] = make_employee_id(name) or name
        if not data['employee_id']:
            return jsonify({'ok': False, 'error': '无法生成员工ID（请检查姓名）'}), 400
    if not data.get('employee_id'):
        return jsonify({'ok': False, 'error': '缺少必填字段'}), 400
    data['operator_id'] = session.get('username', 'unknown')
    # P23 R2: 加班事件校验——必填 date/start_time/end_time，起止时间后端重算合法
    if data.get('event_type') == 'overtime':
        from core.database import _calc_overtime_hours
        _pl = data.get('payload') or {}
        if not (_pl.get('date') and _pl.get('start_time') and _pl.get('end_time')):
            return jsonify({'ok': False, 'error': '加班申请缺少日期或起止时间'}), 400
        _h = _calc_overtime_hours(_pl.get('start_time'), _pl.get('end_time'))
        if _h <= 0:
            return jsonify({'ok': False, 'error': '加班起止时间无效或超出 12 小时上限'}), 400
        # 以后端计算为准（防前端伪造），同时保证 hours>0
        _pl['hours'] = _h
        data['payload'] = _pl
        # R2: 非 super_admin 提交加班时，日期须在提交日前后 2 天内（含）
        if session.get('role') != 'super_admin':
            try:
                ot_date = datetime.strptime(str(_pl.get('date', '')), '%Y-%m-%d').date()
            except (TypeError, ValueError):
                return jsonify({'ok': False, 'error': '加班日期无效'}), 400
            today = datetime.now(EAT).date()
            if abs((ot_date - today).days) > 2:
                _lang = (data.get('lang') or 'zh')
                return jsonify({'ok': False, 'error': _ot_window_err_msg(_lang)}), 400
    # P13: 按事件类型取指定审批人写入（未设定为 ''）
    from core.database import get_approver_for_event
    data['approver'] = get_approver_for_event(app.config['DATA_FOLDER'], data.get('event_type', ''))
    # payload / snapshot 是 dict，需要序列化为 JSON 字符串
    data['payload'] = json.dumps(data.get('payload', {}), ensure_ascii=False)
    data['snapshot'] = json.dumps(data.get('snapshot', {}), ensure_ascii=False)
    event_id = create_event(app.config['DATA_FOLDER'], data)
    log_audit(app.config['DATA_FOLDER'], 'oa_create_event',
              data['employee_id'], json.dumps(data),
              operator=session.get('username',''))
    return jsonify({'ok': True, 'event_id': event_id})

def _oa_read_gate():
    """P29 T4 A4: OA 读接口双键门控（oa:view || oa:apply），spec §7。

    返回 (operator_filter, denied)：
    - 无两键 → (None, (body, 403))，403 形状与 @require_permission 一致
    - 持 oa:view → (None, None)（全量视图）
    - 仅 oa:apply → (username, None)（只看自己提交的：pending/history/count 过滤 operator_id，
      revoked 对本人可见；详情接口据此拦截他人事件）"""
    from core.database import check_permission
    u = session.get('username', '')
    has_view = check_permission(app.config['DATA_FOLDER'], u, 'oa', 'view')
    has_apply = check_permission(app.config['DATA_FOLDER'], u, 'oa', 'apply')
    if not (has_view or has_apply):
        _audit('perm_denied', '', json.dumps({'user': u, 'module': 'oa', 'action': 'view'}))
        return None, (jsonify({'ok': False, 'error': 'forbidden', 'need_permission': 'oa'}), 403)
    return (None if has_view else u), None

@app.route('/api/oa/pending', methods=['GET'])
@login_required
def oa_pending():
    """待审批事件列表（P13: 按当前用户为审批人过滤，super_admin 全可见）
    P29 T4 A4: (oa:view || oa:apply) 门控；仅 apply 者只看自己提交的"""
    from core.database import get_pending_events
    op_filter, denied = _oa_read_gate()
    if denied:
        return denied
    events = get_pending_events(
        app.config['DATA_FOLDER'],
        approver=session.get('username', ''),
        is_super_admin=(session.get('role') == 'super_admin'),
        operator_filter=op_filter)
    return jsonify({'events': events})

@app.route('/api/oa/pending/count', methods=['GET'])
@login_required
def oa_pending_count():
    """待审批数量（P13: 与列表同一过滤规则；P29 T4 A4 同款双键门控+own 过滤）"""
    from core.database import get_pending_events
    op_filter, denied = _oa_read_gate()
    if denied:
        return denied
    events = get_pending_events(
        app.config['DATA_FOLDER'],
        approver=session.get('username', ''),
        is_super_admin=(session.get('role') == 'super_admin'),
        operator_filter=op_filter)
    return jsonify({'count': len(events)})

@app.route('/api/oa/history', methods=['GET'])
@login_required
def oa_history():
    """P8: 已处理事件列表（approved/rejected）；P22 R2: 支持 ?type= 按事件类型筛选
    P29 T4 A4: (oa:view || oa:apply) 门控；仅 apply 者只看自己提交的（含 revoked）"""
    from core.database import get_processed_events
    op_filter, denied = _oa_read_gate()
    if denied:
        return denied
    ev_type = (request.args.get('type') or '').strip() or None
    events = get_processed_events(app.config['DATA_FOLDER'], event_type=ev_type,
                                  operator_filter=op_filter)
    return jsonify({'events': events})

@app.route('/api/oa/events/<int:event_id>/approve', methods=['POST'])
@editor_required
@require_permission('oa', 'approve')
def oa_approve_event(event_id):
    """批准 OA 事件"""
    from core.database import approve_event, get_event, log_audit, apply_approved_event
    username = session.get('username', '')
    event = get_event(app.config['DATA_FOLDER'], event_id)
    if not event:
        return jsonify({'ok': False, 'error': '事件不存在'}), 404
    # P13: 自审限制——super_admin 例外，可批准自己提交的事件
    if event['operator_id'] == username and session.get('role') != 'super_admin':
        return jsonify({'ok': False, 'error': '不能批准自己提交的事件'}), 400
    # P28: 须在 approve_event 抢占状态之前入账——放后面会把批准卡死在无法回退的半途
    from core.database import accrue_comp_leave_monthly
    try:
        accrue_comp_leave_monthly(app.config['DATA_FOLDER'])
    except Exception as e:
        return jsonify({'ok': False, 'error': f'调休入账失败，请稍后重试: {e}'}), 500
    # 并发安全：先用 approve_event 原子抢占（pending→approved，WHERE status='pending'），
    # 抢到的请求才执行副作用；失败回滚状态。防同一事件并发双审批/双扣。
    ok = approve_event(app.config['DATA_FOLDER'], event_id, username)
    if not ok:
        return jsonify({'ok': False, 'error': '事件状态已变化，请刷新后重试'}), 409
    try:
        apply_approved_event(app.config['DATA_FOLDER'], event)
    except Exception as e:
        from core.database import unapprove_event
        unapprove_event(app.config['DATA_FOLDER'], event_id)
        log_audit(app.config['DATA_FOLDER'], 'oa_apply_failed', event['employee_id'],
                  json.dumps({'event_id': event_id, 'event_type': event['event_type'],
                              'error': str(e)}),
                  operator=session.get('username',''))
        return jsonify({'ok': False, 'error': f'批准失败: {str(e)}'}), 400
    log_audit(app.config['DATA_FOLDER'], 'oa_approve',
              event['employee_id'], json.dumps({'event_id': event_id}),
              operator=session.get('username',''))
    # P27 按月隔离修复：按事件生效月份精准刷新，避免全局 APP_STATE.month 错月污染
    eff_month = (event.get('effective_date') or '')[:7]
    if eff_month and MONTH_RE.match(eff_month):
        _refresh_employees_cache(eff_month)
    else:
        _refresh_employees_cache()  # 回退：无生效日期时全局刷新
    return jsonify({'ok': True})

@app.route('/api/oa/events/<int:event_id>/reject', methods=['POST'])
@editor_required
@require_permission('oa', 'approve')
def oa_reject_event(event_id):
    """驳回 OA 事件"""
    from core.database import reject_event, get_event, log_audit
    data = request.get_json()
    reason = (data or {}).get('reject_reason', '')
    if not reason:
        return jsonify({'ok': False, 'error': '驳回原因不能为空'}), 400
    username = session.get('username', '')
    ok = reject_event(app.config['DATA_FOLDER'], event_id, username, reason)
    if ok:
        event = get_event(app.config['DATA_FOLDER'], event_id)
        if event:
            log_audit(app.config['DATA_FOLDER'], 'oa_reject',
                      event['employee_id'], json.dumps({'event_id': event_id, 'reason': reason}),
                      operator=session.get('username',''))
    return jsonify({'ok': ok})

@app.route('/api/oa/events/<int:event_id>', methods=['GET'])
@login_required
def oa_event_detail(event_id):
    """P21 M4: 单个事件详情（前端编辑/撤销预填用）
    P29 T4 A4: (oa:view || oa:apply) 门控；仅 apply 者不可看他人事件"""
    from core.database import get_event
    op_filter, denied = _oa_read_gate()
    if denied:
        return denied
    ev = get_event(app.config['DATA_FOLDER'], event_id)
    if not ev:
        return jsonify({'error': '事件不存在'}), 404
    if op_filter and ev.get('operator_id') != op_filter:
        _audit('perm_denied', '', json.dumps({'user': session.get('username', ''),
                                              'module': 'oa', 'action': 'view'}))
        return jsonify({'ok': False, 'error': 'forbidden', 'need_permission': 'oa'}), 403
    return jsonify({'event': ev})

@app.route('/api/oa/events/<int:event_id>/revoke', methods=['POST'])
@editor_required
@require_permission('oa', 'approve')  # P29 T4 A5（内联本人待审自撤规则原样保留于 handler 内）
def oa_revoke_event(event_id):
    """P21 M4: 撤销事件（已批需 oa:approve 权限；待审仅申请人本人或 super_admin）"""
    from core.database import get_event, revoke_event, log_audit, check_permission
    username = session.get('username', '')
    event = get_event(app.config['DATA_FOLDER'], event_id)
    if not event:
        return jsonify({'ok': False, 'error': '事件不存在'}), 404
    if event['status'] not in ('approved', 'pending'):
        return jsonify({'ok': False, 'error': '该事件状态不可撤销'}), 400
    # 权限分支
    if event['status'] == 'approved':
        if not check_permission(app.config['DATA_FOLDER'], username, 'oa', 'approve'):
            return jsonify({'ok': False, 'error': '无权撤销已批事件（需 OA 审批权限）'}), 403
    else:
        if event['operator_id'] != username and session.get('role') != 'super_admin':
            return jsonify({'ok': False, 'error': '只能撤销自己提交的待审事件'}), 403
    ok = revoke_event(app.config['DATA_FOLDER'], event_id, username)
    if not ok:
        return jsonify({'ok': False, 'error': '撤销失败（事件可能已处理）'}), 400
    log_audit(app.config['DATA_FOLDER'], 'oa_revoke', event['employee_id'],
              json.dumps({'event_id': event_id, 'event_type': event['event_type']}),
              operator=session.get('username',''))
    # P27 按月隔离修复：按事件生效月份精准回退，避免全局错月
    eff_month = (event.get('effective_date') or '')[:7]
    if eff_month and MONTH_RE.match(eff_month):
        _refresh_employees_cache(eff_month)
    else:
        _refresh_employees_cache()  # 回退：无生效日期时全局刷新
    return jsonify({'ok': True})

@app.route('/api/oa/events/<int:event_id>/edit', methods=['POST'])
@editor_required
@require_permission('oa', 'approve')  # P29 T4 A5（内联本人待审自改规则原样保留于 handler 内）
def oa_edit_event(event_id):
    """R1: 修改 OA 事件（年假/调休/病假/普通请假/加班）
    - 待审: update_pending_event（payload.days + effective_date；加班同步改 payload.date）
    - 已批: 支持跨月跨年修改（edit_approved_event 事务内撤销重建或原地修改）
    """
    from core.database import (get_event, update_pending_event, log_audit, check_permission,
                               edit_approved_event)
    data = request.get_json() or {}
    username = session.get('username', '')
    event = get_event(app.config['DATA_FOLDER'], event_id)
    if not event:
        return jsonify({'ok': False, 'error': '事件不存在'}), 404
    if event['event_type'] not in ('annual_leave', 'comp_leave', 'sick', 'casual', 'overtime'):
        return jsonify({'ok': False, 'error': '该类型事件不支持修改'}), 400
    # 权限（同 revoke）
    if event['status'] == 'approved':
        if not check_permission(app.config['DATA_FOLDER'], username, 'oa', 'approve'):
            return jsonify({'ok': False, 'error': '无权修改已批事件（需 OA 审批权限）'}), 403
    else:
        if event['operator_id'] != username and session.get('role') != 'super_admin':
            return jsonify({'ok': False, 'error': '只能修改自己提交的待审事件'}), 403

    new_date = (data.get('effective_date') or '').strip()
    new_days = data.get('days')
    try:
        new_days = int(new_days) if new_days is not None else None
    except (TypeError, ValueError):
        return jsonify({'ok': False, 'error': '无效的天数'}), 400
    if not new_date:
        return jsonify({'ok': False, 'error': '缺少生效日期'}), 400
    if new_days is not None and new_days < 1:
        return jsonify({'ok': False, 'error': '天数至少为 1'}), 400

    if event['status'] == 'pending':
        try:
            payload = json.loads(event['payload'] or '{}')
        except Exception:
            payload = {}
        if event['event_type'] == 'overtime':
            # 待审加班：同步更新 payload.date
            payload['date'] = new_date
            # R2 防绕过：非 super_admin 编辑待审加班也受 ±2 天窗口约束
            if session.get('role') != 'super_admin':
                try:
                    ot_date = datetime.strptime(new_date, '%Y-%m-%d').date()
                except (TypeError, ValueError):
                    return jsonify({'ok': False, 'error': '加班日期无效'}), 400
                today = datetime.now(EAT).date()
                if abs((ot_date - today).days) > 2:
                    _lang = (data.get('lang') or 'zh')
                    return jsonify({'ok': False, 'error': _ot_window_err_msg(_lang)}), 400
        elif new_days is not None:
            payload['days'] = new_days
        fields = {'effective_date': new_date, 'payload': json.dumps(payload, ensure_ascii=False)}
        ok = update_pending_event(app.config['DATA_FOLDER'], event_id, fields)
        if not ok:
            return jsonify({'ok': False, 'error': '修改失败（事件可能已处理）'}), 400
        log_audit(app.config['DATA_FOLDER'], 'oa_edit', event['employee_id'],
                  json.dumps({'event_id': event_id, 'status': 'pending', 'new_date': new_date, 'days': new_days}),
                  operator=session.get('username',''))
        # P27 按月隔离：待审修改按新旧生效月份失效缓存
        old_month = (event.get('effective_date') or '')[:7]
        new_month = (new_date or '')[:7]
        if old_month and MONTH_RE.match(old_month):
            _invalidate_month_cache(old_month)
        if new_month and MONTH_RE.match(new_month) and new_month != old_month:
            _invalidate_month_cache(new_month)
        elif new_month and MONTH_RE.match(new_month):
            _invalidate_month_cache(new_month)
        return jsonify({'ok': True})

    # 已批：支持跨月跨年修改
    if new_days is None:
        try:
            new_days = int(json.loads(event['payload'] or '{}').get('days', 1) or 1)
        except (TypeError, ValueError):
            new_days = 1
    ok, msg = edit_approved_event(app.config['DATA_FOLDER'], event_id, new_date, new_days, username)
    if not ok:
        return jsonify({'ok': False, 'error': str(msg)}), 400
    # R1: 审计动作用 'oa_edit_date'（区分普通 oa_edit）
    _old_date = event.get('effective_date') or ''
    log_audit(app.config['DATA_FOLDER'], 'oa_edit_date', event['employee_id'],
              json.dumps({'event_id': event_id, 'employee_id': event['employee_id'],
                          'old_date': _old_date, 'new_date': new_date, 'operator': username}),
              operator=session.get('username',''))
    # R1: 跨月跨年缓存刷新——收集旧/新日期范围覆盖的所有月份，逐月刷新
    from datetime import datetime as _dt, timedelta as _td
    try:
        _eff_range = int(json.loads(event.get('payload') or '{}').get('days', 1) or 1)
    except (TypeError, ValueError):
        _eff_range = 1
    if event['event_type'] == 'overtime':
        _eff_range = 1  # 加班仅单日
    _months_to_refresh = set()
    for _eff, _days in [(_old_date, _eff_range), (new_date, new_days)]:
        try:
            _d0 = _dt.strptime(_eff, '%Y-%m-%d')
        except (TypeError, ValueError):
            continue
        for _i in range(_days):
            _ds = (_d0 + _td(days=_i)).strftime('%Y-%m')
            if MONTH_RE.match(_ds):
                _months_to_refresh.add(_ds)
    if _months_to_refresh:
        for _m in sorted(_months_to_refresh):
            _refresh_employees_cache(_m)
    else:
        _refresh_employees_cache()  # 回退：无有效月份时全局刷新
    return jsonify({'ok': True, 'event_id': event_id, 'new_event_id': msg})


# ── P13: 审批人路由（super_admin 后台指定） ─────────────

ALLOWED_APPROVAL_EVENT_TYPES = ('hire', 'transfer', 'dismiss', 'leave', 'overtime', 'sick')  # P28 R5: 增病假

# ── P21 M7/R5: 年假资格错误双语（后端无 i18n 模块，硬编码两套文案） ──
_LEAVE_ERR_MSG = {
    'zh': {
        'no_nssf': '未参加NSSF',
        'no_nida': 'NIDA证件号为空',
        'no_tin': 'TIN号码为空',
        'no_hire_date': '入职日期为空',
        'invalid_hire_date': '入职日期格式无效',
        'no_employee': '员工不存在',
        'under_1year': '入职不满1年({days}天)',
        'prefix': '年假资格不足: ',
    },
    'en': {
        'no_nssf': 'NSSF not enrolled',
        'no_nida': 'NIDA number is empty',
        'no_tin': 'TIN number is empty',
        'no_hire_date': 'Hire date is empty',
        'invalid_hire_date': 'Invalid hire date format',
        'no_employee': 'Employee not found',
        'under_1year': 'Less than 1 year since hire ({days} days)',
        'prefix': 'Annual leave eligibility failed: ',
    },
}

def _leave_err_msg(codes, reasons, lang):
    """按 reason codes 拼双语错误串（codes 来自 check_annual_leave_eligible）"""
    import re
    lang = 'en' if (lang or '') == 'en' else 'zh'
    d = _LEAVE_ERR_MSG[lang]
    parts = []
    for i, code in enumerate(codes):
        if code == 'under_1year':
            days = '?'
            if reasons and i < len(reasons):
                _m = re.search(r'\((\d+)', str(reasons[i]))
                if _m:
                    days = _m.group(1)
            parts.append(d['under_1year'].format(days=days))
        else:
            parts.append(d.get(code, code))
    return d['prefix'] + ', '.join(parts)

# ── R2: 加班日期窗口错误双语 ───────────────────────────────
_OT_WINDOW_ERR_MSG = {
    'zh': '加班日期须在提交日期前后2天内（含），请及时提交加班申请',
    'en': 'Overtime date must be within 2 calendar days of the submission date (inclusive). Please submit your overtime request promptly.',
}

def _ot_window_err_msg(lang):
    """返回加班日期窗口超限的双语错误串"""
    return _OT_WINDOW_ERR_MSG.get('en' if (lang or '') == 'en' else 'zh',
                                   _OT_WINDOW_ERR_MSG['zh'])

def _require_super_admin():
    """内部校验 super_admin，返回错误响应或 None"""
    if session.get('role') != 'super_admin':
        return jsonify({'ok': False, 'error': 'forbidden', 'need_admin': True}), 403
    return None

@app.route('/api/approval-routes', methods=['GET'])
@admin_required
def api_approval_routes_list():
    """P13: 审批人路由全表"""
    from core.database import get_approval_routes
    routes = get_approval_routes(app.config['DATA_FOLDER'])
    return jsonify({'routes': routes})

@app.route('/api/approval-routes', methods=['POST'])
@admin_required
def api_approval_routes_set():
    """P13: 设置事件类型的指定审批人"""
    _block = _require_super_admin()
    if _block:
        return _block
    from core.database import set_approval_route, get_user_role, log_audit
    data = request.get_json() or {}
    event_type = (data.get('event_type') or '').strip()
    approver = (data.get('approver') or '').strip()
    if event_type not in ALLOWED_APPROVAL_EVENT_TYPES:
        return jsonify({'ok': False, 'error': '无效的事件类型'}), 400
    if not approver or not get_user_role(app.config['DATA_FOLDER'], approver):
        return jsonify({'ok': False, 'error': '审批人不存在'}), 400
    set_approval_route(app.config['DATA_FOLDER'], event_type, approver)
    log_audit(app.config['DATA_FOLDER'], 'approval_route_set', '',
              json.dumps({'event_type': event_type, 'approver': approver}),
              operator=session.get('username',''))
    return jsonify({'ok': True})

@app.route('/api/approval-routes/<int:route_id>', methods=['DELETE'])
@admin_required
def api_approval_routes_delete(route_id):
    """P13: 删除审批人路由（清除指定审批人，恢复为所有人可见）"""
    _block = _require_super_admin()
    if _block:
        return _block
    from core.database import delete_approval_route, log_audit
    ok = delete_approval_route(app.config['DATA_FOLDER'], route_id)
    if ok:
        log_audit(app.config['DATA_FOLDER'], 'approval_route_delete', '',
                  json.dumps({'route_id': route_id}),
                  operator=session.get('username',''))
    return jsonify({'ok': ok})


# ═══════════════════════════════════════════════════════════
#  P2 API: 考勤批量提交 + 请假 + 产量录入
# ═══════════════════════════════════════════════════════════

@app.route('/api/attendance/batch', methods=['POST'])
@super_admin_required
@require_permission('attendance', 'edit')
def attendance_batch_submit():
    from core.database import save_attendance_override, log_audit, is_driver, add_driver, get_attendance_status
    data = request.get_json()
    if not data or 'date' not in data or 'marks' not in data:
        return jsonify({'ok': False, 'error': '缺少 date 或 marks'}), 400
    date = data['date']
    count = 0
    for m in data['marks']:
        eid = m.get('employee_id', '')
        status = m.get('status', '')
        if not eid or not status:
            continue
        if status == 'NU':
            return jsonify({'ok': False, 'error': f'NU（年假）状态由审批管理，禁止手动修改（{eid}）'}), 403
        if get_attendance_status(app.config['DATA_FOLDER'], eid, date) == 'NU':
            return jsonify({'ok': False, 'error': f'NU（年假）状态由审批管理，禁止手动修改（{eid}）'}), 403
        try:
            save_attendance_override(app.config['DATA_FOLDER'], eid, date, status, source=1)
        except ValueError as e:  # P28 R3: Y 已取消/非法状态 → 明确报错
            return jsonify({'ok': False, 'error': str(e)}), 400
        count += 1
        if m.get('is_driver') and not is_driver(app.config['DATA_FOLDER'], eid):
            add_driver(app.config['DATA_FOLDER'], eid)
    _invalidate_month_cache((date or '')[:7])
    log_audit(app.config['DATA_FOLDER'], 'attendance_batch', session.get('username',''),
              json.dumps({'date': date, 'count': count}),
              operator=session.get('username',''))
    return jsonify({'ok': True, 'count': count})

@app.route('/api/attendance/roster', methods=['GET'])
@login_required
def attendance_roster():
    dept = request.args.get('department', '')
    from core.database import list_employees_extended
    emps = list_employees_extended(app.config['DATA_FOLDER'], status_filter='active', department=dept)
    return jsonify({'employees': emps})

@app.route('/api/oa/leave', methods=['POST'])
@login_required
@require_permission('oa', 'apply')  # P29 T4 A3（editor 角色地板移除，同上）
def oa_submit_leave():
    from core.database import create_event, log_audit, check_annual_leave_eligible
    data = request.get_json()
    if not data or 'employee_id' not in data or 'event_type' not in data:
        return jsonify({'ok': False, 'error': '缺少必填字段'}), 400
    event_type = data['event_type']
    eid = data['employee_id']
    if event_type == 'annual_leave':
        # P21 R3/M7: 资格检查返回结构化 codes；错误按请求 lang 双语拼装（保留 error 字符串字段）
        chk = check_annual_leave_eligible(app.config['DATA_FOLDER'], eid)
        if not chk['eligible']:
            lang = (data.get('lang') or 'zh')
            return jsonify({'ok': False,
                            'error': _leave_err_msg(chk.get('codes', []), chk.get('reasons', []), lang),
                            'codes': chk.get('codes', [])}), 403
    # P21 R1: comp_leave 由「提交即生效」改为「创建 pending 事件」，审批通过才扣余额+落 T
    data['operator_id'] = session.get('username', 'unknown')
    data['payload'] = json.dumps({
        'days': data.get('days', 1),
        'note': data.get('note', ''),
        'event_type': event_type,
    }, ensure_ascii=False)
    event_id = create_event(app.config['DATA_FOLDER'], {
        'employee_id': eid,
        'event_type': event_type,
        'effective_date': data['effective_date'],
        'payload': data['payload'],
        'snapshot': '{}',
        'operator_id': data['operator_id'],
    })
    log_audit(app.config['DATA_FOLDER'], 'oa_create_event', eid,
              json.dumps({'event_type': event_type, 'event_id': event_id}),
              operator=session.get('username',''))
    return jsonify({'ok': True, 'event_id': event_id})

@app.route('/api/leave/balance/<employee_id>', methods=['GET'])
@login_required
def leave_balance(employee_id):
    import datetime as _dt
    year = request.args.get('year', str(_dt.datetime.now(_dt.timezone(_dt.timedelta(hours=3))).year))
    from core.database import get_leave_balance, accrue_comp_leave_monthly, log_audit
    try:
        accrue_comp_leave_monthly(app.config['DATA_FOLDER'])
    except Exception as e:
        log_audit(app.config['DATA_FOLDER'], 'comp_accrual_lazy_failed', employee_id,
                  json.dumps({'error': str(e)}),
                  operator=session.get('username',''))
    balance = get_leave_balance(app.config['DATA_FOLDER'], employee_id, year)
    return jsonify({'balance': balance})

@app.route('/api/leave/sick', methods=['POST'])
@login_required
@require_permission('oa', 'apply')  # P29 T4 A3（editor 角色地板移除，同上）
def leave_sick():
    """P28 R5: 病假申请改走 OA 审批 — 创建待审事件（审批通过才扣病假余额+逐日落 SK），
    不再免审直批、不再落 P"""
    from core.database import create_event, log_audit, get_approver_for_event, get_employee_profile
    data = request.get_json() or {}
    eid = data.get('employee_id', '')
    date = data.get('effective_date', '')
    if not eid or not date:
        return jsonify({'ok': False, 'error': '缺少员工或日期'}), 400
    if not get_employee_profile(app.config['DATA_FOLDER'], eid):
        return jsonify({'ok': False, 'error': '员工不存在'}), 404
    try:
        days = int(data.get('days', 1))
        if days < 1:
            raise ValueError
        datetime.strptime(date, '%Y-%m-%d')
    except (TypeError, ValueError):
        return jsonify({'ok': False, 'error': '无效的日期或天数'}), 400
    # P13/P28: 按事件类型取指定审批人写入（未设定为 ''，与年假/调休提交同款）
    event_id = create_event(app.config['DATA_FOLDER'], {
        'employee_id': eid,
        'event_type': 'sick',
        'effective_date': date,
        'payload': json.dumps({
            'days': days,
            'note': data.get('note', ''),
            'event_type': 'sick',
        }, ensure_ascii=False),
        'snapshot': '{}',
        'operator_id': session.get('username', 'unknown'),
        'approver': get_approver_for_event(app.config['DATA_FOLDER'], 'sick'),
    })
    log_audit(app.config['DATA_FOLDER'], 'oa_create_event', eid,
              json.dumps({'event_type': 'sick', 'event_id': event_id, 'date': date, 'days': days}),
              operator=session.get('username',''))
    return jsonify({'ok': True, 'message': '已提交审批，待OA审核', 'event_id': event_id})

@app.route('/api/leave/balance/adjust', methods=['POST'])
@admin_required
def leave_balance_adjust():
    """P8: 手动调整员工病假余额（写审计）"""
    from core.database import adjust_leave_balance, log_audit
    import datetime as _dt
    data = request.get_json() or {}
    eid = data.get('employee_id', '')
    year = str(data.get('year', _dt.datetime.now(_dt.timezone(_dt.timedelta(hours=3))).year))
    if not eid:
        return jsonify({'ok': False, 'error': '缺少员工ID'}), 400
    ok = adjust_leave_balance(app.config['DATA_FOLDER'], eid, year,
                              sick_entitled=data.get('sick_entitled'),
                              sick_used=data.get('sick_used'),
                              annual_entitled=data.get('annual_entitled'),
                              annual_used=data.get('annual_used'),
                              comp_entitled=data.get('comp_entitled'),
                              comp_used=data.get('comp_used'))
    if not ok:
        return jsonify({'ok': False, 'error': '无有效调整字段'}), 400
    log_audit(app.config['DATA_FOLDER'], 'leave_balance_adjust', eid,
              json.dumps({'year': year, 'sick_entitled': data.get('sick_entitled'),
                          'sick_used': data.get('sick_used'),
                          'annual_entitled': data.get('annual_entitled'),
                          'annual_used': data.get('annual_used'),
                          'comp_entitled': data.get('comp_entitled'),
                          'comp_used': data.get('comp_used')}),
              operator=session.get('username',''))
    return jsonify({'ok': True})

# ═══════════════════════════════════════════════════════════
#  P9 API: 数据采集
# ═══════════════════════════════════════════════════════════

def _collection_payload_names(payload, form_type):
    """把 payload 中的 eid 集合转成姓名（与 Excel 解析 main_data 同构）"""
    from core.database import get_conn
    name_map = {}
    conn = None
    try:
        conn = get_conn(app.config['DATA_FOLDER'])
        for r in conn.execute("SELECT id, name FROM employees").fetchall():
            name_map[r['id']] = r['name']
    except Exception:
        pass
    finally:
        if conn:
            conn.close()
    def _names(eids):
        return [name_map.get(e, e) for e in (eids or [])]
    return name_map, _names

def _norm_ug_dept(s):
    """UG 部门规范化：去所有空格 + 全角括号归一半角 + 大写（禁止裸字符串比较）"""
    return re.sub(r'\s+', '', (s or '')).replace('（', '(').replace('）', ')').upper()

_UG_NORM_TARGET = _norm_ug_dept('Production TEAM （underground）')

def _is_ug_dept(dept):
    return _norm_ug_dept(dept) == _UG_NORM_TARGET

def _ensure_collection_team_id_column(data_folder):
    """懒迁移：给 collection_submissions 补 team_id 列（UG 出勤按 team 分行需要）"""
    try:
        from core.database import get_conn
        conn = get_conn(data_folder)
        try:
            conn.execute("ALTER TABLE collection_submissions ADD COLUMN team_id INTEGER DEFAULT 0")
            conn.commit()
        except Exception:
            pass
        conn.close()
    except Exception:
        pass


def _build_ug_team_members(data_folder, month=None):
    """C5: 构建 ug_team_members: {team_id: [employee_id,...]} 仅 UG 部门按 team_id 分组。
    month 可选：按月台账解析该月部门/班组归属，避免直接改部门/班组回溯污染历史月份。"""
    team_map = {}
    try:
        from core.database import get_conn, resolve_base_for_month
        conn = get_conn(data_folder)
        for r in conn.execute("SELECT id, department, team_id FROM employees").fetchall():
            base = resolve_base_for_month(data_folder, str(r['id']), month) if month else None
            dept = (base['department'] if base and base.get('department') is not None else r['department'])
            tid = (int(base['team_id'] or 0) if base and base.get('team_id') is not None
                   else int(r['team_id'] or 0))
            if _norm_ug_dept(dept) == _UG_NORM_TARGET:
                if tid:
                    team_map.setdefault(tid, []).append(str(r['id']))
        conn.close()
    except Exception:
        pass
    return team_map

def _merge_collection_to_main_data(main_data, form_type, date, payload):
    """P9: 单条采集提交合并进 main_data（Web 采集覆盖 Excel 同日期）"""
    _, _names = _collection_payload_names(payload, form_type)
    if form_type == 'underground':
        # C1/C2: 检测新格式 teams 存在即按团队数组构建，否则走旧 day/night 路径（保持 byte-identical）
        if 'teams' in payload:
            teams = []
            for t in (payload.get('teams') or []):
                try:
                    tid = int(t.get('team_id', 0) or 0)
                except Exception:
                    tid = 0
                teams.append({
                    'team_id': tid,
                    'prod': {'NICKEL（H）': t.get('nh', 0), 'NICKEL（L）': t.get('nl', 0), 'MAWE': t.get('mw', 0)},
                    'exempt': bool(t.get('exempt', False)),
                    'remark': str(t.get('remark') or ''),
                })
            rec = {'date': date, 'teams': teams}
            shift = main_data.setdefault('shift_production', [])
            for i, x in enumerate(shift):
                if x.get('date') == date:
                    shift[i] = rec
                    return
            shift.append(rec)
            return
        day = payload.get('day') or {}
        night = payload.get('night') or {}
        rec = {
            'date': date,
            'day_prod': {'NICKEL（H）': day.get('nh', 0), 'NICKEL（L）': day.get('nl', 0), 'MAWE': day.get('mw', 0)},
            'night_prod': {'NICKEL（H）': night.get('nh', 0), 'NICKEL（L）': night.get('nl', 0), 'MAWE': night.get('mw', 0)},
            'day_emps': _names(day.get('emps')),
            'night_emps': _names(night.get('emps')),
            'day_team': int(day.get('team_id', 0) or 0),
            'night_team': int(night.get('team_id', 0) or 0),
            'day_exempt': bool(day.get('exempt', False)),
            'night_exempt': bool(night.get('exempt', False)),
        }
        shift = main_data.setdefault('shift_production', [])
        for i, x in enumerate(shift):
            if x.get('date') == date:
                shift[i] = rec
                return
        shift.append(rec)
    elif form_type == 'driller':
        driller = main_data.setdefault('driller_production', [])
        _, _names2 = _collection_payload_names(payload, form_type)
        for team in (payload.get('teams') or []):
            cap = team.get('captain', '')
            if not cap:
                continue
            cap_name = _names2([cap])[0]
            rec = {
                'date': date,
                'captain': cap_name,
                'slot': 0,
                'nh': team.get('nh', 0), 'nl': team.get('nl', 0), 'mw': team.get('mw', 0),
                'futa': 0, 'waya': 0, 'kibiriti': 0,
                'members': _names2(team.get('members')),
                'slots': [],
                'has_members': bool(team.get('members')),
            }
            for i, x in enumerate(driller):
                if x.get('date') == date and x.get('captain') == cap_name:
                    driller[i] = rec
                    break
            else:
                driller.append(rec)
    elif form_type == 'crush':
        crush = main_data.setdefault('crush_production', [])
        rec = {'date': date, 'bags': payload.get('bags', 0), 'personnel': _names(payload.get('emps'))}
        for i, x in enumerate(crush):
            if x.get('date') == date:
                crush[i] = rec
                return
        crush.append(rec)

def rebuild_main_data_from_collections(main_data):
    """P9: /reload 后从 collection_submissions 最新版本回填 main_data（Web 采集覆盖 Excel 同日期）
    B1: 全量重建语义——先清空产量数组，再从 DB 最新提交重建，杜绝双日期残留"""
    from core.database import get_collection_submissions
    main_data['shift_production'] = []
    main_data['driller_production'] = []
    main_data['crush_production'] = []
    subs = get_collection_submissions(app.config['DATA_FOLDER'])
    for s in subs:
        try:
            payload = json.loads(s.get('payload', '{}'))
        except (TypeError, ValueError):
            continue
        if s.get('form_type') in ('underground', 'driller', 'crush'):
            _merge_collection_to_main_data(main_data, s['form_type'], s['submission_date'], payload)

def _reapply_driver_flags_for_date(data_folder, date):
    """C4: 清除某日 driver 标志后从所有来源重设：
    (a) 旧 underground payload day/night.drivers
    (b) attendance submissions drivers array"""
    from core.database import get_collection_submissions, mark_driver_flag, clear_driver_flags_for_date
    clear_driver_flags_for_date(data_folder, date)
    subs = get_collection_submissions(data_folder)
    for s in subs:
        if s.get('submission_date') != date:
            continue
        try:
            payload = json.loads(s.get('payload', '{}'))
        except (TypeError, ValueError):
            continue
        if s.get('form_type') == 'underground':
            drivers = []
            for _shift in ('day', 'night'):
                drivers += (payload.get(_shift) or {}).get('drivers') or []
            for _eid in drivers:
                mark_driver_flag(data_folder, _eid, date)
        elif s.get('form_type') == 'attendance':
            for _eid in (payload.get('drivers') or []):
                mark_driver_flag(data_folder, _eid, date)

def _reapply_driver_flags():
    """B1: 全量重建 main_data 后重设井下 driver 标志（遍历日期委托到单日 helper）"""
    from core.database import get_collection_submissions
    subs = get_collection_submissions(app.config['DATA_FOLDER'])
    ug_dates = {s.get('submission_date') for s in subs
                if s.get('form_type') == 'underground' and s.get('submission_date')}
    att_dates = set()
    for s in subs:
        if s.get('form_type') == 'attendance':
            try:
                pl = json.loads(s.get('payload', '{}'))
            except Exception:
                continue
            if pl.get('drivers'):
                att_dates.add(s.get('submission_date'))
    all_dates = sorted(ug_dates | att_dates)
    for _dt in all_dates:
        _reapply_driver_flags_for_date(app.config['DATA_FOLDER'], _dt)

def _filter_marks_by_department(marks, dept, team_id=None):
    """A5: 出勤收集兜底校验 — 非 UG 按 department 过滤；UG 时按 team_id 过滤（C3）"""
    if not dept:
        return marks or [], 0
    is_ug = _is_ug_dept(dept)
    if is_ug and team_id is not None:
        # UG 按班组过滤：仅保留同 team_id 的员工（直接查 DB 避免 APP_STATE 陈旧）
        from core.database import get_conn
        team_map = {}
        conn = None
        try:
            conn = get_conn(app.config['DATA_FOLDER'])
            team_map = {str(r['id']): int(r['team_id'] or 0) for r in conn.execute(
                "SELECT id, team_id FROM employees").fetchall()}
        except Exception:
            pass
        finally:
            if conn:
                conn.close()
        kept, discarded = [], 0
        try:
            tid_int = int(team_id)
        except Exception:
            tid_int = 0
        for m in marks or []:
            eid = str(m.get('employee_id') or '')
            emp_team = team_map.get(eid) if eid else None
            if emp_team is not None and emp_team != tid_int:
                discarded += 1
                continue
            kept.append(m)
        return kept, discarded
    # 非 UG：原 department 过滤
    dept_map = {}
    try:
        dept_map = {str(e.get('id')): (e.get('department') or '') for e in (APP_STATE.get('employees') or [])}
    except Exception:
        pass
    if not dept_map:
        from core.database import get_conn
        conn = None
        try:
            conn = get_conn(app.config['DATA_FOLDER'])
            dept_map = {str(r['id']): (r['department'] or '') for r in conn.execute(
                "SELECT id, department FROM employees").fetchall()}
        except Exception:
            pass
        finally:
            if conn:
                conn.close()
    kept, discarded = [], 0
    for m in marks or []:
        eid = str(m.get('employee_id') or '')
        emp_dept = dept_map.get(eid) if eid else None
        if emp_dept is not None and emp_dept != dept:
            discarded += 1
            continue
        kept.append(m)
    return kept, discarded

@app.route('/api/collection/submit', methods=['POST'])
@login_required
def collection_submit():
    """P9: 数据采集提交 — 写 collection_submissions + 合并 main_data + 重算"""
    from datetime import datetime
    from core.database import insert_collection_submission, update_collection_submission, \
        get_collection_submissions, save_attendance_override, mark_driver_flag, log_audit
    data = request.get_json() or {}
    form_type = data.get('form_type', '')
    date = data.get('submission_date', '')
    payload = data.get('payload') or {}
    if form_type not in ('underground', 'driller', 'crush', 'attendance'):
        return jsonify({'ok': False, 'error': '无效表单类型'}), 400
    # P29 T4 A6: 按表单类型动态鉴权 collection:<type>（403 形状与 @require_permission 一致）
    from core.database import check_permission as _check_perm
    _u = session.get('username', '')
    if not _check_perm(app.config['DATA_FOLDER'], _u, 'collection', form_type):
        _audit('perm_denied', '', json.dumps({'user': _u, 'module': 'collection', 'action': form_type}))
        return jsonify({'ok': False, 'error': 'forbidden', 'need_permission': 'collection'}), 403
    if not date:
        return jsonify({'ok': False, 'error': '缺少日期'}), 400
    if date > datetime.now(EAT).strftime('%Y-%m-%d'):
        return jsonify({'ok': False, 'error': '不能提交未来日期'}), 400
    username = session.get('username', 'unknown')
    dept = (payload.get('department') or '').strip()

    # 出勤收集：写 attendance_overrides（batch 语义），collection 仅作留痕
    discarded = 0
    att_team_id = None
    if form_type == 'attendance':
        marks = payload.get('marks') or []
        is_ug_att = _is_ug_dept(dept)
        if is_ug_att:
            att_team_id = payload.get('team_id')
            try:
                att_team_id = int(att_team_id) if att_team_id is not None else None
            except Exception:
                att_team_id = None
        # A5 + C3: 非 UG 按 department 过滤；UG 按 team_id 过滤
        if dept:
            if is_ug_att:
                marks, discarded = _filter_marks_by_department(marks, dept, team_id=att_team_id)
            else:
                marks, discarded = _filter_marks_by_department(marks, dept)
            payload['marks'] = marks
        # B→P 遗留映射
        for m in marks:
            if m.get('status') == 'B':
                m['status'] = 'P'
        # P21: NU（年假）由审批管理，采集提交不得覆盖——命中即拒绝整批（防部分写入）
        # P31: E 已从采集摘除（产量豁免仅在井下出渣采集 teams[].exempt），L/SK 转 OA pending（T 直写不转OA）
        from core.database import get_attendance_status
        for m in marks:
            eid = m.get('employee_id', '')
            st = m.get('status', '')
            if eid and get_attendance_status(app.config['DATA_FOLDER'], eid, date) == 'NU':
                return jsonify({'ok': False,
                                'error': f'NU（年假）状态由审批管理，禁止覆盖（员工 {eid} · {date}）'}), 403
            if st == 'E':
                return jsonify({'ok': False, 'error': 'E（豁免）已从出勤采集摘除：设备豁免请在井下出渣产量中勾选，出勤豁免请走 OA 或由管理员在出勤网格标记'}), 400
            if st == 'NU':
                return jsonify({'ok': False, 'error': f'NU（年假）状态由审批管理，禁止采集提交（员工 {eid} · {date}）'}), 403
        # 采集分流：P/A/T 直写，L/SK 转 OA pending
        _OA_MAP = {'L': 'casual', 'SK': 'sick'}
        _oa_created = []
        _oa_skipped = []
        _direct_marks = []
        for m in marks:
            st = m.get('status', '')
            if st in _OA_MAP:
                eid = m.get('employee_id', '')
                if not eid:
                    continue
                etype = _OA_MAP[st]
                # 去重：同 eid+date+type pending 已存在则跳过
                try:
                    from core.database import get_conn as _gc
                    _conn = _gc(app.config['DATA_FOLDER'])
                    _pend = _conn.execute(
                        "SELECT id FROM employee_events WHERE employee_id=? AND effective_date=? AND event_type=? AND status='pending'",
                        (eid, date, etype)).fetchone()
                    _conn.close()
                    if _pend:
                        _oa_skipped.append({'employee_id': eid, 'status': st, 'event_type': etype, 'existing_event_id': _pend['id']})
                        continue
                except Exception:
                    pass
                try:
                    from core.database import create_event, get_approver_for_event, log_audit as _log
                    _approver = get_approver_for_event(app.config['DATA_FOLDER'], etype) or 'maua'
                    # 校验 maua 账号存在性，不存在则回退 ''（所有 oa:approve 可见，KEJU 超管必见）
                    try:
                        from core.database import get_user_role as _gur
                        if _approver == 'maua' and not _gur(app.config['DATA_FOLDER'], 'maua'):
                            _approver = ''
                    except Exception:
                        pass
                    _eid2 = create_event(app.config['DATA_FOLDER'], {
                        'employee_id': eid,
                        'event_type': etype,
                        'effective_date': date,
                        'payload': json.dumps({'days': 1, 'note': f'由出勤采集自动转OA({st})', 'event_type': etype, 'source': 'collection_routing'}, ensure_ascii=False),
                        'snapshot': '{}',
                        'operator_id': username,
                        'approver': _approver,
                    })
                    _log(app.config['DATA_FOLDER'], 'oa_create_event', eid,
                         json.dumps({'event_type': etype, 'event_id': _eid2, 'date': date, 'source': 'collection_routing', 'orig_status': st}, ensure_ascii=False))
                    _oa_created.append({'employee_id': eid, 'status': st, 'event_type': etype, 'event_id': _eid2, 'approver': _approver})
                except Exception as e:
                    return jsonify({'ok': False, 'error': f'创建OA事件失败({eid}/{st}): {e}'}), 500
            else:
                _direct_marks.append(m)
        # 重置 marks 为仅 P/A 直写集合（payload 留痕仍保留原始 marks 供审计，但直写仅 P/A）
        marks = _direct_marks
        # drivers 校验：subset of marks（仅 P/A 集合内）
        drivers = payload.get('drivers') or []
        if drivers:
            marks_ids = {str(m.get('employee_id') or '') for m in marks}
            for d in drivers:
                if str(d) not in marks_ids:
                    return jsonify({'ok': False, 'error': f'驾驶员 {d} 不在当天出勤名单中（仅 P/A 人员可标记驾驶，L/SK/T 已转审批）'}), 400
        for m in marks:
            eid = m.get('employee_id', '')
            status = m.get('status', '')
            if not eid or not status:
                continue
            if status not in ('P', 'A', 'T'):
                return jsonify({'ok': False, 'error': f'采集仅支持 P/A/T 直写，{status} 已转OA或不支持'}), 400
            try:
                save_attendance_override(app.config['DATA_FOLDER'], eid, date, status, source=0)
                if status == 'T':
                    from core.database import deduct_comp_leave
                    _year = int(date[:4])
                    if not deduct_comp_leave(app.config['DATA_FOLDER'], eid, _year, 1):
                        delete_attendance_override(app.config['DATA_FOLDER'], eid, date)
                        return jsonify({'ok': False, 'error': f'员工 {eid} 调休余额不足，无法标记 T'}), 400
            except ValueError as e:
                return jsonify({'ok': False, 'error': str(e)}), 400
        # 标记 driver flag（通过 helper 重设，保证与旧地下 drivers 合并一致）
        if drivers:
            for _eid in drivers:
                try:
                    mark_driver_flag(app.config['DATA_FOLDER'], str(_eid), date)
                except Exception:
                    pass
        # 供外层返回体使用
        _collection_oa_created = _oa_created
        _collection_oa_skipped = _oa_skipped

    # upsert collection_submissions by (form_type, date) [attendance 另加 department+team_id 维度]
    _ensure_collection_team_id_column(app.config['DATA_FOLDER'])
    existing = get_collection_submissions(app.config['DATA_FOLDER'], form_type=form_type)
    if form_type == 'attendance' and dept:
        if _is_ug_dept(dept) and att_team_id is not None:
            ex = next((e for e in existing if e['submission_date'] == date
                       and (e.get('department') or '') == dept and int(e.get('team_id') or 0) == int(att_team_id)), None)
        else:
            ex = next((e for e in existing if e['submission_date'] == date
                       and (e.get('department') or '') == dept), None)
    else:
        ex = next((e for e in existing if e['submission_date'] == date), None)
    if ex:
        update_collection_submission(app.config['DATA_FOLDER'], ex['id'], payload, username)
        sid = ex['id']
    else:
        # insert 需要处理 team_id 列——直接 SQL 以兼容旧 DB（get_conn 方式）
        if form_type == 'attendance' and _is_ug_dept(dept) and att_team_id is not None:
            from core.database import get_conn
            conn = get_conn(app.config['DATA_FOLDER'])
            try:
                cur = conn.execute(
                    "INSERT INTO collection_submissions (form_type, submission_date, payload, operator_id, month, department, team_id, version) VALUES (?,?,?,?,?,?,?,1)",
                    (form_type, date, json.dumps(payload, ensure_ascii=False), username, date[:7], dept, int(att_team_id)))
                conn.commit()
                sid = cur.lastrowid
            except Exception:
                # 回退到旧方法（team_id 写入 payload，查询时靠 payload.team_id）
                sid = insert_collection_submission(app.config['DATA_FOLDER'], form_type, date, payload, username, department=dept)
            finally:
                conn.close()
        else:
            sid = insert_collection_submission(app.config['DATA_FOLDER'], form_type, date, payload, username,
                                               department=dept)

    # map: POST /api/collection/submit -> MONTH_CACHE[submission_date[:7]] per-month evict + __all__ base evict
    _invalidate_month_cache(date[:7])
    _invalidate_all_cache_base()
    # C4: UG 出勤 drivers 变更后重设当日的 driver flags（通过单日 helper 保证多源合并）
    if form_type == 'attendance' and _is_ug_dept(dept):
        try:
            _reapply_driver_flags_for_date(app.config['DATA_FOLDER'], date)
        except Exception:
            pass

    log_audit(app.config['DATA_FOLDER'], 'collection_submit', '',
              json.dumps({'form_type': form_type, 'date': date, 'sid': sid, 'department': dept, 'month': date[:7],
                          'oa_created': _collection_oa_created if form_type=='attendance' and '_collection_oa_created' in locals() else [],
                          'oa_skipped': _collection_oa_skipped if form_type=='attendance' and '_collection_oa_skipped' in locals() else []}, ensure_ascii=False),
              operator=session.get('username',''))
    result = {'ok': True, 'submission_id': sid}
    if form_type == 'attendance' and '_collection_oa_created' in locals():
        if _collection_oa_created:
            result['oa_created'] = _collection_oa_created
        if _collection_oa_skipped:
            result['oa_skipped'] = _collection_oa_skipped
    if discarded:
        result['discarded'] = discarded
        result['warning'] = '已忽略 %d 名部门不符的员工' % discarded
    return jsonify(result)

@app.route('/api/collection/history', methods=['GET'])
@login_required
def collection_history():
    """P9: 采集提交历史（按 form_type/month/date/operator 过滤）
    P29 T4 A7: 门控改 (collection:view || 任一表单键)；仅表单键者强制 operator=本人
    （原 role==='admin' 硬编码收敛为权限判定，view 持有者保留看全部语义）"""
    from core.database import get_collection_submissions, check_permission
    form_type = request.args.get('form_type')
    month = resolve_month(request)
    date = request.args.get('date')
    operator = request.args.get('operator')
    u = session.get('username', '')
    has_view = check_permission(app.config['DATA_FOLDER'], u, 'collection', 'view')
    if not has_view:
        has_form_key = any(check_permission(app.config['DATA_FOLDER'], u, 'collection', a)
                           for a in ('underground', 'driller', 'crush', 'attendance'))
        if not has_form_key:
            _audit('perm_denied', '', json.dumps({'user': u, 'module': 'collection', 'action': 'view'}))
            return jsonify({'ok': False, 'error': 'forbidden', 'need_permission': 'collection'}), 403
        operator = u
    subs = get_collection_submissions(app.config['DATA_FOLDER'], form_type=form_type, month=month,
                                      date=date, operator=operator)
    from core.database import load_dismissed
    try:
        dismissed_ids = sorted(load_dismissed(app.config['DATA_FOLDER']))
    except Exception:
        dismissed_ids = []
    return jsonify({'submissions': subs, 'dismissed_ids': dismissed_ids})

@app.route('/api/collection/driller-teams', methods=['GET'])
@editor_required
def collection_driller_teams():
    """P9: 钻工队长→组员 eid 映射（聚合 main_data + 采集历史），供成员选择弹层默认勾选"""
    from core.database import get_conn, get_collection_submissions
    md = APP_STATE.get('main_data', {})
    name_to_id = {}
    conn = None
    try:
        conn = get_conn(app.config['DATA_FOLDER'])
        for r in conn.execute("SELECT id, name FROM employees").fetchall():
            name_to_id[r['name']] = r['id']
    except Exception:
        pass
    finally:
        if conn:
            conn.close()
    teams = {}
    for d in md.get('driller_production', []):
        cap = d.get('captain', '')
        if not cap:
            continue
        cid = name_to_id.get(cap, cap)
        mids = [name_to_id.get(m, m) for m in (d.get('members') or [])]
        teams.setdefault(cid, [])
        for m in mids:
            if m not in teams[cid]:
                teams[cid].append(m)
    # 采集历史补充（覆盖/新增的队长组员）
    for s in get_collection_submissions(app.config['DATA_FOLDER'], form_type='driller'):
        try:
            payload = json.loads(s.get('payload', '{}'))
        except (TypeError, ValueError):
            continue
        for t in (payload.get('teams') or []):
            cid = t.get('captain', '')
            if not cid:
                continue
            teams.setdefault(cid, [])
            for m in (t.get('members') or []):
                if m not in teams[cid]:
                    teams[cid].append(m)
    return jsonify({'teams': teams})

@app.route('/api/collection/edit/<int:submission_id>', methods=['POST'])
@login_required
def collection_edit(submission_id):
    """P9: 再编辑采集提交（仅本人或 admin+），版本+1 + 旧版写 history + 重新合并 main_data"""
    from datetime import datetime
    from core.database import get_collection_submission, get_collection_submissions, update_collection_submission, \
        delete_collection_submission, log_audit, delete_attendance_override, save_attendance_override
    username = session.get('username', 'unknown')
    sub = get_collection_submission(app.config['DATA_FOLDER'], submission_id)
    if not sub:
        return jsonify({'ok': False, 'error': '提交不存在'}), 404
    # P29 T4 A8: 编辑者（含 admin）须持该表单键 collection:<form_type>，先于 owner-or-admin 判定
    from core.database import check_permission as _check_perm
    if not _check_perm(app.config['DATA_FOLDER'], username, 'collection', sub['form_type']):
        _audit('perm_denied', '', json.dumps({'user': username, 'module': 'collection',
                                              'action': sub['form_type']}))
        return jsonify({'ok': False, 'error': 'forbidden', 'need_permission': 'collection'}), 403
    is_admin = (session.get('role') in ('admin', 'super_admin'))
    if sub['operator_id'] != username and not is_admin:
        return jsonify({'ok': False, 'error': '只能编辑本人提交或管理员可改'}), 403
    data = request.get_json() or {}
    payload = data.get('payload') or {}
    form_type = sub['form_type']
    old_date = sub['submission_date']
    new_date = data.get('submission_date') or old_date
    if new_date > datetime.now(EAT).strftime('%Y-%m-%d'):
        return jsonify({'ok': False, 'error': '不能提交未来日期'}), 400

    # P21: NU（年假）由审批管理，编辑出勤收集不得覆盖——在任何 DB 修改之前拦截
    # C3: UG attendance 按 team 过滤 + B→P + drivers 校验
    edit_att_team_id = None
    if form_type == 'attendance':
        dept_check = (payload.get('department') or '').strip()
        is_ug_edit = _is_ug_dept(dept_check)
        if is_ug_edit:
            try:
                edit_att_team_id = int(payload.get('team_id')) if payload.get('team_id') is not None else None
            except Exception:
                edit_att_team_id = None
        # 兜底过滤（与 submit 一致）
        marks_tmp = payload.get('marks') or []
        if dept_check:
            if is_ug_edit:
                marks_tmp, _ = _filter_marks_by_department(marks_tmp, dept_check, team_id=edit_att_team_id)
            else:
                marks_tmp, _ = _filter_marks_by_department(marks_tmp, dept_check)
            payload['marks'] = marks_tmp
        for m in payload.get('marks') or []:
            if m.get('status') == 'B':
                m['status'] = 'P'
        # drivers 校验
        drivers_edit = payload.get('drivers') or []
        if drivers_edit:
            marks_ids_edit = {str(m.get('employee_id') or '') for m in (payload.get('marks') or [])}
            for d in drivers_edit:
                if str(d) not in marks_ids_edit:
                    return jsonify({'ok': False, 'error': f'驾驶员 {d} 不在当天出勤名单中'}), 400
        from core.database import get_attendance_status
        for m in (payload.get('marks') or []):
            eid = m.get('employee_id', '')
            st = m.get('status', '')
            if eid and get_attendance_status(app.config['DATA_FOLDER'], eid, new_date) == 'NU':
                return jsonify({'ok': False,
                                'error': f'NU（年假）状态由审批管理，禁止覆盖（员工 {eid} · {new_date}）'}), 403
            if st == 'E':
                return jsonify({'ok': False, 'error': 'E（豁免）已从出勤采集摘除'}), 400
            if st == 'NU':
                return jsonify({'ok': False, 'error': f'NU 禁止采集提交（员工 {eid}）'}), 403

    # B1: 日期变更 → 若目标日期已有同 form_type 提交则覆盖合并（更新目标行、删除被编辑旧行），
    #     否则仅更新本行日期。同步更新 submission_date + month 列（payload 内 date 不再作为唯一来源）
    merged_target = None  # B1b: 覆盖合并的目标行（attendance 分支需清理其旧 marks）
    sub_dept = ''
    sub_team_id = None
    if form_type == 'attendance':
        try:
            _sp = json.loads(sub['payload'] or '{}')
            sub_dept = (_sp.get('department') or '').strip()
            if _is_ug_dept(sub_dept):
                sub_team_id = _sp.get('team_id')
                try:
                    sub_team_id = int(sub_team_id) if sub_team_id is not None else None
                except Exception:
                    sub_team_id = None
        except Exception:
            sub_dept = ''
            sub_team_id = None
    if new_date != old_date:
        existing = get_collection_submissions(app.config['DATA_FOLDER'], form_type=form_type)
        if form_type == 'attendance' and sub_dept:
            if _is_ug_dept(sub_dept) and sub_team_id is not None:
                # UG 按 department+team_id 合并
                ex = next((e for e in existing if e['id'] != submission_id
                           and e['submission_date'] == new_date
                           and (e.get('department') or '') == sub_dept
                           and int(e.get('team_id') or 0) == int(sub_team_id)), None)
            else:
                # 出勤留痕按部门×日期：改日期只与同部门的行合并，避免误并到其他部门行
                ex = next((e for e in existing if e['id'] != submission_id
                           and e['submission_date'] == new_date
                           and (e.get('department') or '') == sub_dept), None)
        else:
            ex = next((e for e in existing if e['id'] != submission_id and e['submission_date'] == new_date), None)
        if ex:
            ok = update_collection_submission(app.config['DATA_FOLDER'], ex['id'], payload, username, date=new_date)
            delete_collection_submission(app.config['DATA_FOLDER'], submission_id)
            submission_id = ex['id']
            merged_target = ex
        else:
            ok = update_collection_submission(app.config['DATA_FOLDER'], submission_id, payload, username, date=new_date)
    else:
        ok = update_collection_submission(app.config['DATA_FOLDER'], submission_id, payload, username)
    if not ok:
        return jsonify({'ok': False, 'error': '更新失败'}), 500
    if form_type == 'attendance':
        # 出勤收集编辑: 先删旧 marks 覆盖再写新(避免残留),与 submit 语义一致
        # B1: 日期变更时旧日期的 marks 也要清理，新 marks 落到新日期
        # P21: 删除时跳过 NU 天（年假由审批管理，不随采集编辑被清掉）
        from core.database import get_attendance_status as _att_st
        try:
            old_payload = json.loads(sub['payload'] or '{}')
            for m in (old_payload.get('marks') or []):
                if m.get('employee_id') and _att_st(app.config['DATA_FOLDER'], m['employee_id'], old_date) != 'NU':
                    if m.get('status') == 'T':
                        from core.database import restore_comp_leave
                        restore_comp_leave(app.config['DATA_FOLDER'], m['employee_id'], int(old_date[:4]), 1)
                    delete_attendance_override(app.config['DATA_FOLDER'], m['employee_id'], old_date)
        except Exception:
            pass
        # B1b: 覆盖合并场景——目标行(new_date)原来的 marks 也要删除，避免残留
        if merged_target:
            try:
                target_payload = json.loads(merged_target['payload'] or '{}')
                for m in (target_payload.get('marks') or []):
                    if m.get('employee_id') and _att_st(app.config['DATA_FOLDER'], m['employee_id'], new_date) != 'NU':
                        if m.get('status') == 'T':
                            from core.database import restore_comp_leave
                            restore_comp_leave(app.config['DATA_FOLDER'], m['employee_id'], int(new_date[:4]), 1)
                        delete_attendance_override(app.config['DATA_FOLDER'], m['employee_id'], new_date)
            except Exception:
                pass
        _OA_MAP_EDIT = {'L': 'casual', 'SK': 'sick'}
        _oa_created_edit = []
        _oa_skipped_edit = []
        for m in (payload.get('marks') or []):
            eid = m.get('employee_id', '')
            status = m.get('status', '')
            if not eid or not status:
                continue
            if status in _OA_MAP_EDIT:
                etype = _OA_MAP_EDIT[status]
                try:
                    from core.database import get_conn as _gc2
                    _conn2 = _gc2(app.config['DATA_FOLDER'])
                    _pend2 = _conn2.execute(
                        "SELECT id FROM employee_events WHERE employee_id=? AND effective_date=? AND event_type=? AND status='pending'",
                        (eid, new_date, etype)).fetchone()
                    _conn2.close()
                    if _pend2:
                        _oa_skipped_edit.append({'employee_id': eid, 'status': status, 'existing_event_id': _pend2['id']})
                        continue
                except Exception:
                    pass
                try:
                    from core.database import create_event as _ce2, get_approver_for_event as _g2, log_audit as _lg2
                    _ap2 = _g2(app.config['DATA_FOLDER'], etype) or 'maua'
                    try:
                        from core.database import get_user_role as _gur2
                        if _ap2 == 'maua' and not _gur2(app.config['DATA_FOLDER'], 'maua'):
                            _ap2 = ''
                    except Exception:
                        pass
                    _eid2 = _ce2(app.config['DATA_FOLDER'], {
                        'employee_id': eid,
                        'event_type': etype,
                        'effective_date': new_date,
                        'payload': json.dumps({'days': 1, 'note': f'由出勤采集编辑转OA({status})', 'event_type': etype, 'source': 'collection_edit_routing'}, ensure_ascii=False),
                        'snapshot': '{}',
                        'operator_id': username,
                        'approver': _ap2,
                    })
                    _lg2(app.config['DATA_FOLDER'], 'oa_create_event', eid,
                         json.dumps({'event_type': etype, 'event_id': _eid2, 'date': new_date, 'source': 'collection_edit_routing', 'orig_status': status}, ensure_ascii=False))
                    _oa_created_edit.append({'employee_id': eid, 'status': status, 'event_id': _eid2})
                except Exception as e:
                    return jsonify({'ok': False, 'error': f'创建OA事件失败({eid}/{status}): {e}'}), 500
                continue
            if status not in ('P', 'A', 'T'):
                return jsonify({'ok': False, 'error': f'采集仅支持 P/A/T 直写，{status} 已转OA或不支持'}), 400
            try:
                save_attendance_override(app.config['DATA_FOLDER'], eid, new_date, status, source=0)
                if status == 'T':
                    from core.database import deduct_comp_leave
                    _year = int(new_date[:4])
                    if not deduct_comp_leave(app.config['DATA_FOLDER'], eid, _year, 1):
                        delete_attendance_override(app.config['DATA_FOLDER'], eid, new_date)
                        return jsonify({'ok': False, 'error': f'员工 {eid} 调休余额不足，无法标记 T'}), 400
            except ValueError as e:
                return jsonify({'ok': False, 'error': str(e)}), 400
        # C4: 重设 driver flags（编辑后 attendance drivers 可能变）
        try:
            _reapply_driver_flags_for_date(app.config['DATA_FOLDER'], new_date)
            if new_date != old_date:
                _reapply_driver_flags_for_date(app.config['DATA_FOLDER'], old_date)
        except Exception:
            pass
    # map: POST /api/collection/edit -> MONTH_CACHE[old_date[:7], new_date[:7]] per-month evict + __all__ base evict
    _invalidate_month_cache(old_date[:7])
    if new_date != old_date:
        _invalidate_month_cache(new_date[:7])
    else:
        _invalidate_month_cache(new_date[:7])
    _invalidate_all_cache_base()
    log_audit(app.config['DATA_FOLDER'], 'collection_edit', '',
              json.dumps({'submission_id': submission_id, 'form_type': form_type,
                          'date': new_date, 'old_date': old_date, 'month': new_date[:7],
                          'oa_created': _oa_created_edit if form_type=='attendance' and '_oa_created_edit' in locals() else [],
                          'oa_skipped': _oa_skipped_edit if form_type=='attendance' and '_oa_skipped_edit' in locals() else []}, ensure_ascii=False),
              operator=session.get('username',''))
    _res = {'ok': True, 'submission_id': submission_id}
    if form_type == 'attendance' and '_oa_created_edit' in locals():
        if _oa_created_edit:
            _res['oa_created'] = _oa_created_edit
        if _oa_skipped_edit:
            _res['oa_skipped'] = _oa_skipped_edit
    return jsonify(_res)

@app.route('/api/collection/roster', methods=['GET'])
@login_required
def api_collection_roster():
    """P29-b: 采集/评分表单专用轻量花名册——collector 等 0 级角色无 employees:view,
    但提交表单需要员工选择器数据;按持有任一采集键或 scoring:edit 放行。
    仅返回 id/name/department 等必要字段,不暴露 overrides/bonus 等薪酬敏感数据。"""
    from core.database import check_permission, list_employees_extended
    u = session.get('username', '')
    ok = any(check_permission(app.config['DATA_FOLDER'], u, 'collection', a)
             for a in ('view', 'underground', 'driller', 'crush', 'attendance'))
    if not ok:
        ok = check_permission(app.config['DATA_FOLDER'], u, 'scoring', 'edit')
    if not ok:
        _audit('perm_denied', '', json.dumps({'user': u, 'module': 'collection', 'action': 'roster'}))
        return jsonify({'ok': False, 'error': 'forbidden', 'need_permission': 'collection'}), 403
    emps = list_employees_extended(app.config['DATA_FOLDER'], status_filter='active') or []
    keep = ('id', 'name', 'department', 'default_type', 'team_id', 'custom_number', 'alias')
    slim = [{k: e.get(k) for k in keep} for e in emps]
    # 出勤采集「已批假期」排除：可选 ?date=YYYY-MM-DD，命中者 approved_leave=true
    # （仅布尔标记，不回传具体 status / 薪酬字段；不传 date 行为完全不变）
    date = (request.args.get('date') or '').strip()
    if date:
        from core.database import get_approved_leave_statuses
        approved = set(str(eid) for eid in get_approved_leave_statuses(app.config['DATA_FOLDER'], date).keys())
        for e in slim:
            e['approved_leave'] = str(e.get('id')) in approved
    return jsonify({'ok': True, 'employees': slim})

@app.route('/api/collection/cleanup-routed-leave', methods=['POST'])
def api_collection_cleanup_routed_leave():
    """清理「出勤采集自动路由产生、但与被批假期冲突」的 pending 请假 OA。

    出勤采集提交时把 L/SK 自动转为 casual/sick 的 pending 事件（payload 标记
    source='collection_routing'）。若该员工生效日其实已有审批通过假期
    （attendance_overrides ∈ NU/T/SK/L），则该事件为重复应清除。
    仅清理 collection_routing 来源的 pending，绝不触碰用户正常提交的请假。

    body: {dry_run: true|false}（默认 true 只预览不写库）
    """
    _block = _require_super_admin()
    if _block:
        _audit('perm_denied', '', json.dumps({'user': session.get('username', ''), 'module': 'collection', 'action': 'cleanup'}))
        return _block
    data = request.get_json() or {}
    dry_run = bool(data.get('dry_run', True))
    from core.database import get_conn, find_conflicting_routed_leave_events, log_audit as _log_audit
    username = session.get('username', '')
    folder = app.config['DATA_FOLDER']
    conflicts = find_conflicting_routed_leave_events(folder)
    if not dry_run and conflicts:
        conn = get_conn(folder)
        for c in conflicts:
            conn.execute("DELETE FROM employee_events WHERE id=?", (c['event_id'],))
        conn.commit()
        for c in conflicts:
            _log_audit(folder, 'oa_purge_dup_collection', c['employee_id'],
                       json.dumps({'event_id': c['event_id'], 'event_type': c['event_type'],
                                   'date': c['date'], 'operator': username}, ensure_ascii=False))
        conn.close()
    return jsonify({'ok': True, 'dry_run': dry_run, 'checked': len(conflicts),
                    'conflicts': len(conflicts), 'purged': 0 if dry_run else len(conflicts),
                    'items': conflicts})

@app.route('/api/collection/exempt/<int:submission_id>', methods=['POST'])
def api_collection_exempt(submission_id):
    _block = _require_super_admin()
    if _block:
        _audit('perm_denied', '', json.dumps({'user': session.get('username',''), 'module': 'collection', 'action': 'exempt'}))
        return _block
    from core.database import get_collection_submission, update_collection_submission
    sub = get_collection_submission(app.config['DATA_FOLDER'], submission_id)
    if not sub:
        return jsonify({'ok': False, 'error': '提交不存在'}), 404
    if sub.get('form_type') != 'underground':
        return jsonify({'ok': False, 'error': '仅井下采集支持豁免切换'}), 404
    data = request.get_json() or {}
    # 判定新/旧格式：新含 team_id，旧含 shift
    try:
        payload = json.loads(sub.get('payload') or '{}')
    except Exception:
        payload = {}
    is_new = 'teams' in payload
    month = (sub.get('submission_date') or '')[:7]
    if is_new:
        if 'team_id' not in data or 'exempt' not in data:
            return jsonify({'ok': False, 'error': '缺少 team_id 或 exempt'}), 400
        try:
            tid = int(data.get('team_id'))
        except Exception:
            return jsonify({'ok': False, 'error': 'team_id 非法'}), 400
        exempt_val = data.get('exempt')
        if not isinstance(exempt_val, bool):
            return jsonify({'ok': False, 'error': 'exempt 必须为布尔值'}), 400
        # 找到对应 team
        target = None
        old_val = None
        for t in (payload.get('teams') or []):
            try:
                if int(t.get('team_id', 0)) == tid:
                    target = t
                    old_val = bool(t.get('exempt', False))
                    break
            except Exception:
                continue
        if target is None:
            return jsonify({'ok': False, 'error': '团队不存在'}), 404
        target['exempt'] = exempt_val
        payload['teams'] = payload.get('teams')
    else:
        if 'shift' not in data or 'exempt' not in data:
            return jsonify({'ok': False, 'error': '缺少 shift 或 exempt'}), 400
        shift = str(data.get('shift') or '').strip()
        if shift not in ('day', 'night'):
            return jsonify({'ok': False, 'error': 'shift 必须为 day 或 night'}), 404
        exempt_val = data.get('exempt')
        if not isinstance(exempt_val, bool):
            return jsonify({'ok': False, 'error': 'exempt 必须为布尔值'}), 400
        old_val = bool((payload.get(shift) or {}).get('exempt', False))
        if shift not in payload or not isinstance(payload.get(shift), dict):
            payload[shift] = {}
        payload[shift]['exempt'] = exempt_val
    ok = update_collection_submission(app.config['DATA_FOLDER'], submission_id, payload, session.get('username','unknown'))
    if not ok:
        return jsonify({'ok': False, 'error': '更新失败'}), 500
    # 读取新 version
    sub2 = get_collection_submission(app.config['DATA_FOLDER'], submission_id)
    new_ver = sub2.get('version', 0) if sub2 else 0
    _invalidate_month_cache(month)
    _invalidate_all_cache_base()
    # audit
    try:
        from core.database import log_audit
        audit_detail = {'submission_id': submission_id, 'date': sub.get('submission_date'), 'old': old_val, 'new': exempt_val}
        if is_new:
            audit_detail['team_id'] = tid
        else:
            audit_detail['shift'] = shift
        log_audit(app.config['DATA_FOLDER'], 'collection_exempt_edit', '', json.dumps(audit_detail, ensure_ascii=False),
                  operator=session.get('username',''))
    except Exception:
        pass
    return jsonify({'ok': True, 'version': new_ver})

@app.route('/api/production/shift', methods=['POST'])
@editor_required
def production_shift_entry():
    data = request.get_json()
    from core.database import get_conn, log_audit
    conn = get_conn(app.config['DATA_FOLDER'])
    conn.execute(
        "INSERT OR REPLACE INTO shift_additions (employee_id, date, shift) VALUES (?,?,?)",
        (data['employee_id'], data['date'], data.get('shift', 'D')))
    conn.commit()
    conn.close()
    log_audit(app.config['DATA_FOLDER'], 'production_shift', data['employee_id'],
              json.dumps(data),
              operator=session.get('username',''))
    return jsonify({'ok': True})


# ═══════════════════════════════════════════════════════════
#  P10 API: 评分班组
# ═══════════════════════════════════════════════════════════

@app.route('/api/employee_groups', methods=['GET'])
@login_required
def api_employee_groups_list():
    """P10: 班组列表"""
    from core.database import list_employee_groups
    groups = list_employee_groups(app.config['DATA_FOLDER'])
    return jsonify({'groups': groups})

@app.route('/api/employee_groups', methods=['POST'])
@admin_required
def api_employee_groups_create():
    """P10: 创建班组（admin+）"""
    from core.database import create_employee_group, log_audit
    data = request.get_json() or {}
    name = (data.get('name') or '').strip()
    if not name:
        return jsonify({'ok': False, 'error': '班组名不能为空'}), 400
    gid = create_employee_group(app.config['DATA_FOLDER'], name, data.get('description', ''))
    if not gid:
        return jsonify({'ok': False, 'error': '班组名已存在'}), 400
    log_audit(app.config['DATA_FOLDER'], 'employee_group_create', '',
              json.dumps({'name': name}),
              operator=session.get('username',''))
    return jsonify({'ok': True, 'group_id': gid})

@app.route('/api/employee_groups/<int:group_id>', methods=['PUT'])
@admin_required
def api_employee_groups_update(group_id):
    """P10: 改名班组（admin+）"""
    from core.database import update_employee_group, log_audit
    data = request.get_json() or {}
    name = (data.get('name') or '').strip()
    if not name:
        return jsonify({'ok': False, 'error': '班组名不能为空'}), 400
    ok = update_employee_group(app.config['DATA_FOLDER'], group_id, name,
                               data.get('description'))
    if not ok:
        return jsonify({'ok': False, 'error': '班组不存在或名称冲突'}), 400
    log_audit(app.config['DATA_FOLDER'], 'employee_group_update', '',
              json.dumps({'group_id': group_id, 'name': name}),
              operator=session.get('username',''))
    return jsonify({'ok': True})

@app.route('/api/employee_groups/<int:group_id>', methods=['DELETE'])
@admin_required
def api_employee_groups_delete(group_id):
    """P10: 删除班组（admin+，解除员工关联）"""
    from core.database import delete_employee_group, log_audit
    delete_employee_group(app.config['DATA_FOLDER'], group_id)
    log_audit(app.config['DATA_FOLDER'], 'employee_group_delete', '',
              json.dumps({'group_id': group_id}),
              operator=session.get('username',''))
    return jsonify({'ok': True})


# ═══════════════════════════════════════════════════════════
#  P12 API: 钻工队长名单 driller_captains
# ═══════════════════════════════════════════════════════════

@app.route('/api/driller-captains', methods=['GET'])
@login_required
def api_driller_captains_list():
    """P12: 钻工队长名单（钻工采集队长下拉数据源）"""
    from core.database import get_driller_captains
    captains = get_driller_captains(app.config['DATA_FOLDER'])
    return jsonify({'captains': captains})

@app.route('/api/driller-captains', methods=['POST'])
@admin_required
def api_driller_captains_create():
    """P12: 新增钻工队长（admin+）；name 可空则从 employees 查"""
    from core.database import get_conn, add_driller_captain, log_audit
    data = request.get_json() or {}
    eid = (data.get('employee_id') or '').strip()
    if not eid:
        return jsonify({'ok': False, 'error': '缺少员工ID'}), 400
    name = (data.get('name') or '').strip()
    if not name:
        conn = None
        try:
            conn = get_conn(app.config['DATA_FOLDER'])
            r = conn.execute("SELECT name FROM employees WHERE id=?", (eid,)).fetchone()
            name = r['name'] if r else ''
        except Exception:
            name = ''
        finally:
            if conn:
                conn.close()
    if not name:
        return jsonify({'ok': False, 'error': '缺少姓名（员工不存在）'}), 400
    cid = add_driller_captain(app.config['DATA_FOLDER'], eid, name)
    if not cid:
        return jsonify({'ok': False, 'error': '该员工已在钻工队长名单'}), 400
    log_audit(app.config['DATA_FOLDER'], 'driller_captain_create', eid,
              json.dumps({'name': name}),
              operator=session.get('username',''))
    return jsonify({'ok': True, 'captain': {'id': cid, 'employee_id': eid, 'name': name}})

@app.route('/api/driller-captains/<int:captain_id>', methods=['PUT'])
@admin_required
def api_driller_captains_update(captain_id):
    """P12: 更新钻工队长（name/sort_order，admin+）"""
    from core.database import update_driller_captain, log_audit
    data = request.get_json() or {}
    ok = update_driller_captain(app.config['DATA_FOLDER'], captain_id,
                                name=data.get('name'), sort_order=data.get('sort_order'))
    if not ok:
        return jsonify({'ok': False, 'error': '队长不存在或无有效字段'}), 400
    log_audit(app.config['DATA_FOLDER'], 'driller_captain_update', '',
              json.dumps({'id': captain_id, **data}),
              operator=session.get('username',''))
    return jsonify({'ok': True})

@app.route('/api/driller-captains/<int:captain_id>', methods=['DELETE'])
@admin_required
def api_driller_captains_delete(captain_id):
    """P12: 删除钻工队长（admin+）"""
    from core.database import delete_driller_captain, log_audit
    delete_driller_captain(app.config['DATA_FOLDER'], captain_id)
    log_audit(app.config['DATA_FOLDER'], 'driller_captain_delete', '',
              json.dumps({'id': captain_id}),
              operator=session.get('username',''))
    return jsonify({'ok': True})

@app.route('/api/driver-roster', methods=['GET'])
@editor_required
def api_driver_roster():
    """P12: 司机名单 employee_id 列表（井下采集驾驶勾选数据源）"""
    from core.database import list_drivers
    drivers = [d['employee_id'] for d in list_drivers(app.config['DATA_FOLDER'])]
    return jsonify({'drivers': drivers})


# ═══════════════════════════════════════════════════════════
#  P3 API: 评分系统
# ═══════════════════════════════════════════════════════════

@app.route('/api/scoring/card', methods=['POST'])
@editor_required
def scoring_submit_card():
    from core.database import (submit_scoring_card, save_scoring_card_entries,
                               delete_scoring_card_entries, log_audit)
    data = request.get_json()
    if not data:
        return jsonify({'ok': False, 'error': '缺少数据'}), 400
    # P14.8: 新格式 rows（一张卡 9 行明细）→ 先删再插，覆盖重录
    if 'rows' in data:
        week = data.get('week')
        team_id = data.get('team_id')
        card_no = data.get('card_no')
        source = data.get('source', '工友')
        month = (data.get('month') or '').strip() or resolve_month(request)
        rows = data.get('rows') or []
        if week is None or team_id is None or not card_no:
            return jsonify({'ok': False, 'error': '缺少 week/team_id/card_no'}), 400
        if not rows:
            return jsonify({'ok': False, 'error': '缺少评分行'}), 400
        # P15: 每班驾驶 ≤2 人（匿名原则：一张卡内最多 2 名驾驶员有驾驶维评分；driving<=0 不计）
        driver_count = sum(1 for r in rows if int(r.get('driving') or 0) > 0)
        if driver_count > 2:
            return jsonify({'ok': False, 'error': '每班驾驶最多勾选 2 人'}), 400
        delete_scoring_card_entries(app.config['DATA_FOLDER'], week, team_id, card_no, source, month)
        save_scoring_card_entries(app.config['DATA_FOLDER'], week, team_id, card_no, source, rows, month)
        log_audit(app.config['DATA_FOLDER'], 'scoring_card_entries', session.get('username',''),
                  json.dumps({'week': week, 'team_id': team_id, 'card_no': card_no,
                              'source': source, 'month': month, 'count': len(rows)}),
                  operator=session.get('username',''))
        return jsonify({'ok': True, 'count': len(rows)})
    # 旧格式 entries（P3 逐行）兼容保留
    if 'entries' not in data:
        return jsonify({'ok': False, 'error': '缺少评分数据'}), 400
    week = data['week']
    team = data['team']
    card_no = data['card_no']
    source = data.get('source', '工友')
    entries = data['entries']
    card_id = submit_scoring_card(app.config['DATA_FOLDER'], week, team, card_no, source, entries)
    log_audit(app.config['DATA_FOLDER'], 'scoring_card', session.get('username',''),
              json.dumps({'card_id': card_id, 'week': week, 'team': team, 'source': source}),
              operator=session.get('username',''))
    return jsonify({'ok': True, 'card_id': card_id})

# ── P14.8: 评分原始记录查询/删除（周次/班组/卡号/被评人） ──

@app.route('/api/scoring/entries', methods=['GET'])
@editor_required
@require_permission('scoring', 'view')
def scoring_entries_list():
    """P14.8: 评分原始记录列表（team/week/card/source 组合过滤）"""
    from core.database import get_scoring_card_entries
    team = request.args.get('team', type=int)
    week = request.args.get('week', type=int)
    card = request.args.get('card', '') or None
    source = request.args.get('source', '') or None
    rows = get_scoring_card_entries(app.config['DATA_FOLDER'],
                                    team_id=team, week=week, card_no=card, source=source)
    return jsonify({'ok': True, 'entries': rows, 'count': len(rows)})

@app.route('/api/scoring/card', methods=['GET'])
@editor_required
@require_permission('scoring', 'view')
def scoring_card_get():
    """P14.8: 查询单卡全部行（9 行明细）"""
    from core.database import get_scoring_card_entries
    team = request.args.get('team', type=int)
    week = request.args.get('week', type=int)
    card = request.args.get('card', '')
    source = request.args.get('source', '工友')
    if team is None or week is None or not card:
        return jsonify({'ok': False, 'error': '缺少 team/week/card'}), 400
    rows = get_scoring_card_entries(app.config['DATA_FOLDER'],
                                    team_id=team, week=week, card_no=card, source=source)
    return jsonify({'ok': True, 'rows': rows, 'count': len(rows)})

@app.route('/api/scoring/card', methods=['DELETE'])
@editor_required
def scoring_card_delete():
    """P14.8: 删除单卡（重录/废弃）"""
    from core.database import delete_scoring_card_entries, log_audit
    data = request.get_json(silent=True) or {}
    week = data.get('week') if data.get('week') is not None else request.args.get('week', type=int)
    team_id = data.get('team_id') if data.get('team_id') is not None else request.args.get('team', type=int)
    card_no = data.get('card_no') or request.args.get('card', '')
    source = data.get('source', request.args.get('source', '工友'))
    if week is None or team_id is None or not card_no:
        return jsonify({'ok': False, 'error': '缺少 week/team_id/card_no'}), 400
    delete_scoring_card_entries(app.config['DATA_FOLDER'], week, team_id, card_no, source,
                                month=data.get('month') or request.args.get('month', ''))
    log_audit(app.config['DATA_FOLDER'], 'scoring_card_delete', '',
              json.dumps({'week': week, 'team_id': team_id, 'card_no': card_no,
                          'source': source, 'month': data.get('month') or request.args.get('month', '')}),
              operator=session.get('username',''))
    return jsonify({'ok': True})

# ── P10: 评分录入（班组+月份，一张卡一人） ──

@app.route('/api/scoring/team/<int:team_id>/month/<month>', methods=['GET'])
@editor_required
@require_permission('scoring', 'view')
def scoring_team_month(team_id, month):
    """P10: 班组全员（custom_number 升序，无工号排后）+ 该月已提交评分（按 source 分组，预填回显）"""
    _resolved = resolve_month(request)
    _md_check = _get_month_data(month or _resolved)
    from core.database import get_conn, get_employee_group
    group = get_employee_group(app.config['DATA_FOLDER'], team_id)
    if not group:
        return jsonify({'ok': False, 'error': '班组不存在'}), 404
    conn = get_conn(app.config['DATA_FOLDER'])
    emps = conn.execute("""
        SELECT id, name, custom_number, team_id FROM employees
        WHERE team_id=? AND status='active'
        ORDER BY CASE WHEN custom_number='' OR custom_number IS NULL THEN 1 ELSE 0 END, custom_number
    """, (team_id,)).fetchall()
    rows = conn.execute("""
        SELECT se.*, sc.source FROM scoring_entries se
        JOIN scoring_cards sc ON se.card_id = sc.id
        WHERE sc.team=? AND sc.month=?
    """, (team_id, month)).fetchall()
    conn.close()
    by_source = {}
    for r in rows:
        by_source.setdefault(r['source'], {})[r['target_employee_id']] = {
            'initiative': r['initiative'], 'diligence': r['diligence'],
            'discipline': r['discipline'], 'cooperation': r['cooperation'],
            'safety': r['safety'], 'driving': r['driving'],
        }
    employees = [{'id': e['id'], 'name': e['name'], 'custom_number': e['custom_number'] or ''}
                 for e in emps]
    return jsonify({'group': group, 'employees': employees, 'submitted': by_source})

@app.route('/api/scoring/card/batch', methods=['POST'])
@editor_required
def scoring_card_batch():
    """P10: 批量提交评分卡（team+month+source 先删旧再插新 → 覆盖更新）
    body: {team_id, month, source, cards:[{employee_id, wid, 6 维, driving, note}]}"""
    from core.database import get_conn, log_audit
    data = request.get_json() or {}
    team = data.get('team_id')
    month = data.get('month', '')
    source = data.get('source', '工友')
    cards = data.get('cards') or []
    if not team or not month:
        return jsonify({'ok': False, 'error': '缺少班组或月份'}), 400
    if not cards:
        return jsonify({'ok': False, 'error': '缺少评分数据'}), 400
    username = session.get('username', 'unknown')
    conn = get_conn(app.config['DATA_FOLDER'])
    try:
        # 删除该班组+月份+来源的旧卡（先删 entries 再删卡）
        old = conn.execute(
            "SELECT id FROM scoring_cards WHERE team=? AND month=? AND source=?",
            (team, month, source)).fetchall()
        for oc in old:
            conn.execute("DELETE FROM scoring_entries WHERE card_id=?", (oc['id'],))
        conn.execute("DELETE FROM scoring_cards WHERE team=? AND month=? AND source=?",
                     (team, month, source))
        # 一人一卡：card_no = custom_number 或 employee_id 兜底
        for card in cards:
            eid = card.get('employee_id', '')
            if not eid:
                continue
            wid = card.get('wid', '') or card.get('custom_number', '') or eid
            cur = conn.execute("""
                INSERT OR REPLACE INTO scoring_cards (week, team, card_no, source, month)
                VALUES (0,?,?,?,?)
            """, (team, wid, source, month))
            card_id = cur.lastrowid
            conn.execute("""
                INSERT OR REPLACE INTO scoring_entries (card_id, target_wid, target_employee_id,
                    initiative, diligence, discipline, cooperation, safety, driving)
                VALUES (?,?,?,?,?,?,?,?,?)
            """, (card_id, wid, eid,
                  int(card.get('initiative') or 0), int(card.get('diligence') or 0),
                  int(card.get('discipline') or 0), int(card.get('cooperation') or 0),
                  int(card.get('safety') or 0),
                  card.get('driving') if card.get('driving') is not None else None))
        conn.commit()
    finally:
        conn.close()
    log_audit(app.config['DATA_FOLDER'], 'scoring_card_batch', '',
              json.dumps({'team': team, 'month': month, 'source': source,
                          'count': len(cards), 'operator': username}),
              operator=session.get('username',''))
    return jsonify({'ok': True, 'count': len(cards)})

@app.route('/api/scoring/week/<int:team>/<int:week>', methods=['GET'])
@login_required
@require_permission('scoring', 'view')
def scoring_week_cards(team, week):
    from core.database import get_week_cards
    cards = get_week_cards(app.config['DATA_FOLDER'], team, week)
    return jsonify({'cards': cards})

@app.route('/api/scoring/summary/<int:team>', methods=['GET'])
@login_required
@require_permission('scoring', 'view')
def scoring_summary(team):
    """P15: 评分汇总（与 _get_scoring_bonus 共用共享函数，含去极值 + 三闸 R7 豁免）
    ?month= 缺省用当前月；响应含 pool 块 + individuals[].bonus"""
    from core.calculator import compute_scoring_pool, compute_team_bonuses
    from core.database import get_scoring_config
    from core.pricing import load_config
    data_folder = app.config['DATA_FOLDER']
    month = resolve_month(request)
    requested = (request.args.get('month') or '').strip()
    if requested and MONTH_RE.match(requested[:7]):
        month = requested[:7]
    _md_scoring = _get_month_data(month)
    cfg = (_md_scoring.get('config_snapshot') if _md_scoring else None) or APP_STATE.get('config') or {}
    ug_mode = cfg.get('underground_mode') or 'piecework'
    # P25-Q2: 周视图（week=1-5）→ 单周评分数据（无奖金池/三闸）；缺省/0 → 全月汇总
    week_arg = request.args.get('week', '')
    if week_arg not in ('', '0'):
        try:
            week = int(week_arg)
        except (TypeError, ValueError):
            return jsonify({'error': '无效周次'}), 400
        if not (1 <= week <= 5):
            return jsonify({'error': '周次需为 1-5'}), 400
        from core.calculator import compute_scoring_week
        indiv = compute_scoring_week(data_folder, team, week, month)
        result = []
        for eid, ind in indiv.items():
            row = dict(ind)
            row['employee_id'] = eid
            result.append(row)
        return jsonify({'individuals': result, 'week': week, 'month': month, 'underground_mode': ug_mode})
    # 产量层：与计薪同源 main_data（month 已过滤）
    pricing = load_config(data_folder)
    main_data = (_md_scoring.get('main_data') if _md_scoring else {}) if _md_scoring is not None else {}
    pool = compute_scoring_pool(main_data, pricing)
    # 单班全量（新表优先 + 旧表回退 + 守恒），内部已做分票/去极值/1.5票加权/系数
    tb = compute_team_bonuses(data_folder, team, month, pool)
    result = []
    for eid, ind in tb['individuals'].items():
        row = dict(ind)
        row['employee_id'] = eid
        row['bonus'] = tb['bonuses'].get(eid, 0)
        result.append(row)
    # 三闸 + R7 豁免（全班 final_behavior>=85 且 max(deviation)<10 → 最高档上限失效）
    behaviors = [r['final_behavior'] for r in result]
    deviations = [r['deviation'] for r in result]
    variance_range = max(behaviors) - min(behaviors) if len(behaviors) > 1 else 0
    config = get_scoring_config(data_folder)
    max_tier_count = sum(1 for r in result if r['coefficient'] >= 1.1)
    max_tier_ratio = max_tier_count / max(len(result), 1)
    deviation_threshold = config.get('mgmt_deviation_threshold', 15)
    dev_count = sum(1 for r in result if r['deviation'] > deviation_threshold)
    r7_exempt = bool(result) and all(r['final_behavior'] >= 85 for r in result) and max(deviations) < 10
    gates = {
        'zero_variance_triggered': variance_range <= config.get('zero_variance_threshold', 8) and len(behaviors) > 1,
        'max_tier_triggered': (max_tier_ratio > config.get('max_tier_ratio', 0.3)) and not r7_exempt,
        'mgmt_deviation_triggered': dev_count > 0,
        'variance_range': round(variance_range, 2),
        'max_tier_count': max_tier_count, 'max_tier_ratio': round(max_tier_ratio, 4),
        'deviation_count': dev_count,
        'r7_exempt': r7_exempt,
    }
    pool_block = {
        'nh_count': pool['nh_count'],
        'total_pool': pool['total_pool'],
        'half_pool': pool['half_pool'],
        'monthly_s': tb['monthly_s'],
        'distribution_ratio': tb['distribution_ratio'],
        'actual_pool': tb['actual_pool'],
        'sum_coef': tb['sum_coef'],
        'objective_missing': tb['objective_missing'],
        'conserved': tb['conserved'],
    }
    return jsonify({'individuals': result, 'gates': gates, 'pool': pool_block, 'month': month, 'underground_mode': ug_mode})

@app.route('/api/objective/entry', methods=['POST'])
@editor_required
def objective_entry():
    from core.database import save_objective_entry, log_audit
    data = request.get_json()
    # 实际出渣量改为手动录入（产量采集无按班组提交来源，不再从产量自动带出）
    data['actual_output'] = float(data.get('actual_output') or 0)
    daily_s = save_objective_entry(app.config['DATA_FOLDER'], data['record_date'],
        data['team'], data['planned_output'], data['actual_output'],
        data['total_hours'], data['effective_hours'], data['week'])
    log_audit(app.config['DATA_FOLDER'], 'objective_entry', session.get('username',''),
              json.dumps(data),
              operator=session.get('username',''))
    return jsonify({'ok': True, 'daily_s': daily_s})

@app.route('/api/objective/daily/<int:team>', methods=['GET'])
@login_required
def objective_daily(team):
    from core.database import get_objective_records
    records = get_objective_records(app.config['DATA_FOLDER'], team)
    return jsonify({'records': records})

@app.route('/api/objective/monthly/<int:team>', methods=['GET'])
@login_required
def objective_monthly(team):
    from core.database import get_monthly_objective
    month = resolve_month(request)
    requested = (request.args.get('month') or '').strip()
    if requested and MONTH_RE.match(requested[:7]):
        month = requested[:7]
    md = _get_month_data(month)
    summary = get_monthly_objective(app.config['DATA_FOLDER'], team, month or None)
    summary['month'] = month
    return jsonify(summary)

@app.route('/api/scoring/config', methods=['GET'])
@login_required
@require_permission('scoring', 'view')
def scoring_config_get():
    from core.database import get_scoring_config
    config = get_scoring_config(app.config['DATA_FOLDER'])
    return jsonify(config)

@app.route('/api/scoring/config', methods=['POST'])
@editor_required
def scoring_config_save():
    from core.database import save_scoring_config, log_audit
    data = request.get_json()
    save_scoring_config(app.config['DATA_FOLDER'], data)
    log_audit(app.config['DATA_FOLDER'], 'scoring_config', session.get('username',''), json.dumps(data),
              operator=session.get('username',''))
    return jsonify({'ok': True})

# ═══════════════════════════════════════════════════════════
#  API: NSSF（社保）
# ═══════════════════════════════════════════════════════════

@app.route('/nssf/list', methods=['GET'])
@login_required
def get_nssf_list():
    """获取 NSSF 参保状态列表"""
    from core.nssf import load_nssf_enrollment
    enrollment = load_nssf_enrollment(app.config['DATA_FOLDER'])
    sdl = APP_STATE.get('nssf_sdl_members', {})
    return jsonify({
        'enrollment': enrollment,
        'sdl_members': {k: v['name'] for k, v in sdl.items()},
    })

@app.route('/nssf/toggle', methods=['POST'])
@editor_required
def toggle_nssf():
    """切换某人的 NSSF 参保状态"""
    data = request.get_json()
    eid = data.get('employee_id')
    enrolled = data.get('enrolled', False)
    from core.nssf import save_nssf_enrollment
    save_nssf_enrollment(app.config['DATA_FOLDER'], eid, enrolled)
    # 同步内存状态
    for emp in APP_STATE.get('employees', []):
        if emp['id'] == eid:
            emp['nssf_enrolled'] = enrolled
            break
    _audit('nssf_toggle', eid, json.dumps({'enrolled': enrolled}))
    return jsonify({'ok': True})

# ═══════════════════════════════════════════════════════════
#  API: 通讯录
# ═══════════════════════════════════════════════════════════

@app.route('/addressbook', methods=['GET'])
@login_required
def get_addressbook():
    book = APP_STATE.get('address_book', {})
    from collections import defaultdict
    by_dept = defaultdict(list)
    for eid, info in book.items():
        dept = info.get('department', '未分类')
        by_dept[dept].append({'id': eid, **info})
    sorted_depts = sorted(by_dept.items(), key=lambda x: -len(x[1]))
    return jsonify({
        'total': len(book),
        'departments': [{'name': d, 'count': len(p), 'people': p} for d, p in sorted_depts]
    })

# ═══════════════════════════════════════════════════════════
#  API: 计算参数配置
# ═══════════════════════════════════════════════════════════

@app.route('/config', methods=['GET'])
@login_required
@require_permission('system', 'view')
def get_config():
    from core.pricing import load_config
    return jsonify(load_config(app.config['DATA_FOLDER']))

@app.route('/config', methods=['POST'])
@admin_required
def save_config():
    from core.pricing import load_config, save_config as _save_cfg
    incoming = request.get_json() or {}
    config = load_config(app.config.get('DATA_FOLDER'))
    config.update(incoming)
    _save_cfg(app.config.get('DATA_FOLDER'), config)
    APP_STATE.update({'config': config})
    with MONTH_CACHE_LOCK:
        for _v in MONTH_CACHE.values():
            _v['config_snapshot'] = _copy.deepcopy(config)
    _audit('config_update', '', json.dumps({'keys': list(incoming.keys())}))
    return jsonify({'ok': True, 'config': config})

# ═══════════════════════════════════════════════════════════
#  API: 计算/薪资
# ═══════════════════════════════════════════════════════════

@app.route('/recalculate', methods=['POST'])
@admin_required
def recalculate():
    # map: POST /recalculate -> MONTH_CACHE[g.view_month] per-month recompute
    result = _recalc_internal()
    if result is None:
        return jsonify({'ok': False, 'error': '请先加载数据'})
    return jsonify({'ok': True, 'result': result})

def _recalc_internal(month: str | None = None):
    """Per-month recompute via _run_pipeline (includes V2 apply_v2_month_end per month)."""
    if not APP_STATE.get('parsed'):
        return None
    # resolve affected month: explicit arg > g.view_month > APP_STATE month
    if month is None:
        try:
            month = g.view_month
        except Exception:
            month = APP_STATE.get('month')
    # per-month evict then rebuild (full pipeline ensures V2 coefficient per month)
    _invalidate_month_cache(month)
    md = _run_pipeline(month_filter=month if month != 'all' else None)
    if md is None:
        return None
    result = md.get('salary_result')
    _audit('recalculate', '', json.dumps({'month': month, 'total_gross': (result or {}).get('total_gross', 0)}))
    return result

@app.route('/salary', methods=['GET'])
@login_required
@require_permission('salary', 'view')
def get_salary():
    month = resolve_month(request)
    md = _get_month_data(month)
    if md is None or md.get('salary_result') is None:
        return jsonify({'result': None, 'month': month, 'headless': True})
    res = md.get('salary_result')
    if isinstance(res, dict):
        res = {**res, 'month': month}
    return jsonify({'result': res, 'month': month, 'headless': bool(md.get('headless', False))})

@app.route('/api/salary/inline-edit', methods=['POST'])
@login_required
@require_permission('employees', 'edit')
def api_salary_inline_edit():
    """P29-c: 薪资总表行内编辑——奖金/罚款合并写 bonus_penalties,预支直改 employees.advance。
    前端保存后重拉 GET /salary(GET 每次全量重算)即见自动计算结果,无需 /recalculate 权限。"""
    from core.database import save_bonus_penalty, load_bonus_penalties, get_conn
    data = request.get_json(silent=True) or {}
    eid = (data.get('employee_id') or '').strip()
    field = (data.get('field') or '').strip()
    month = (data.get('month') or '').strip() or datetime.now(EAT).strftime('%Y-%m')
    try:
        value = int(data.get('value'))
    except (TypeError, ValueError):
        return jsonify({'ok': False, 'error': '数值非法'}), 400
    if value < 0:
        return jsonify({'ok': False, 'error': '数值不能为负'}), 400
    if field not in ('bonus', 'penalty', 'advance'):
        return jsonify({'ok': False, 'error': '不支持的字段'}), 400
    if not eid:
        return jsonify({'ok': False, 'error': '缺少 employee_id'}), 400
    old = None
    conn = get_conn(app.config['DATA_FOLDER'])
    row = conn.execute('SELECT 1 FROM employees WHERE id=?', (eid,)).fetchone()
    conn.close()
    if not row:
        return jsonify({'ok': False, 'error': '员工不存在'}), 404
    bp = load_bonus_penalties(app.config['DATA_FOLDER'], month) or {}
    cur = bp.get(eid) or {}
    old = int(cur.get(field, 0) or 0)
    bonus = value if field == 'bonus' else int(cur.get('bonus', 0) or 0)
    penalty = value if field == 'penalty' else int(cur.get('penalty', 0) or 0)
    advance = value if field == 'advance' else int(cur.get('advance', 0) or 0)
    save_bonus_penalty(app.config['DATA_FOLDER'], eid, month, bonus, penalty, advance)
    # map: POST /api/salary/inline-edit -> MONTH_CACHE[month[:7]] per-month evict
    _invalidate_month_cache(month[:7])
    _audit('salary_inline_edit', eid,
           json.dumps({'user': session.get('username', ''), 'month': month,
                       'field': field, 'old': old, 'new': value, 'invalidated': month[:7]}))
    return jsonify({'ok': True})

# ═══════════════════════════════════════════════════════════
#  API: 产量
# ═══════════════════════════════════════════════════════════

@app.route('/production', methods=['GET'])
@login_required
@require_permission('dashboard', 'view')  # P29 T4 A9: production:view → dashboard:view
def get_production():
    month = resolve_month(request)
    _md = _get_month_data(month)
    md = (_md.get('main_data') if _md else {}) if _md is not None else {}
    shift_prod = md.get('shift_production', [])
    driller_prod = md.get('driller_production', [])

    shift_daily = []
    for d in shift_prod:
        dp = d.get('day_prod') or {}
        np = d.get('night_prod') or {}
        shift_daily.append({
            'date': d['date'],
            'nh': (dp.get('NICKEL（H）', 0) or 0) + (np.get('NICKEL（H）', 0) or 0),
            'nl': (dp.get('NICKEL（L）', 0) or 0) + (np.get('NICKEL（L）', 0) or 0),
            'mw': (dp.get('MAWE', 0) or 0) + (np.get('MAWE', 0) or 0),
        })

    from collections import defaultdict
    cap_totals = defaultdict(lambda: {'nh': 0, 'nl': 0, 'mw': 0, 'futa': 0, 'waya': 0, 'kibiriti': 0, 'name': ''})
    for d in driller_prod:
        cap = d['captain']
        cap_totals[cap]['name'] = cap
        cap_totals[cap]['nh'] += d['nh']
        cap_totals[cap]['nl'] += d['nl']
        cap_totals[cap]['mw'] += d['mw']
        cap_totals[cap]['futa'] += d['futa']
        cap_totals[cap]['waya'] += d['waya']
        cap_totals[cap]['kibiriti'] += d['kibiriti']

    driller_summary = [v for _, v in sorted(cap_totals.items())]
    return jsonify({
        'month': month,
        'shift_production': shift_daily,
        'driller_production': driller_summary,
        'driller_consumables': [{'name': v['name'],
                                  'futa': v['futa'], 'waya': v['waya'],
                                  'kibiriti': v['kibiriti']}
                                for v in driller_summary],
    })

# ═══════════════════════════════════════════════════════════
#  API: 数据台产量仪表盘（P15: 白夜班分离 + 钻工逐日明细 + 破碎）
# ═══════════════════════════════════════════════════════════

@app.route('/api/production/dashboard', methods=['GET'])
@login_required
@require_permission('dashboard', 'view')  # P29 T4 A9: production:view → dashboard:view
def get_production_dashboard():
    month = resolve_month(request)
    _md = _get_month_data(month)
    md = (_md.get('main_data') if _md else {}) if _md is not None else {}
    shift_prod = md.get('shift_production', [])
    driller_prod = md.get('driller_production', [])
    crush_prod = md.get('crush_production', [])

    # ── 井下产量: 白班/夜班/合计 三者分离（旧） / 按团队（新 C8） ──
    group_names = {}
    try:
        from core.database import list_employee_groups as _list_groups
        for g in _list_groups(app.config['DATA_FOLDER']):
            group_names[int(g['id'])] = g['name']
    except Exception:
        pass
    # 井下计件单价：V2 凸性加速（v2 生效月起）用 accel_prices+加速系数；否则用 underground_prices 线性兜底
    # 班组薪资 = 产量 × 单价 × 加速系数（系数=当日总车次/accel_target，豁免日系数=1.0），与出勤无关
    from core.pricing import load_config as _load_cfg
    _cfg = _load_cfg(app.config['DATA_FOLDER']) or {}
    _mode = _cfg.get('underground_mode', 'piecework')
    _v2_from = _cfg.get('v2_effective_from', '') or ''
    _v2_active = (_mode == 'v2' and (not _v2_from or month >= _v2_from))
    _target = int(_cfg.get('accel_target', 40) or 40) if _v2_active else 40
    _accel_price = (_cfg.get('accel_prices') or {}) or {}
    _up_price = (_cfg.get('underground_prices') or {}) or {}
    if _v2_active:
        _ph_p = _accel_price.get('NICKEL（H）', 8000) or 8000
        _pl_p = _accel_price.get('NICKEL（L）', 5000) or 5000
        _pm_p = _accel_price.get('MAWE', 3000) or 3000
    else:
        _ph_p = _up_price.get('NICKEL（H）', 6000) or 6000
        _pl_p = _up_price.get('NICKEL（L）', 5000) or 5000
        _pm_p = _up_price.get('MAWE', 4000) or 4000
    shift_daily = []
    for d in shift_prod:
        if 'teams' in d:
            teams_out = []
            for t in (d.get('teams') or []):
                prod = t.get('prod') or {}
                nh = prod.get('NICKEL（H）', 0) or 0
                nl = prod.get('NICKEL（L）', 0) or 0
                mw = prod.get('MAWE', 0) or 0
                tid = int(t.get('team_id', 0) or 0)
                # V2 班组薪资池：Σ(产量×单价)×加速系数（系数=当日总车次/accel_target；豁免日=1.0）
                _cars = nh + nl + mw
                _mult = 1.0 if t.get('exempt', False) else ((_cars / _target) if _target else 1.0)
                nh_amt = round(nh * _ph_p * _mult)
                nl_amt = round(nl * _pl_p * _mult)
                mw_amt = round(mw * _pm_p * _mult)
                teams_out.append({
                    'team_id': tid,
                    'team_name': group_names.get(tid, ''),
                    'nh': nh, 'nl': nl, 'mw': mw,
                    'total': _cars,
                    'coef': round(_mult, 4),
                    'nh_amt': nh_amt, 'nl_amt': nl_amt, 'mw_amt': mw_amt,
                    'salary': nh_amt + nl_amt + mw_amt,
                })
            shift_daily.append({'date': d['date'], 'teams': teams_out})
        else:
            dp = d.get('day_prod') or {}
            np = d.get('night_prod') or {}
            shift_daily.append({
                'date': d['date'],
                'day_nh': dp.get('NICKEL（H）', 0) or 0,
                'day_nl': dp.get('NICKEL（L）', 0) or 0,
                'day_mw': dp.get('MAWE', 0) or 0,
                'night_nh': np.get('NICKEL（H）', 0) or 0,
                'night_nl': np.get('NICKEL（L）', 0) or 0,
                'night_mw': np.get('MAWE', 0) or 0,
                'total_nh': (dp.get('NICKEL（H）', 0) or 0) + (np.get('NICKEL（H）', 0) or 0),
                'total_nl': (dp.get('NICKEL（L）', 0) or 0) + (np.get('NICKEL（L）', 0) or 0),
                'total_mw': (dp.get('MAWE', 0) or 0) + (np.get('MAWE', 0) or 0),
            })

    # ── 钻工产量: 逐日明细（非队长汇总）──
    driller_daily = []
    for d in driller_prod:
        driller_daily.append({
            'date': d['date'],
            'captain': d['captain'],
            'nh': d['nh'],
            'nl': d['nl'],
            'mw': d['mw'],
            'members': d.get('members', []),
        })

    # ── 破碎产量: 逐日 ──
    crush_daily = []
    for c in crush_prod:
        crush_daily.append({
            'date': c['date'],
            'bags': c['bags'],
            'personnel_count': len(c.get('personnel', [])),
        })

    return jsonify({
        'month': month,
        'shift_production': shift_daily,
        'driller_production': driller_daily,
        'crush_production': crush_daily,
    })


# ═══════════════════════════════════════════════════════════
#  API: 产量核验（逐日对比钻工组与井下合计）
# ═══════════════════════════════════════════════════════════

@app.route('/production-verify', methods=['GET'])
@login_required
def get_production_verify():
    """返回逐日钻工组产量与井下产量对比（P31 A1: teams 新格式分支 + 双键门控）"""
    # P31 A1: 权限放宽 salary:view OR dashboard:view（沿用 P29 _oa_read_gate 双键写法，
    # 403 形状与 @require_permission 一致；核验卡随 P31 对 dashboard:view 用户开放）
    from core.database import check_permission
    _u = session.get('username', '')
    if not (check_permission(app.config['DATA_FOLDER'], _u, 'salary', 'view')
            or check_permission(app.config['DATA_FOLDER'], _u, 'dashboard', 'view')):
        _audit('perm_denied', '', json.dumps({'user': _u, 'module': 'salary', 'action': 'view'}))
        return jsonify({'ok': False, 'error': 'forbidden', 'need_permission': 'salary'}), 403
    month = resolve_month(request)
    _md = _get_month_data(month)
    md = (_md.get('main_data') if _md else {}) if _md is not None else {}
    shift_prod = md.get('shift_production', [])
    driller_prod = md.get('driller_production', [])

    # 井下逐日合计（P31 A1 格式分支铁律：rec 含 teams 键=新格式 Σ班组 / 否则旧格式 白+夜逐字保留）
    shift_daily = {}
    _has_teams_fmt = False
    for d in shift_prod:
        dt = d['date']
        if 'teams' in d:
            _has_teams_fmt = True
            if dt not in shift_daily:
                shift_daily[dt] = {'nh': 0, 'nl': 0, 'mw': 0}
            for _t in (d.get('teams') or []):
                _prod = _t.get('prod') or {}
                shift_daily[dt]['nh'] += _prod.get('NICKEL（H）', 0) or 0
                shift_daily[dt]['nl'] += _prod.get('NICKEL（L）', 0) or 0
                shift_daily[dt]['mw'] += _prod.get('MAWE', 0) or 0
        else:
            dp = d.get('day_prod') or {}
            np = d.get('night_prod') or {}
            if dt not in shift_daily:
                shift_daily[dt] = {'nh': 0, 'nl': 0, 'mw': 0}
            shift_daily[dt]['nh'] += (dp.get('NICKEL（H）', 0) or 0) + (np.get('NICKEL（H）', 0) or 0)
            shift_daily[dt]['nl'] += (dp.get('NICKEL（L）', 0) or 0) + (np.get('NICKEL（L）', 0) or 0)
            shift_daily[dt]['mw'] += (dp.get('MAWE', 0) or 0) + (np.get('MAWE', 0) or 0)

    # 钻工逐日分组
    from collections import defaultdict
    driller_daily = defaultdict(lambda: defaultdict(lambda: {'nh': 0, 'nl': 0, 'mw': 0}))
    for d in driller_prod:
        dt = d['date']
        cap = d['captain']
        driller_daily[dt][cap]['nh'] += d['nh']
        driller_daily[dt][cap]['nl'] += d['nl']
        driller_daily[dt][cap]['mw'] += d['mw']

    # 构建返回数据
    all_dates = sorted(set(list(shift_daily.keys()) + list(driller_daily.keys())))
    result = {}
    for dt in all_dates:
        dtot = {'nh': 0, 'nl': 0, 'mw': 0}
        groups = []
        for cap, g in sorted(driller_daily.get(dt, {}).items()):
            groups.append({'captain': cap, 'nh': g['nh'], 'nl': g['nl'], 'mw': g['mw']})
            dtot['nh'] += g['nh']; dtot['nl'] += g['nl']; dtot['mw'] += g['mw']
        st = shift_daily.get(dt, {'nh': 0, 'nl': 0, 'mw': 0})
        result[dt] = {
            'driller_groups': groups,
            'driller_total': dtot,
            'shift_total': st,
            'match': dtot['nh'] == st['nh'] and dtot['nl'] == st['nl'] and dtot['mw'] == st['mw'],
        }
    result['month'] = month
    # P31 A2: 顶层汇总（仅 teams 新格式月下发；旧格式月响应与改造前逐字段一致）
    if _has_teams_fmt:
        _days = [v for k, v in result.items() if k != 'month']
        result['verify_days'] = len(_days)
        result['match_days'] = sum(1 for v in _days if v['match'])
        result['mismatch_days'] = result['verify_days'] - result['match_days']
    return jsonify(result)

# ═══════════════════════════════════════════════════════════
#  API: 逐日工资明细
# ═══════════════════════════════════════════════════════════

@app.route('/daily-wages', methods=['GET'])
@login_required
@require_permission('salary', 'view')
def get_daily_wages():
    from core.calculator import compute_daily_breakdown
    from core.exceptions import load_overrides, load_daily_exclusions
    month = resolve_month(request)
    _md = _get_month_data(month)
    cur_md_dw = (_md.get('main_data') if _md else {}) if _md is not None else {}
    if not cur_md_dw:
        return jsonify({})
    _emps_dw = (_md.get('employees') if _md else []) if _md is not None else []
    _cfg_dw = (_md.get('config_snapshot') if _md else {}) if _md is not None else {}
    _ug_dw = _build_ug_team_members(app.config.get('DATA_FOLDER'))
    try:
        result = compute_daily_breakdown(
            main_data=cur_md_dw,
            employees=_emps_dw,
            overrides=load_overrides(app.config['DATA_FOLDER'], month=month),
            exclusions=load_daily_exclusions(app.config['DATA_FOLDER']),
            pricing=_cfg_dw,
            data_folder=app.config.get('DATA_FOLDER'),
            ug_team_members=_ug_dw,
        )
    except TypeError:
        result = compute_daily_breakdown(
            main_data=cur_md_dw,
            employees=_emps_dw,
            overrides=load_overrides(app.config['DATA_FOLDER'], month=month),
            exclusions=load_daily_exclusions(app.config['DATA_FOLDER']),
            pricing=_cfg_dw,
            data_folder=app.config.get('DATA_FOLDER'),
        )
    # 合并出勤手动覆盖（P/A/L）到逐日工资结果
    import sqlite3, os
    att_ov = {}
    db_path = os.path.join(app.config['DATA_FOLDER'], 'kilwa.db')
    if os.path.exists(db_path):
        conn = sqlite3.connect(db_path)
        try:
            for r in conn.execute("SELECT employee_id, date FROM attendance_overrides").fetchall():
                key = f"{r[0]}|{r[1]}"
                att_ov[key] = True
        except: pass
        conn.close()
    # 构建每个员工的 att_override_dates
    for eid, e in result.items():
        att_dates = []
        for dt in e.get('daily', {}):
            if f"{eid}|{dt}" in att_ov:
                att_dates.append(dt)
        if att_dates:
            e['att_override_dates'] = att_dates
        else:
            e['att_override_dates'] = []
    return jsonify(result)

# ═══════════════════════════════════════════════════════════
#  API: 钻工队长列表（用于临时例外弹窗）
# ═══════════════════════════════════════════════════════════

@app.route('/driller-captains', methods=['GET'])
@login_required
def get_driller_captains():
    """返回所有钻工队长列表（不按日期过滤，用于临时例外弹窗）"""
    md = APP_STATE.get('main_data', {})
    captains = set()
    for d in md.get('driller_production', []):
        captains.add(d['captain'])
    return jsonify(sorted(captains))

# ═══════════════════════════════════════════════════════════
#  API: 出勤网格
# ═══════════════════════════════════════════════════════════

def _build_attendance_grid(md=None, employees=None):
    if md is None:
        try:
            cur = g.view_month
        except Exception:
            try:
                cur = resolve_month(request)
            except Exception:
                cur = APP_STATE.get('month')
        md_data = _get_month_data(cur) if cur else None
        md = md_data.get('main_data') if md_data else APP_STATE.get('main_data', {})
        if employees is None and md_data:
            employees = md_data.get('employees')
    if employees is None:
        try:
            _cur2 = g.view_month
            _md2 = _get_month_data(_cur2)
            if _md2 and _md2.get('employees'):
                employees = _md2.get('employees')
            else:
                employees = APP_STATE.get('employees', [])
        except Exception:
            employees = APP_STATE.get('employees', [])
    from collections import defaultdict
    from core.namematch import make_employee_id
    shift_prod = md.get('shift_production', []) if md else []
    driller_prod = md.get('driller_production', []) if md else []
    attendance_data = md.get('attendance', []) if md else []

    # 收集所有日期（含钻工+破碎计件日期）
    all_dates = sorted(set(
        list(set(d['date'] for d in shift_prod)) +
        list(set(d['date'] for d in driller_prod)) +
        list(set(d.get('date', '') for d in attendance_data)) +
        list(set(d.get('date', '') for d in (md.get('crush_production') or []))) +
        list(md.get('dates', []))
    ))
    # 补全当月全部自然日（确保没有数据时也能看到所有日期）
    if all_dates:
        from calendar import monthrange
        ym = all_dates[0][:7]
        if ym and len(ym) == 7:
            y, m = int(ym[:4]), int(ym[5:7])
            _, last_day = monthrange(y, m)
            all_dates = sorted(set(all_dates) | set(f"{ym}-{d:02d}" for d in range(1, last_day + 1)))

    # 收集每人每天的状态（原始来源）
    day_status = defaultdict(dict)
    day_origin = defaultdict(dict)  # 'auto' or 'manual'

    # 加载手动覆盖
    from core.database import load_attendance_overrides
    manual = load_attendance_overrides(app.config['DATA_FOLDER'])

    # 生产薪资：白班=D 夜班=N 全天=B
    for d in shift_prod:
        dt = d['date']
        for e in d.get('day_emps', []):
            from core.namematch import make_employee_id
            eid = make_employee_id(e)
            if eid:
                day_status[eid][dt] = 'D'
                day_origin[eid][dt] = 'auto'
        for e in d.get('night_emps', []):
            eid = make_employee_id(e)
            if eid:
                existing = day_status.get(eid, {}).get(dt, '')
                day_status[eid][dt] = 'B' if existing == 'D' else 'N'
                day_origin[eid][dt] = 'auto'

    # 钻工出勤
    for d in driller_prod:
        dt = d['date']
        from core.namematch import make_employee_id, canonical
        cap_id = make_employee_id(d['captain'])
        if cap_id and dt not in day_status.get(cap_id, {}):
            day_status[cap_id][dt] = 'R'
            day_origin[cap_id][dt] = 'auto'
        for m in d.get('members', []):
            mid = make_employee_id(m)
            if mid and dt not in day_status.get(mid, {}):
                day_status[mid][dt] = 'R'
                day_origin[mid][dt] = 'auto'

    # 破碎计件出勤
    crush_data = md.get('crush_production', [])
    for d in crush_data:
        dt = d['date']
        for e in d.get('personnel', []):
            eid = make_employee_id(e)
            if eid and dt not in day_status.get(eid, {}):
                day_status[eid][dt] = 'C'
                day_origin[eid][dt] = 'auto'

    # 日薪出勤
    for d in attendance_data:
        dt = d['date']
        for e in d.get('normal', []):
            eid = make_employee_id(e)
            if eid and dt not in day_status.get(eid, {}):
                day_status[eid][dt] = 'P'
                day_origin[eid][dt] = 'auto'

    # 月薪默认出勤
    type_labels = {'piece_crush': '破碎计件','piece_underground':'生产薪资','piece_driller':'钻工计件','day_rate':'日薪','monthly':'月薪','advance_only':'仅预支','address_book':'通讯录'}
    rows = []

    for emp in employees:
        eid = emp['id']
        status_row = {}
        origin_row = {}
        auto_row = {}

        for dt in all_dates:
            # 手动覆盖优先
            mkey = f'{eid}|{dt}'
            if mkey in manual:
                status_row[dt] = manual[mkey]['status']
                origin_row[dt] = 'manual' if manual[mkey].get('source', 0) == 1 else 'collection'
            else:
                auto_val = day_status.get(eid, {}).get(dt, '')
                # 顶层部门月薪人员：数据为空时默认全勤
                if not auto_val and emp.get('department') == 'ENPRIZON LINDI PROJECT' and (emp.get('override_type') == 'monthly' or emp.get('default_type') == 'monthly'):
                    auto_val = 'P'
                status_row[dt] = auto_val
                origin_row[dt] = 'auto' if auto_val else ''
            # 原始自动值（用于前端判断是否与手动不同）
            raw_auto = day_status.get(eid, {}).get(dt, '')
            # 顶层部门月薪人员：标记(P)显示灰色背景
            if not raw_auto and emp.get('department') == 'ENPRIZON LINDI PROJECT' and (emp.get('override_type') == 'monthly' or emp.get('default_type') == 'monthly'):
                raw_auto = '(P)'

            auto_row[dt] = raw_auto

        is_super_admin = session.get('role') == 'super_admin'
        rows.append({
            'id': eid,
            'name': emp.get('name', ''),
            'type': type_labels.get(emp.get('override_type') or emp.get('default_type', ''), emp.get('default_type', '')),
            'department': emp.get('department', ''),
            'custom_number': emp.get('custom_number', '') or '',
            'days': status_row,
            'origin': origin_row,
            'auto': auto_row,
            'editable': is_super_admin,
        })

    return {'dates': all_dates, 'rows': rows}


@app.route('/attendance', methods=['GET'])
@login_required
@require_permission('attendance', 'view')
def get_attendance():
    """返回出勤网格：每人每天的状态。P=出勤 A=旷工 L=请假"""
    month = resolve_month(request)
    md = _get_month_data(month)
    if md is not None:
        grid = _build_attendance_grid(md.get('main_data'), md.get('employees'))
        grid['month'] = month
        return jsonify(grid)
    return jsonify(_build_attendance_grid())


@app.route('/attendance/toggle', methods=['POST'])
@super_admin_required
@require_permission('attendance', 'edit')
def toggle_attendance():
    """手动标某人某天的状态：P出勤 A旷工 L请假"""
    import json as _json
    month = resolve_month(request)
    md = _get_month_data(month)
    data = request.get_json()
    eid = data.get('employee_id')
    date = data.get('date')
    status = data.get('status', 'P')

    from core.database import save_attendance_override, get_attendance_status
    if status == 'NU':
        return jsonify({'ok': False, 'error': 'NU（年假）状态由审批管理，禁止手动修改'}), 403
    if get_attendance_status(app.config['DATA_FOLDER'], eid, date) == 'NU':
        return jsonify({'ok': False, 'error': 'NU（年假）状态由审批管理，禁止手动修改'}), 403
    try:
        save_attendance_override(app.config['DATA_FOLDER'], eid, date, status, source=1)
    except ValueError as e:
        return jsonify({'ok': False, 'error': str(e)}), 400
    _invalidate_month_cache((date or '')[:7] or month)
    _invalidate_all_cache_base()
    _audit('attendance_toggle', eid, json.dumps({'date': date, 'status': status, 'month': month, 'invalidated': (date or '')[:7] or month}))
    return jsonify({'ok': True, 'month': month})

# ═══════════════════════════════════════════════════════════
#  API: 审计日志
# ═══════════════════════════════════════════════════════════

@app.route('/audit-log', methods=['GET'])
@admin_required
def get_audit_log():
    from core.database import get_audit_logs
    logs = get_audit_logs(app.config['DATA_FOLDER'])
    return jsonify(logs)


# ═══════════════════════════════════════════════════════════
#  API: 导出 Excel
# ═══════════════════════════════════════════════════════════

@app.route('/export', methods=['POST'])
@login_required
@require_permission('salary', 'export')
def export_salary():
    month = resolve_month(request)
    md_wrap = _get_month_data(month)
    if EXPORT_STRICT_MONTH:
        stamp = _month_stamp()
        if stamp and month != stamp:
            return jsonify({'ok': False, 'error': 'month_mismatch', 'expected': stamp}), 409
    if md_wrap is None or md_wrap.get('salary_result') is None:
        return jsonify({'ok': False, 'error': '请先计算薪资'})
    result = md_wrap.get('salary_result')
    md = md_wrap.get('main_data', {})
    eff_month = month

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Salary Summary'

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill('solid', fgColor='185FA5')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                          top=Side(style='thin'), bottom=Side(style='thin'))

    headers = ['Name', 'Type', 'Underground Piece Rate(TZS)', 'Driller Piece(TZS)', 'Crush Piece(TZS)',
               'Day Rate(TZS)', 'Monthly(TZS)', 'Overtime(TZS)', 'Bonus(TZS)', 'Driver Allowance(TZS)',
               'Gross Total(TZS)', 'NSSF(TZS)', 'PAYE(TZS)', 'Company PAYE (50%)(TZS)',
               'Penalty(TZS)', 'Advance Deduction(TZS)', 'Net Salary(TZS)']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = header_align; cell.border = thin_border

    type_map = {'piece_crush': 'Crush Piece', 'piece_underground': 'Underground Piece Rate', 'piece_driller': 'Driller Piece',
                'day_rate': 'Day Rate', 'monthly': 'Monthly', 'both': 'Unspecified', 'advance_only': 'Advance Only', 'address_book': 'Address Book'}
    total_fill = PatternFill('solid', fgColor='FFF3CD')

    for i, emp in enumerate(result['employees'], 2):
        gross = emp.get('gross', 0) or 0
        bonus = int(emp.get('bonus', 0) or 0)
        penalty = int(emp.get('penalty', 0) or 0)
        nssf = emp.get('nssf', 0) or 0
        driver = int(emp.get('driver_allowance', 0) or 0)
        overtime = int(emp.get('overtime', 0) or 0)
        paye = int(emp.get('paye', 0) or 0)
        paye_half = int(emp.get('paye_half', 0) or 0)
        vals = [
            emp['name'] or '', type_map.get(emp.get('salary_type', ''), emp.get('salary_type', '')),
            int(emp.get('piece_underground', 0) or 0), int(emp.get('piece_driller', 0) or 0),
            int(emp.get('piece_crush', 0) or 0),
            int(emp.get('day_rate', 0) or 0), int(emp.get('monthly', 0) or 0),
            overtime, bonus, driver,
            int(gross), int(nssf),
            paye, paye_half, penalty, int(emp.get('advance', 0) or 0),
            int(emp.get('net', 0) or 0),
        ]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(i, col, v); cell.border = thin_border
            cell.alignment = Alignment(horizontal='left' if col == 1 else 'right')
            if col > 1: cell.number_format = '#,##0'

    total_row = len(result['employees']) + 2
    ws.cell(total_row, 1, 'Total').font = Font(bold=True, size=11)
    ws.cell(total_row, 1).fill = total_fill; ws.cell(total_row, 1).border = thin_border

    # 井下(C), 钻工(D), 破碎(E), 日薪(F), 月薪(G), 加班(H), 奖金(I), 司机(J), 应发(K), NSSF(L), PAYE(M), 公司代付(N), 罚款(O), 预支(P) → SUM公式
    for ci in [3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16]:
        letter = chr(64 + ci)
        cell = ws.cell(total_row, ci, f'=SUM({letter}2:{letter}{total_row-1})')
        cell.font = Font(bold=True); cell.fill = total_fill; cell.border = thin_border
        cell.number_format = '#,##0'

    # 实发(17=Q) = Σ逐行后端 net
    ws.cell(total_row, 17, f'=SUM(Q2:Q{total_row-1})')
    ws.cell(total_row, 17).font = Font(bold=True)
    ws.cell(total_row, 17).fill = total_fill; ws.cell(total_row, 17).border = thin_border
    ws.cell(total_row, 17).number_format = '#,##0'

    for i, w in enumerate([18, 12, 16, 16, 16, 16, 16, 16, 14, 16, 16, 14, 16, 16, 16, 16, 16], 1):
        ws.column_dimensions[chr(64+i)].width = w

    # Sheet 2: 产量
    ws2 = wb.create_sheet('Production Summary')
    for ci, h in enumerate(['Date', 'NICKEL(H)', 'NICKEL(L)', 'MAWE'], 1):
        c = ws2.cell(1, ci, h); c.font = header_font; c.fill = header_fill

    md2 = md if 'md' in locals() and isinstance(md, dict) else (md_wrap.get('main_data', {}) if 'md_wrap' in locals() and md_wrap else {})
    for i, d in enumerate(md2.get('shift_production', []), 2):
        dp = d.get('day_prod') or {}; np = d.get('night_prod') or {}
        ws2.cell(i, 1, d['date'])
        ws2.cell(i, 2, (dp.get('NICKEL（H）', 0) or 0) + (np.get('NICKEL（H）', 0) or 0))
        ws2.cell(i, 3, (dp.get('NICKEL（L）', 0) or 0) + (np.get('NICKEL（L）', 0) or 0))
        ws2.cell(i, 4, (dp.get('MAWE', 0) or 0) + (np.get('MAWE', 0) or 0))

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    fname = f'ENPRIZON_LINDI_Salary_{eff_month}.xlsx' if eff_month else 'ENPRIZON_LINDI_Salary.xlsx'
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=fname)

# ═══════════════════════════════════════════════════════════
#  API: 导出员工信息表
# ═══════════════════════════════════════════════════════════

@app.route('/export/employees', methods=['POST'])
@login_required
@require_permission('employees', 'export')
def export_employees():
    """导出员工信息表（薪资类型、日薪基数、月薪基数、预支）"""
    month = resolve_month(request)
    md_wrap = _get_month_data(month)
    if EXPORT_STRICT_MONTH:
        stamp = _month_stamp()
        if stamp and month != stamp:
            return jsonify({'ok': False, 'error': 'month_mismatch', 'expected': stamp}), 409
    eff_month = month
    employees = (md_wrap.get('employees') if md_wrap else []) if md_wrap is not None else []
    if not employees:
        return jsonify({'ok': False, 'error': '无员工数据'})

    from core.exceptions import load_overrides
    overrides = load_overrides(app.config['DATA_FOLDER'], month=eff_month)

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Employee Info'

    hf = Font(bold=True, color='FFFFFF', size=11)
    hfill = PatternFill('solid', fgColor='185FA5')
    ha = Alignment(horizontal='center', vertical='center')
    tb = Border(left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))

    headers = ['Name', 'Department', 'Type', 'NIDA Number', 'NSSF Number', 'TIN Number', 'Day Rate(TZS)', 'Monthly Base(TZS)', 'Advance This Month(TZS)', 'Notes']
    for ci, h in enumerate(headers, 1):
        c = ws.cell(1, ci, h); c.font = hf; c.fill = hfill; c.alignment = ha; c.border = tb

    type_map = {'piece_underground':'Underground Piece Rate','piece_driller':'Driller Piece',
                'day_rate':'Day Rate','monthly':'Monthly','both':'Unspecified','advance_only':'Advance Only','address_book':'Address Book'}
    total_fill = PatternFill('solid', fgColor='FFF3CD')

    for i, emp in enumerate(employees, 2):
        eid = emp['id']
        note = emp.get('_note', '')
        if not note and eid in overrides:
            note = '; '.join(f"{o.get('start_date','')}~{o.get('end_date','')} {o.get('salary_type','')}" for o in overrides[eid])

        vals = [
            emp.get('name', ''),
            emp.get('department', ''),
            type_map.get(emp.get('default_type',''), emp.get('default_type','')),
            emp.get('nida_number','') or '', emp.get('nssf_number','') or '', emp.get('tin_number','') or '',
            int(emp.get('day_rate', 0) or 0),
            int(emp.get('monthly_salary', 0) or 0),
            int(emp.get('advance_total', 0) or 0),
            note,
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(i, ci, v); c.border = tb
            c.alignment = Alignment(horizontal='left' if ci in (1,2,3,4,5,6,10) else 'right')
            if 7 <= ci <= 9: c.number_format = '#,##0'

    for i, w in enumerate([16, 22, 12, 24, 16, 14, 16, 16, 16, 30], 1):
        ws.column_dimensions[chr(64+i)].width = w

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    fname = f'ENPRIZON_LINDI_Employees_{eff_month}.xlsx' if eff_month else 'ENPRIZON_LINDI_Employees.xlsx'
    return send_file(buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True, download_name=fname)

# ═══════════════════════════════════════════════════════════
#  API: 导出出勤表
# ═══════════════════════════════════════════════════════════

@app.route('/export/attendance', methods=['GET'])
@login_required
@require_permission('attendance', 'export')
def export_attendance():
    """导出出勤网格为 Excel，含状态颜色标记 — P3 归一：数据层委托 _build_attendance_grid 唯一真源"""
    month = resolve_month(request)
    md_wrap = _get_month_data(month)
    if EXPORT_STRICT_MONTH:
        stamp = _month_stamp()
        if stamp and month != stamp:
            return jsonify({'ok': False, 'error': 'month_mismatch', 'expected': stamp}), 409
    eff_month = month
    if md_wrap is not None:
        _grid = _build_attendance_grid(md_wrap.get('main_data'), md_wrap.get('employees'))
    else:
        _grid = _build_attendance_grid()
    all_dates = _grid['dates']
    # Adapt grid rows to legacy shape (name/type/days) for existing Excel coloring logic
    type_labels = {'piece_crush': 'Crush Piece', 'piece_underground': 'Underground Piece Rate', 'piece_driller': 'Driller Piece',
                   'day_rate': 'Day Rate', 'monthly': 'Monthly', 'advance_only': 'Advance Only', 'address_book': 'Address Book'}
    rows = []
    for r in _grid['rows']:
        # _grid already resolves monthly default (P) vs '(P)'; legacy export distinguished is_monthly check
        # Preserve downstream `is_monthly` coloring by keeping original days; no recompute
        # r['type'] already via type_labels; keep for coloring trigger
        rows.append({
            'name': r.get('name', ''),
            'type': r.get('type', ''),
            'days': r.get('days', {}),
        })
    if not rows:
        return jsonify({'ok': False, 'error': '无出勤数据'})

    # ── 生成 Excel ──
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Attendance'

    hfont = Font(bold=True, color='FFFFFF', size=11)
    hfill = PatternFill('solid', fgColor='185FA5')
    ha = Alignment(horizontal='center', vertical='center', wrap_text=True)
    tb = Border(left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))

    # 状态 → 颜色（与 UI 保持一致）
    fill_map = {
        'D':  PatternFill('solid', fgColor='3B82F6'),   # Day Shift
        'N':  PatternFill('solid', fgColor='06B6D4'),   # Night Shift
        'B':  PatternFill('solid', fgColor='8B5CF6'),   # Both (Day+Night)
        'P':  PatternFill('solid', fgColor='10B981'),   # Present
        'A':  PatternFill('solid', fgColor='EF4444'),   # Absent
        'L':  PatternFill('solid', fgColor='F59E0B'),   # Leave
        'R':  PatternFill('solid', fgColor='14B8A6'),   # Driller
        'C':  PatternFill('solid', fgColor='F97316'),   # Crush
        '(P)': PatternFill('solid', fgColor='9CA3AF'),  # Monthly Default
    }
    text_color = Font(color='FFFFFF', bold=True)

    # ── Headers ──
    headers = ['Name', 'Type'] + all_dates
    for ci, h in enumerate(headers, 1):
        c = ws.cell(1, ci, h); c.font = hfont; c.fill = hfill; c.alignment = ha; c.border = tb

    # ── 数据行 ──
    for ri, row in enumerate(rows, 2):
        ws.cell(ri, 1, row['name']).border = tb
        c_type = ws.cell(ri, 2, row['type']); c_type.border = tb
        c_type.alignment = Alignment(horizontal='center')

        for di, dt in enumerate(all_dates):
            status = row['days'].get(dt, '')
            c = ws.cell(ri, 3 + di, status)
            c.border = tb
            c.alignment = Alignment(horizontal='center', vertical='center')
            # 状态着色
            sf = fill_map.get(status)
            if sf:
                c.fill = sf
                c.font = text_color
            elif status == '':
                c.fill = PatternFill('solid', fgColor='F3F4F6')

    # ── 列宽 ──
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 12
    ws.freeze_panes = 'C2'

    # 日期列宽度
    for di in range(len(all_dates)):
        col_letter = chr(67 + di) if di < 24 else ''
        if col_letter:
            ws.column_dimensions[col_letter].width = 7

    # ── 添加图例 sheet ──
    ws2 = wb.create_sheet('Legend')
    legend = [
        ('D', 'Day Shift', '3B82F6'),
        ('N', 'Night Shift', '06B6D4'),
        ('B', 'Both (Day+Night)', '8B5CF6'),
        ('P', 'Present', '10B981'),
        ('A', 'Absent', 'EF4444'),
        ('L', 'Leave', 'F59E0B'),
        ('R', 'Driller', '14B8A6'),
        ('C', 'Crush', 'F97316'),
        ('(P)', 'Monthly Default', '9CA3AF'),
    ]
    ws2.cell(1, 1, 'Code').font = Font(bold=True)
    ws2.cell(1, 2, 'Meaning').font = Font(bold=True)
    ws2.column_dimensions['A'].width = 10
    ws2.column_dimensions['B'].width = 26
    for i, (code, meaning, color) in enumerate(legend, 2):
        c1 = ws2.cell(i, 1, code)
        c1.fill = PatternFill('solid', fgColor=color)
        c1.font = Font(color='FFFFFF', bold=True)
        c1.alignment = Alignment(horizontal='center')
        ws2.cell(i, 2, meaning)

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    fname = f'ENPRIZON_LINDI_Attendance_{eff_month}.xlsx' if eff_month else 'ENPRIZON_LINDI_Attendance.xlsx'
    return send_file(buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True, download_name=fname)


@app.route('/export/attendance-report', methods=['GET'])
@login_required
@require_permission('attendance', 'export')
def export_attendance_report():
    """导出出勤数据报表（分部门横版）：每个部门一个 Sheet，首行标题=部门，列宽自适应，便于打印"""
    month = resolve_month(request)
    md_wrap = _get_month_data(month)
    if EXPORT_STRICT_MONTH:
        stamp = _month_stamp()
        if stamp and month != stamp:
            return jsonify({'ok': False, 'error': 'month_mismatch', 'expected': stamp}), 409
    eff_month = month
    if md_wrap is not None:
        _md_rep = md_wrap.get('main_data')
        _emps_rep = md_wrap.get('employees')
        data = _build_attendance_grid(_md_rep, _emps_rep)
    else:
        data = _build_attendance_grid()
    dates = data['dates']
    rows = data['rows']
    if not rows:
        return jsonify({'ok': False, 'error': '无出勤数据'})
    from core.atten_report import build_attendance_report
    wb = build_attendance_report(dates, rows, eff_month)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f'ENPRIZON_LINDI_Attendance_Report_{eff_month}.xlsx' if eff_month else 'ENPRIZON_LINDI_Attendance_Report.xlsx'
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=fname)

# ═══════════════════════════════════════════════════════════
#  API: 统一导出（所有报表合并为一个文件）
# ═══════════════════════════════════════════════════════════

@app.route('/export/all', methods=['POST'])
@login_required
@require_permission('salary', 'export')
def export_all():
    """一次性导出：员工信息 → 薪资总表 → 出勤表 → 日工资分布 → 产量汇总"""
    try:
        month = resolve_month(request)
        md_wrap = _get_month_data(month)
        if EXPORT_STRICT_MONTH:
            stamp = _month_stamp()
            if stamp and month != stamp:
                return jsonify({'ok': False, 'error': 'month_mismatch', 'expected': stamp}), 409
        eff_month = month
        eff_result = (md_wrap.get('salary_result') if md_wrap else None) if md_wrap is not None else None
        eff_md = (md_wrap.get('main_data') if md_wrap else None) if md_wrap is not None else None
        return _do_export_all(eff_month=eff_month, eff_result=eff_result, eff_md=eff_md)
    except Exception as e:
        import traceback, sys
        print(f'[EXPORT ERROR] {e}', file=sys.stderr, flush=True)
        traceback.print_exc(file=sys.stderr)
        return jsonify({'error': str(e), 'ok': False}), 500


def _do_export_all(eff_month=None, eff_result=None, eff_md=None):
    """导出逻辑体，方便包装错误处理"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from datetime import datetime

    # ── 公共样式 ──
    hfont = Font(bold=True, color='FFFFFF', size=11)
    hfill = PatternFill('solid', fgColor='185FA5')
    ha = Alignment(horizontal='center', vertical='center', wrap_text=True)
    tb = Border(left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))
    total_fill = PatternFill('solid', fgColor='FFF3CD')
    type_map = {'piece_underground':'Underground Piece Rate','piece_driller':'Driller Piece',
                'day_rate':'Day Rate','monthly':'Monthly','both':'Unspecified','advance_only':'Advance Only','address_book':'Address Book'}

    wb = openpyxl.Workbook()
    # 删除默认空 sheet
    wb.remove(wb.active)

    # ── 辅助：解析日期字符串为 datetime ──
    def parse_dt(s):
        try: return datetime.strptime(str(s)[:10], '%Y-%m-%d')
        except: return s

    # ═══════════════════════════════════════════════════════
    #  Sheet 1: 员工信息
    # ═══════════════════════════════════════════════════════
    _eff_month = eff_month if eff_month is not None else (getattr(g, 'view_month', '') or resolve_month(request) if request else '')
    _md_all = _get_month_data(_eff_month) if _eff_month else None
    employees = (_md_all.get('employees') if _md_all else []) if _md_all is not None else []
    if eff_result is not None:
        _eff_result = eff_result
        _eff_md = eff_md if eff_md is not None else (_md_all.get('main_data', {}) if _md_all else {})
    else:
        _eff_result = (_md_all.get('salary_result') if _md_all else None) if _md_all is not None else None
        _eff_md = (_md_all.get('main_data', {}) if _md_all else {}) if _md_all is not None else {}
    if employees:
        from core.exceptions import load_overrides
        overrides = load_overrides(app.config['DATA_FOLDER'], month=_eff_month)
        ws1 = wb.create_sheet('Employee Info')
        headers1 = ['Name', 'Department', 'Type', 'NIDA Number', 'NSSF Number', 'TIN Number', 'Day Rate(TZS)', 'Monthly Base(TZS)', 'Advance(TZS)', 'Notes']
        for ci, h in enumerate(headers1, 1):
            c = ws1.cell(1, ci, h); c.font = hfont; c.fill = hfill; c.alignment = ha; c.border = tb
        for i, emp in enumerate(employees, 2):
            eid = emp['id']
            note = emp.get('_note', '')
            if not note and eid in overrides:
                note = '; '.join(f"{o.get('start_date','')}~{o.get('end_date','')} {o.get('salary_type','')}" for o in overrides[eid])
            vals = [
                emp.get('name',''), emp.get('department',''),
                type_map.get(emp.get('default_type',''), emp.get('default_type','')),
                emp.get('nida_number','') or '', emp.get('nssf_number','') or '', emp.get('tin_number','') or '',
                int(emp.get('day_rate',0) or 0), int(emp.get('monthly_salary',0) or 0),
                int(emp.get('advance_total',0) or 0), note,
            ]
            for ci, v in enumerate(vals, 1):
                c = ws1.cell(i, ci, v); c.border = tb
                c.alignment = Alignment(horizontal='left' if ci in (1,2,3,4,5,6,10) else 'right')
                if 7 <= ci <= 9: c.number_format = '#,##0'
        for i, w in enumerate([18, 22, 12, 24, 16, 14, 16, 16, 16, 30], 1):
            ws1.column_dimensions[chr(64+i)].width = w
        ws1.freeze_panes = 'A2'

    # ═══════════════════════════════════════════════════════
    #  Sheet 2: 薪资总表
    # ═══════════════════════════════════════════════════════
    result = _eff_result
    if result:
        ws2 = wb.create_sheet('Salary Summary')
        headers2 = ['Name', 'Type', 'Underground Piece Rate(TZS)', 'Driller Piece(TZS)', 'Crush Piece(TZS)',
                    'Day Rate(TZS)', 'Monthly(TZS)', 'Overtime(TZS)', 'Bonus(TZS)', 'Driver Allowance(TZS)',
                    'Gross Total(TZS)', 'NSSF(TZS)', 'PAYE(TZS)', 'Company PAYE (50%)(TZS)',
                    'Penalty(TZS)', 'Advance Deduction(TZS)', 'Net Salary(TZS)']
        for ci, h in enumerate(headers2, 1):
            c = ws2.cell(1, ci, h); c.font = hfont; c.fill = hfill; c.alignment = ha; c.border = tb

        _type_map2 = {'piece_crush':'Crush Piece','piece_underground':'Underground Piece Rate','piece_driller':'Driller Piece',
                      'day_rate':'Day Rate','monthly':'Monthly','both':'Unspecified','advance_only':'Advance Only','address_book':'Address Book'}
        for i, emp in enumerate(result['employees'], 2):
            gross = emp.get('gross', 0) or 0
            bonus = int(emp.get('bonus', 0) or 0)
            penalty = int(emp.get('penalty', 0) or 0)
            nssf = emp.get('nssf', 0) or 0
            driver = int(emp.get('driver_allowance', 0) or 0)
            overtime = int(emp.get('overtime', 0) or 0)
            paye = int(emp.get('paye', 0) or 0)
            paye_half = int(emp.get('paye_half', 0) or 0)
            vals = [
                emp.get('name','') or '', _type_map2.get(emp.get('salary_type',''), emp.get('salary_type','')),
                int(emp.get('piece_underground',0) or 0), int(emp.get('piece_driller',0) or 0),
                int(emp.get('piece_crush',0) or 0),
                int(emp.get('day_rate',0) or 0), int(emp.get('monthly',0) or 0),
                overtime, bonus, driver,
                int(gross), int(nssf),
                paye, paye_half, penalty, int(emp.get('advance',0) or 0),
                int(emp.get('net', 0) or 0),
            ]
            for ci, v in enumerate(vals, 1):
                c = ws2.cell(i, ci, v); c.border = tb
                c.alignment = Alignment(horizontal='left' if ci == 1 else 'right')
                if ci > 1: c.number_format = '#,##0'

        tr = len(result['employees']) + 2
        ws2.cell(tr, 1, 'Total').font = Font(bold=True, size=11)
        ws2.cell(tr, 1).fill = total_fill; ws2.cell(tr, 1).border = tb
        for ci in [3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16]:
            lt = chr(64 + ci)
            c = ws2.cell(tr, ci, f'=SUM({lt}2:{lt}{tr-1})')
            c.font = Font(bold=True); c.fill = total_fill; c.border = tb
            c.number_format = '#,##0'
        ws2.cell(tr, 17, f'=SUM(Q2:Q{tr-1})')
        ws2.cell(tr, 17).font = Font(bold=True)
        ws2.cell(tr, 17).fill = total_fill; ws2.cell(tr, 17).border = tb
        ws2.cell(tr, 17).number_format = '#,##0'
        for i, w in enumerate([18, 12, 16, 16, 16, 16, 16, 16, 14, 16, 16, 14, 16, 16, 16, 16, 16], 1):
            ws2.column_dimensions[chr(64+i)].width = w
        ws2.freeze_panes = 'A2'

    # ═══════════════════════════════════════════════════════
    #  Sheet 3: 出勤表
    # ═══════════════════════════════════════════════════════
    md = _eff_md
    if md and employees:
        from collections import defaultdict
        from core.namematch import make_employee_id
        shift_prod = md.get('shift_production', [])
        driller_prod = md.get('driller_production', [])
        attendance_data = md.get('attendance', [])
        all_dates = sorted(set(
            list(set(d['date'] for d in shift_prod)) +
            list(set(d['date'] for d in driller_prod)) +
            list(set(d.get('date', '') for d in attendance_data)) +
            list(set(d.get('date', '') for d in (md.get('crush_production') or []))) +
            list(md.get('dates', []))
        ))

        # 收集自动状态
        day_status = defaultdict(dict)
        for d in shift_prod:
            dt = d['date']
            for e in d.get('day_emps', []):
                eid = make_employee_id(e)
                if eid: day_status[eid][dt] = 'D'
            for e in d.get('night_emps', []):
                eid = make_employee_id(e)
                if eid:
                    existing = day_status.get(eid, {}).get(dt, '')
                    day_status[eid][dt] = 'B' if existing == 'D' else 'N'
        for d in driller_prod:
            dt = d['date']
            cap_id = make_employee_id(d['captain'])
            if cap_id and dt not in day_status.get(cap_id, {}):
                day_status[cap_id][dt] = 'R'
            for m in d.get('members', []):
                mid = make_employee_id(m)
                if mid and dt not in day_status.get(mid, {}):
                    day_status[mid][dt] = 'R'
        # 破碎计件出勤
        crush_data = md.get('crush_production', [])
        for d in crush_data:
            dt = d['date']
            for e in d.get('personnel', []):
                eid = make_employee_id(e)
                if eid and dt not in day_status.get(eid, {}):
                    day_status[eid][dt] = 'C'
        for d in attendance_data:
            dt = d['date']
            for e in d.get('normal', []):
                eid = make_employee_id(e)
                if eid and dt not in day_status.get(eid, {}):
                    day_status[eid][dt] = 'P'

        from core.database import load_attendance_overrides
        manual = load_attendance_overrides(app.config['DATA_FOLDER'])

        # 状态 → 颜色
        fill_map = {
            'D': PatternFill('solid', fgColor='3B82F6'),
            'N': PatternFill('solid', fgColor='06B6D4'),
            'B': PatternFill('solid', fgColor='8B5CF6'),
            'P': PatternFill('solid', fgColor='10B981'),
            'A': PatternFill('solid', fgColor='EF4444'),
            'L': PatternFill('solid', fgColor='F59E0B'),
            'R': PatternFill('solid', fgColor='14B8A6'),
            'C': PatternFill('solid', fgColor='F97316'),
            '(P)': PatternFill('solid', fgColor='9CA3AF'),
        }
        white_bold = Font(color='FFFFFF', bold=True)
        grey_fill = PatternFill('solid', fgColor='F3F4F6')
        date_fmt = 'yyyy-mm-dd'

        att_rows = []
        for emp in employees:
            eid = emp.get('id', '')
            emp_type = emp.get('override_type') or emp.get('default_type', '')
            is_monthly = (emp_type == 'monthly')
            is_top_dept_monthly = is_monthly and emp.get('department') == 'ENPRIZON LINDI PROJECT'
            row_days = {}
            for dt in all_dates:
                kid = f"{eid}|{dt}"
                if kid in manual:
                    row_days[dt] = manual[kid]['status']
                elif eid in day_status and dt in day_status[eid]:
                    row_days[dt] = day_status[eid][dt]
                elif is_top_dept_monthly:
                    row_days[dt] = 'P'
                else:
                    row_days[dt] = ''
            att_rows.append({
                'name': emp.get('name', ''),
                'type': type_map.get(emp_type, emp_type),
                'dept': emp.get('department', ''),
                'days': row_days,
            })

        if att_rows:
            ws3 = wb.create_sheet('Attendance')
            ws3.cell(1, 1, 'Name').font = hfont; ws3.cell(1, 1).fill = hfill
            ws3.cell(1, 1).alignment = ha; ws3.cell(1, 1).border = tb
            ws3.cell(1, 2, 'Type').font = hfont; ws3.cell(1, 2).fill = hfill
            ws3.cell(1, 2).alignment = ha; ws3.cell(1, 2).border = tb
            ws3.cell(1, 3, 'Department').font = hfont; ws3.cell(1, 3).fill = hfill
            ws3.cell(1, 3).alignment = ha; ws3.cell(1, 3).border = tb
            for di, dt in enumerate(all_dates):
                c = ws3.cell(1, 4 + di, parse_dt(dt))
                c.font = hfont; c.fill = hfill; c.alignment = ha; c.border = tb
                c.number_format = date_fmt

            for ri, row in enumerate(att_rows, 2):
                ws3.cell(ri, 1, row['name']).border = tb
                c_type = ws3.cell(ri, 2, row['type']); c_type.border = tb
                c_type.alignment = Alignment(horizontal='center')
                ws3.cell(ri, 3, row['dept']).border = tb
                for di, dt in enumerate(all_dates):
                    status = row['days'].get(dt, '')
                    c = ws3.cell(ri, 4 + di, status)
                    c.border = tb
                    c.alignment = Alignment(horizontal='center', vertical='center')
                    sf = fill_map.get(status)
                    if sf:
                        c.fill = sf; c.font = white_bold
                    elif status == '':
                        c.fill = grey_fill

            ws3.column_dimensions['A'].width = 22
            ws3.column_dimensions['B'].width = 12
            ws3.column_dimensions['C'].width = 22
            ws3.freeze_panes = 'D2'

    # ═══════════════════════════════════════════════════════
    #  Sheet 4: 日工资分布
    # ═══════════════════════════════════════════════════════
    if _eff_md and employees:
        from core.calculator import compute_daily_breakdown
        from core.exceptions import load_overrides as _ld_ov, load_daily_exclusions
        _ug_exp = _build_ug_team_members(app.config['DATA_FOLDER'])
        try:
            dw_result = compute_daily_breakdown(
                main_data=_eff_md, employees=employees,
                overrides=_ld_ov(app.config['DATA_FOLDER'], month=_eff_month),
                exclusions=load_daily_exclusions(app.config['DATA_FOLDER']),
                pricing=(_md_all.get('config_snapshot', {}) if _md_all else {}) if _md_all is not None else {},
                data_folder=app.config['DATA_FOLDER'],
                ug_team_members=_ug_exp,
            )
        except TypeError:
            dw_result = compute_daily_breakdown(
                main_data=_eff_md, employees=employees,
                overrides=_ld_ov(app.config['DATA_FOLDER'], month=_eff_month),
                exclusions=load_daily_exclusions(app.config['DATA_FOLDER']),
                pricing=(_md_all.get('config_snapshot', {}) if _md_all else {}) if _md_all is not None else {},
                data_folder=app.config['DATA_FOLDER'],
            )
        # 合并 att_override_dates
        import sqlite3 as _sq, os as _os
        att_ov_map = {}
        db_path = _os.path.join(app.config['DATA_FOLDER'], 'kilwa.db')
        if _os.path.exists(db_path):
            conn = _sq.connect(db_path)
            try:
                for r in conn.execute("SELECT employee_id, date FROM attendance_overrides").fetchall():
                    att_ov_map[f"{r[0]}|{r[1]}"] = True
            except: pass
            conn.close()
        for eid, e in dw_result.items():
            e['att_override_dates'] = [dt for dt in e.get('daily', {}) if f"{eid}|{dt}" in att_ov_map]

        # 收集所有日期
        dw_all_dates = set()
        for e in dw_result.values():
            dw_all_dates.update(e.get('daily', {}).keys())
        dw_dates = sorted(dw_all_dates)

        if dw_dates and dw_result:
            ws4 = wb.create_sheet('Daily Wages')
            ws4.cell(1, 1, 'Name').font = hfont; ws4.cell(1, 1).fill = hfill
            ws4.cell(1, 1).alignment = ha; ws4.cell(1, 1).border = tb
            ws4.cell(1, 2, 'Type').font = hfont; ws4.cell(1, 2).fill = hfill
            ws4.cell(1, 2).alignment = ha; ws4.cell(1, 2).border = tb
            for di, dt in enumerate(dw_dates):
                c = ws4.cell(1, 3 + di, parse_dt(dt))
                c.font = hfont; c.fill = hfill; c.alignment = ha; c.border = tb
                c.number_format = date_fmt
            # 合计列
            total_col = 3 + len(dw_dates)
            c = ws4.cell(1, total_col, 'Total(TZS)')
            c.font = hfont; c.fill = hfill; c.alignment = ha; c.border = tb

            override_fill = PatternFill('solid', fgColor='FFF9C4')  # 黄色标记
            grey_font = Font(color='808080')
            ri = 2
            for emp in employees:
                eid = emp['id']
                e = dw_result.get(eid)
                if not e: continue
                emp_type2 = emp.get('override_type') or emp.get('default_type', '')
                is_monthly2 = (emp_type2 == 'monthly')

                ws4.cell(ri, 1, e['name']).border = tb
                ct = ws4.cell(ri, 2, type_map.get(emp_type2, emp_type2))
                ct.border = tb; ct.alignment = Alignment(horizontal='center')

                for di, dt in enumerate(dw_dates):
                    amt = e.get('daily', {}).get(dt, 0)
                    shift = e.get('daily_shifts', {}).get(dt, '')
                    label = f"{int(amt)} {shift}" if shift and amt > 0 else (int(amt) if amt > 0 else '')
                    c = ws4.cell(ri, 3 + di, label if label else None)
                    c.border = tb
                    c.alignment = Alignment(horizontal='right')
                    if amt > 0: c.number_format = '#,##0'
                    # 覆盖日期黄色标记
                    if dt in e.get('override_dates', []) or dt in e.get('att_override_dates', []):
                        c.fill = override_fill
                    # 月薪灰色
                    if is_monthly2 and dt in e.get('daily', {}) and dt not in e.get('override_dates', []):
                        if not e.get('att_override_dates') or dt not in e.get('att_override_dates', []):
                            c.font = grey_font

                # 合计 = SUM 公式
                col_end = 3 + len(dw_dates) - 1
                col_end_letter = chr(64 + col_end) if col_end <= 26 else ''
                if col_end_letter:
                    c_total = ws4.cell(ri, total_col,
                        f'=SUM(C{ri}:{col_end_letter}{ri})' if col_end >= 3 else None)
                else:
                    c_total = ws4.cell(ri, total_col, e.get('total', 0))
                c_total.border = tb; c_total.alignment = Alignment(horizontal='right')
                c_total.number_format = '#,##0'
                c_total.font = Font(bold=True)
                ri += 1

            # 合计行
            ws4.cell(ri, 1, 'Total').font = Font(bold=True)
            ws4.cell(ri, 1).fill = total_fill; ws4.cell(ri, 1).border = tb
            ws4.cell(ri, 2, '').fill = total_fill; ws4.cell(ri, 2).border = tb
            for di in range(len(dw_dates)):
                col = 3 + di
                c = ws4.cell(ri, col, f'=SUM({chr(64+col)}2:{chr(64+col)}{ri-1})' if col <= 26 else 0)
                c.font = Font(bold=True); c.fill = total_fill; c.border = tb
                c.number_format = '#,##0'
            c = ws4.cell(ri, total_col, f'=SUM({chr(64+total_col)}2:{chr(64+total_col)}{ri-1})' if total_col <= 26 else 0)
            c.font = Font(bold=True); c.fill = total_fill; c.border = tb
            c.number_format = '#,##0'

            ws4.column_dimensions['A'].width = 22
            ws4.column_dimensions['B'].width = 12
            if total_col <= 26:
                ws4.column_dimensions[chr(64+total_col)].width = 14
            ws4.freeze_panes = 'C2'

    # ═══════════════════════════════════════════════════════
    #  Sheet 5: 产量汇总
    # ═══════════════════════════════════════════════════════
    if _eff_md and _eff_md.get('shift_production'):
        ws5 = wb.create_sheet('Production Summary')
        for ci, h in enumerate(['Date', 'NICKEL(H)', 'NICKEL(L)', 'MAWE'], 1):
            c = ws5.cell(1, ci, h); c.font = hfont; c.fill = hfill; c.alignment = ha; c.border = tb
        for i, d in enumerate(_eff_md.get('shift_production', []), 2):
            dp = d.get('day_prod') or {}; np = d.get('night_prod') or {}
            c_dt = ws5.cell(i, 1, parse_dt(d['date']))
            c_dt.number_format = date_fmt; c_dt.border = tb
            ws5.cell(i, 2, (dp.get('NICKEL（H）', 0) or 0) + (np.get('NICKEL（H）', 0) or 0)).border = tb
            ws5.cell(i, 3, (dp.get('NICKEL（L）', 0) or 0) + (np.get('NICKEL（L）', 0) or 0)).border = tb
            ws5.cell(i, 4, (dp.get('MAWE', 0) or 0) + (np.get('MAWE', 0) or 0)).border = tb
        for i, w in enumerate([14, 14, 14, 10], 1):
            ws5.column_dimensions[chr(64+i)].width = w
        ws5.freeze_panes = 'A2'

    # ═══════════════════════════════════════════════════════
    #  Sheet 6: 钻工计件出勤明细
    # ═══════════════════════════════════════════════════════
    if _eff_md and _eff_md.get('driller_production'):
        # ── 队长名规范化（通过员工账号匹配，避免通讯录别名差异）──
        from core.namematch import make_employee_id as _neid

        def _norm_captain(name):
            """标准化队长名称为统一格式，通过 make_employee_id 匹配"""
            import re as _re
            eid = _neid(name)
            # 已知钻工队长 (账号 → 标准显示名)
            _leader_map = {
                _neid('SHEDRACK PINIEL LAIZER'): 'SHEDRACK PINIEL LAIZER',
                _neid('JOHN BOAY BURA'): 'JOHN BOAY BURA',
                _neid('BARAKA LAIZER'): 'BARAKA LAIZER',
                _neid('JOSEPH DONALD'): 'JOSEPH DONALD',
            }
            if eid and eid in _leader_map:
                return _leader_map[eid]
            # 回退：去空格大写匹配
            key = _re.sub(r'\s+', '', _re.sub(r'\s*\([^)]*\)\s*', '', str(name))).upper()
            return key

        from collections import defaultdict
        captain_groups = defaultdict(list)
        for d in _eff_md['driller_production']:
            cap = _norm_captain(d['captain'])
            captain_groups[cap].append(d)

        captain_order = ['SHEDRACK PINIEL LAIZER', 'JOHN BOAY BURA', 'BARAKA LAIZER', 'JOSEPH DONALD']

        if any(captain_groups.get(c) for c in captain_order):
            ws7 = wb.create_sheet('Driller Team Details')
            section_font = Font(bold=True, size=12, color='185FA5')
            subtotal_fill = PatternFill('solid', fgColor='E8F4FD')

            r = 1
            grand_nh = grand_nl = grand_mw = grand_amt = 0

            for cap_name in captain_order:
                records = captain_groups.get(cap_name, [])
                if not records:
                    continue
                records.sort(key=lambda d: d['date'])

                # ── 队长标题行 ──
                ws7.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
                c = ws7.cell(r, 1, f'Captain: {cap_name} ({len(records)} days)')
                c.font = section_font
                r += 1

                # ── 表头 ──
                sub_headers = ['Date', 'NH', 'NL', 'MW', 'Amount(TZS)', 'Headcount', 'Personnel']
                for ci, h in enumerate(sub_headers, 1):
                    c = ws7.cell(r, ci, h)
                    c.font = hfont; c.fill = hfill; c.alignment = ha; c.border = tb
                r += 1

                # ── 数据行 ──
                cap_nh = cap_nl = cap_mw = cap_amt = 0
                for rec in records:
                    nh = rec.get('nh', 0) or 0
                    nl = rec.get('nl', 0) or 0
                    mw = rec.get('mw', 0) or 0
                    amt = nh * 5000 + nl * 4000 + mw * 3000
                    # 出勤人员列表（队长 + 成员）
                    member_list = [_norm_captain(rec['captain'])] + (rec.get('members', []) or [])
                    cap_nh += nh; cap_nl += nl; cap_mw += mw; cap_amt += amt

                    ws7.cell(r, 1, parse_dt(rec['date'])).border = tb
                    ws7.cell(r, 2, nh).border = tb
                    ws7.cell(r, 3, nl).border = tb
                    ws7.cell(r, 4, mw).border = tb
                    c = ws7.cell(r, 5, amt); c.border = tb; c.number_format = '#,##0'
                    ws7.cell(r, 6, len(member_list)).border = tb
                    ws7.cell(r, 7, ', '.join(member_list)).border = tb
                    r += 1

                # ── 小计行 ──
                ws7.cell(r, 1, f'{cap_name.split()[0]} Subtotal').font = Font(bold=True)
                for ci in [1, 2, 3, 4, 5, 6, 7]:
                    ws7.cell(r, ci).fill = subtotal_fill; ws7.cell(r, ci).border = tb
                ws7.cell(r, 2, cap_nh).number_format = '#,##0'
                ws7.cell(r, 3, cap_nl).number_format = '#,##0'
                ws7.cell(r, 4, cap_mw).number_format = '#,##0'
                ws7.cell(r, 5, cap_amt).number_format = '#,##0'
                grand_nh += cap_nh; grand_nl += cap_nl; grand_mw += cap_mw; grand_amt += cap_amt
                r += 2  # 空行分隔

            # ── 总计行 ──
            ws7.cell(r, 1, 'Grand Total').font = Font(bold=True, size=11)
            for ci in [1, 2, 3, 4, 5, 6, 7]:
                ws7.cell(r, ci).fill = total_fill; ws7.cell(r, ci).border = tb
            ws7.cell(r, 2, grand_nh).number_format = '#,##0'
            ws7.cell(r, 3, grand_nl).number_format = '#,##0'
            ws7.cell(r, 4, grand_mw).number_format = '#,##0'
            ws7.cell(r, 5, grand_amt).number_format = '#,##0'

            # ── 列宽 ──
            ws7.column_dimensions['A'].width = 14
            ws7.column_dimensions['B'].width = 10
            ws7.column_dimensions['C'].width = 10
            ws7.column_dimensions['D'].width = 10
            ws7.column_dimensions['E'].width = 18
            ws7.column_dimensions['F'].width = 10
            ws7.column_dimensions['G'].width = 60
            ws7.freeze_panes = 'A2'

    # ── 文件名 ──
    fname = f'ENPRIZON_LINDI_{_eff_month}.xlsx' if _eff_month else 'ENPRIZON_LINDI_Report.xlsx'
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True, download_name=fname)


# ═══════════════════════════════════════════════════════════
#  PDF 工资单导出
# ═══════════════════════════════════════════════════════════

def _build_slip_data(emp, salary_result, employees_db, month):
    from core.database import get_employee_profile
    profile = get_employee_profile(app.config['DATA_FOLDER'], emp.get('employee_id', ''))
    
    gross = (emp.get('piece_underground', 0) or 0) + (emp.get('piece_driller', 0) or 0) + \
            (emp.get('piece_crush', 0) or 0) + (emp.get('day_rate', 0) or 0) + (emp.get('monthly', 0) or 0) + \
            (emp.get('overtime', 0) or 0)
    
    type_labels = {
        'piece_underground': 'Underground', 'piece_driller': 'Driller',
        'piece_crush': 'Crush', 'day_rate': 'Day Rate', 'monthly': 'Monthly'
    }
    
    return {
        'employee_id': emp.get('employee_id', ''),
        'name': emp.get('name', ''),
        'department': profile.get('department', '') if profile else '',
        'position': profile.get('position', '') if profile else '',
        'nssf_number': profile.get('nssf_number', '') if profile else '',
        'tin_number': profile.get('tin_number', '') if profile else '',
        'nssf_enrolled': profile.get('nssf_enrolled', False) if profile else False,
        'month': month,
        'salary_type_label': type_labels.get(emp.get('salary_type', ''), emp.get('salary_type', '')),
        'piece_underground': int(emp.get('piece_underground', 0) or 0),
        'piece_driller': int(emp.get('piece_driller', 0) or 0),
        'piece_crush': int(emp.get('piece_crush', 0) or 0),
        'day_rate': int(emp.get('day_rate', 0) or 0),
        'monthly': int(emp.get('monthly', 0) or 0),
        'overtime': int(emp.get('overtime', 0) or 0),
        'bonus': int(emp.get('bonus', 0) or 0),
        'driver_allowance': int(emp.get('driver_allowance', 0) or 0),
        'gross': int(gross),
        'nssf': int(emp.get('nssf', 0) or 0),
        'paye': int(emp.get('paye', 0) or 0),
        'paye_half': int((emp.get('paye', 0) or 0) // 2),  # 公司代付一半
        'advance': int(emp.get('advance', 0) or 0),
        'penalty': int(emp.get('penalty', 0) or 0),
        'nssf_rate': 0.10,
        'total_deductions': int((emp.get('nssf', 0) or 0) + (emp.get('paye', 0) or 0) - (emp.get('paye', 0) or 0) // 2 + \
                                (emp.get('advance', 0) or 0) + (emp.get('penalty', 0) or 0)),
        'taxable_income': int(gross - (emp.get('nssf', 0) or 0)),
        'net': int(emp.get('net', 0) or 0),
        'attendance_days': 0,
    }


# ═══════════════════════════════════════════════════════════
#  PDF 自适应渲染（工资单导出用）
# ═══════════════════════════════════════════════════════════

_PDF_SCALE_LADDER = [1.0, 0.95, 0.9, 0.85, 0.8, 0.75, 0.7, 0.65, 0.6]
_PDF_SCALE_FLOOR = 0.6


def _render_pdf_fit_pages(template_name, max_pages, **ctx):
    """渲染模板并按缩放阶梯重试，直到 PDF 页数 <= max_pages。

    适用于"内容溢出必然增加页数"的模板（如 payslip_single.html：
    .slip 固定为整页高，条目再多也只会溢出到下一页）。
    返回 (pdf_bytes, 使用的缩放比例)。
    """
    from weasyprint import HTML

    doc = None
    used = _PDF_SCALE_FLOOR
    for s in _PDF_SCALE_LADDER:
        ctx['scale'] = s
        doc = HTML(string=render_template(template_name, **ctx)).render()
        used = s
        if len(doc.pages) <= max_pages:
            return doc.write_pdf(), used
    print(f'[PDF ADAPTIVE] {template_name}: 已到缩放下限 {used}，'
          f'页数 {len(doc.pages)} 仍超过上限 {max_pages}', file=sys.stderr, flush=True)
    return doc.write_pdf(), used


def _count_payslip_rows(slip):
    """统计单张工资单的条目行数（用于估算批量导出的初始缩放）。"""
    earn_keys = ('piece_underground', 'piece_driller', 'piece_crush', 'day_rate',
                 'monthly', 'overtime', 'bonus', 'driver_allowance')
    ded_keys = ('paye_half', 'advance', 'penalty')
    earn = sum(1 for k in earn_keys if (slip.get(k) or 0) > 0) + 1        # + GROSS
    ded = 2 + sum(1 for k in ded_keys if (slip.get(k) or 0) > 0) + 1      # NSSF/PAYE + TOTAL
    return earn + ded


def _render_batch_payslips_pdf(slips):
    """批量工资单自适应渲染：每页固定 4 人，且每格内容完整不裁切。

    批量模板的 .page 固定 277mm 高，格子内容溢出只会发生格间重叠而
    不增加页数——页数信号失效，因此用"探针渲染"验证：把每张工资单
    放入与网格单元格等大（90mm × 133.5mm）的独立页面渲染，
    页数 == 张数 ⟺ 全部放得下；否则缩小 scale 重试。
    返回 (pdf_bytes, 使用的缩放比例)。
    """
    from weasyprint import HTML

    if not slips:
        return HTML(string=render_template('payslip.html', slips=slips, scale=1.0)).write_pdf(), 1.0

    # 初始缩放按最满一张的行数估算（单元格内容区高约 127.5mm，
    # 固定版面开销约 86mm，每行约 3.5mm）。估算不必精确，探针会逐级校验。
    worst_rows = max(_count_payslip_rows(s) for s in slips)
    budget, overhead, row_h = 127.5, 86.0, 3.5
    start = min(1.0, round(0.98 * budget / (overhead + row_h * worst_rows), 2))
    start = max(start, _PDF_SCALE_FLOOR)

    s = start
    pages = 0
    while True:
        probe_html = render_template('payslip.html', slips=slips, scale=s, probe=True)
        pages = len(HTML(string=probe_html).render().pages)
        if pages <= len(slips) or s <= _PDF_SCALE_FLOOR:
            break
        s = max(round(s - 0.05, 2), _PDF_SCALE_FLOOR)
    if pages > len(slips):
        print(f'[PDF ADAPTIVE] payslip.html: 缩放下限 {s} 下仍有条目放不下，'
              f'请检查模板尺寸', file=sys.stderr, flush=True)

    final_html = render_template('payslip.html', slips=slips, scale=s)
    return HTML(string=final_html).write_pdf(), s


@app.route('/export/payslip/<employee_id>', methods=['GET'])
@login_required
@require_permission('salary', 'export')
def export_payslip_single(employee_id):
    try:
        import io, os
        from datetime import datetime
        
        month = (request.args.get('month') or '').strip() or resolve_month(request)
        download = request.args.get('download', '0') == '1'
        
        md = _get_month_data(month)
        result = md.get('salary_result') if md else None
        if not md or not result:
            return jsonify({'error': 'No salary data. Please recalculate.', 'ok': False}), 400
        
        emp = None
        for e in result.get('employees', []):
            if str(e.get('employee_id', '')) == str(employee_id):
                emp = e
                break
        
        if not emp:
            return jsonify({'error': f'Employee {employee_id} not found', 'ok': False}), 404
        
        slip_data = _build_slip_data(emp, result, md.get('employees', []), month)
        pdf_bytes, used_scale = _render_pdf_fit_pages('payslip_single.html', max_pages=1, slip=slip_data)
        
        pays_dir = os.path.join(app.config['DATA_FOLDER'], 'payslips')
        os.makedirs(pays_dir, exist_ok=True)
        filename = f'payslip_{employee_id}_{month}_{datetime.now().strftime("%Y%m%d%H%M%S")}.pdf'
        filepath = os.path.join(pays_dir, filename)
        with open(filepath, 'wb') as f:
            f.write(pdf_bytes)
        
        buf = io.BytesIO(pdf_bytes)
        buf.seek(0)
        
        return send_file(buf,
            mimetype='application/pdf',
            as_attachment=download,
            download_name=f'payslip_{employee_id}_{month}.pdf' if download else None)
            
    except Exception as e:
        import traceback, sys
        print(f'[PDF EXPORT ERROR] {e}', file=sys.stderr, flush=True)
        traceback.print_exc(file=sys.stderr)
        return jsonify({'error': str(e), 'ok': False}), 500


@app.route('/export/payslips-all', methods=['POST'])
@login_required
@require_permission('salary', 'export')
def export_payslips_all():
    try:
        import io, os
        from datetime import datetime

        data = request.get_json(silent=True) or {}
        month = (data.get('month') or '').strip() or resolve_month(request)
        department = data.get('department', '')
        
        md = _get_month_data(month)
        result = md.get('salary_result') if md else None
        if not md or not result:
            return jsonify({'error': 'No salary data. Please recalculate.', 'ok': False}), 400
        
        employees = result.get('employees', [])
        if not employees:
            return jsonify({'error': 'No employees found', 'ok': False}), 404
        
        if department:
            emp_dept_map = {e.get('id'): e.get('department', '') for e in md.get('employees', [])}
            employees = [e for e in employees if emp_dept_map.get(e.get('employee_id', e.get('id', ''))) == department]
        
        slips = []
        for emp in employees:
            slip_data = _build_slip_data(emp, result, md.get('employees', []), month)
            if slip_data.get('net', 0) > 0:
                slips.append(slip_data)
        
        pdf_bytes, used_scale = _render_batch_payslips_pdf(slips)
        
        pays_dir = os.path.join(app.config['DATA_FOLDER'], 'payslips')
        os.makedirs(pays_dir, exist_ok=True)
        dept_suffix = f'_{department}' if department else ''
        # 部门名含 "/"（如 Logistics/Ground production、Sort Crush/Crush Piece Rate）
        # 会被 os.path.join 当作路径分隔符，导致写入不存在的子目录而 FileNotFoundError，
        # 故输出文件名/下载名中一律用 "_" 替换非法路径字符。
        safe_dept_suffix = dept_suffix.replace('/', '_').replace('\\', '_')
        filename = f'payslips{safe_dept_suffix}_{month}_{datetime.now().strftime("%Y%m%d%H%M%S")}.pdf'
        filepath = os.path.join(pays_dir, filename)
        with open(filepath, 'wb') as f:
            f.write(pdf_bytes)
        
        buf = io.BytesIO(pdf_bytes)
        buf.seek(0)
        
        return send_file(buf,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'payslips{safe_dept_suffix}_{month}.pdf')
            
    except Exception as e:
        import traceback, sys
        print(f'[PDF BATCH EXPORT ERROR] {e}', file=sys.stderr, flush=True)
        traceback.print_exc(file=sys.stderr)
        return jsonify({'error': str(e), 'ok': False}), 500


# ═══════════════════════════════════════════════════════════
#  自动加载源文件
# ═══════════════════════════════════════════════════════════

def find_free_port(start=8080, max_try=100):
    for port in range(start, start + max_try):
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
            if s.connect_ex(('localhost', port)) != 0:
                return port
    return start

def _ensure_viewer_account():
    """确保默认账号存在 + KEJU 为 super_admin"""
    from core.database import get_conn, _hash_password
    conn = get_conn(app.config['DATA_FOLDER'])

    # user 账号（viewer 角色）
    existing = conn.execute("SELECT username FROM admin_users WHERE username='user'").fetchone()
    if not existing:
        pwd_hash = _hash_password('qweasd')
        conn.execute(
            "INSERT INTO admin_users (username, password_hash, role) VALUES (?, ?, ?)",
            ('user', pwd_hash, 'viewer')
        )
        conn.commit()
        print('  ✓ 已创建 viewer 账号 (user / qweasd)')
    else:
        # 修正已存在的 user 角色为 viewer
        conn.execute("UPDATE admin_users SET role='viewer' WHERE username='user' AND role!='viewer'")
        conn.commit()

    # KEJU 升级为 super_admin
    conn.execute("UPDATE admin_users SET role='super_admin' WHERE username='KEJU' AND role!='super_admin'")
    conn.commit()

    conn.close()

# ── P5: 用 Excel 种子新表 ──────────

def _backup_to_archive():
    """首次启动时自动备份 kilwa.db → archived_kilwa.db（如果归档不存在）"""
    import shutil
    db_path = os.path.join(app.config['DATA_FOLDER'], 'kilwa.db')
    archive_path = os.path.join(app.config['DATA_FOLDER'], 'archived_kilwa.db')
    if os.path.exists(db_path) and not os.path.exists(archive_path):
        shutil.copy2(db_path, archive_path)
        print('  ✓ 已自动备份 kilwa.db → archived_kilwa.db')

def seed_new_tables_from_excel():
    """仅首次运行：当 employees 表为空时，从通讯录 Excel 种子 employees"""
    from core.database import get_conn
    conn = get_conn(app.config['DATA_FOLDER'])
    # 检查 employees 是否已有数据（防止重复种子）
    count = conn.execute("SELECT COUNT(*) FROM employees").fetchone()[0]
    if count > 10:
        conn.close()
        return False  # 已有数据，跳过

    print('  ► 从通讯录种子 employees 表...')
    # 纯采集模式：找 data/source 下可用的通讯录文件（仅首次导入用）
    import glob
    ab_path = None
    for _p in sorted(glob.glob(os.path.join(SOURCE_DIR, '*.xlsx'))):
        try:
            import openpyxl
            _wb = openpyxl.load_workbook(_p, data_only=True, read_only=True)
            _sheets = _wb.sheetnames
            _wb.close()
            if any('成员列表' in s or '通讯录' in s for s in _sheets):
                ab_path = _p
                break
        except Exception:
            continue
    if not ab_path:
        conn.close()
        return False

    from core.addressbook import parse_address_book
    from core.database import create_event
    ab_data = parse_address_book(ab_path)
    if not ab_data:
        conn.close()
        return False

    seen = set()
    for eid, info in ab_data.items():
        if not eid or eid in seen or eid in HARD_EXCLUDE_IDS:
            continue
        seen.add(eid)
        dept = strip_dept(info.get('department', ''))
        dtype = info.get('guessed_type') or 'day_rate'
        try:
            conn.execute("""
                INSERT OR IGNORE INTO employees (id, name, department, default_type, day_rate, monthly_salary)
                VALUES (?,?,?,?,?,?)
            """, (eid, info['name'], dept, dtype, 0, 0))
            create_event(app.config['DATA_FOLDER'], {
                'employee_id': eid, 'event_type': 'hire',
                'effective_date': '2026-01-01',
                'snapshot': json.dumps({'name': info['name'], 'department': dept, 'type': dtype}),
                'payload': '{}', 'operator_id': 'system', 'status': 'approved',
            })
        except Exception:
            pass

    print(f'  ✓ 已从通讯录种子 {len(seen)} 名员工 + hire 事件')
    conn.commit()
    conn.close()
    return True

def auto_load_source():
    """纯采集模式：从数据库重建并加载当前月份数据"""
    from datetime import datetime
    from core.database import accrue_comp_leave_monthly
    try:
        r = accrue_comp_leave_monthly(app.config['DATA_FOLDER'])
        if r:
            print(f"  ✓ P28 调休月度入账 {','.join(r['accrued_months'])}（{r['employees']} 人）")
    except Exception as e:
        print(f'  ⚠ P28 调休月度入账失败: {e}')
    current_month = datetime.now(EAT).strftime('%Y-%m')
    chosen_month = current_month

    md = _run_pipeline(month_filter=chosen_month)
    if md is not None:
        print(f'  已加载 {len(md.get("employees", []))} 名员工，应发 {md.get("salary_result", {}).get("total_gross", 0):,} TZS')
        ok = True
    else:
        print('  员工表为空，请先导入通讯录或员工数据')
        ok = False
    _ensure_viewer_account()
    return ok

# ── gunicorn 启动时自动加载数据（python app.py 时跳过，由 __main__ 处理）──
_app_initialized = False

def _gunicorn_init():
    global _app_initialized
    if _app_initialized:
        return
    _app_initialized = True
    from core.database import init_db
    init_db(app.config['DATA_FOLDER'])
    from core.database import init_default_permissions, sync_role_permissions, seed_default_forms
    init_default_permissions(app.config['DATA_FOLDER'])
    sync_role_permissions(app.config['DATA_FOLDER'])
    from core.database import _migrate_permissions_v2
    _migrate_permissions_v2(app.config['DATA_FOLDER'])  # P29: V2 权限目录一次性幂等迁移(播种后执行)
    seed_default_forms(app.config['DATA_FOLDER'])
    _backup_to_archive()
    seed_new_tables_from_excel()
    loaded = auto_load_source()
    if loaded:
        print('  ✓ 源数据已自动加载')
    else:
        print('  ⚠ data/source/ 下缺少主文件')

if __name__ != '__main__':
    _gunicorn_init()

if __name__ == '__main__':
    from core.database import init_db, init_default_permissions, sync_role_permissions, seed_default_forms
    _app_initialized = True
    init_db(app.config['DATA_FOLDER'])
    init_default_permissions(app.config['DATA_FOLDER'])
    sync_role_permissions(app.config['DATA_FOLDER'])
    from core.database import _migrate_permissions_v2
    _migrate_permissions_v2(app.config['DATA_FOLDER'])  # P29: V2 权限目录一次性幂等迁移(播种后执行)
    seed_default_forms(app.config['DATA_FOLDER'])
    _backup_to_archive()
    seed_new_tables_from_excel()

    port = find_free_port(8080)
    print('=' * 50)
    print('  ENPRIZON LINDI PROJECT')
    print(f'  启动地址: http://localhost:{port}')
    loaded = auto_load_source()
    if not loaded:
        print('  ⚠ data/source/ 下缺少主文件，请放入再重启')
    print('=' * 50)
    app.run(debug=False, host='0.0.0.0', port=port)
