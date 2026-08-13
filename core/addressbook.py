"""
通讯录解析模块
解析 ENPRIZON LINDI PROJECT通讯录.xlsx 和 员工基础信息表.xlsx
"""
import openpyxl
from .namematch import (canonical, make_employee_id, load_address_book_index, strip_alias,
                        display_name, _find_header_row, _build_header_col_map, _norm_header)

# ── 部门映射（用于批量设置薪资类型） ──────────────
DEPT_PATTERNS = {
    'piece_driller': ['钻工'],
    'piece_underground': ['生产组井下', 'Production Team underground'],
    'day_rate': ['后勤', '分拣破碎', '机修组', 'Logistics', 'Sort Crush', 'Mechanic Team',
                 '后勤/生产地面', 'Ground Production'],
}

def guess_pay_type(department):
    """根据部门猜测默认薪资类型"""
    if not department:
        return None
    for ptype, patterns in DEPT_PATTERNS.items():
        for pat in patterns:
            if pat in department:
                return ptype
    return None

def parse_address_book(filepath):
    """
    解析通讯录文件 → {
        employee_id: {
            name: str,
            department: str,
            phone: str,
            guessed_type: str|None
        }
    }
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)

    sheet_name = None
    for name in ['成员列表', 'Sheet1']:
        if name in wb.sheetnames:
            sheet_name = name
            break
    if not sheet_name:
        wb.close()
        return {}

    ws = wb[sheet_name]
    # 从通讯录加载员工索引
    load_address_book_index(filepath)
    book = {}
    header_row = _find_header_row(ws)
    if not header_row:
        wb.close()
        return {}
    cm = _build_header_col_map(ws, header_row)
    # 列定位：表头优先，回退到旧硬编码（姓名1/账号2/别名3/部门5/手机7）
    name_col = cm.get('姓名') or 1
    acct_col = cm.get('账号') or 2
    alias_col = cm.get('别名')
    dept_col = cm.get('部门') or 5
    phone_col = cm.get('手机') or 7
    for row in range(header_row + 1, (ws.max_row or 0) + 1):
        name_raw = ws.cell(row, name_col).value
        if not name_raw:
            continue
        name_str = str(name_raw).strip()
        if not name_str:
            continue
        acct = ws.cell(row, acct_col).value
        alias = ws.cell(row, alias_col).value if alias_col else None
        department = ws.cell(row, dept_col).value
        phone = ws.cell(row, phone_col).value
        # 使用账号作为 key
        acct_str = str(acct).strip() if acct else ''
        if not acct_str:
            continue
        eid = acct_str
        alias_str = str(alias).strip() if alias else ''
        disp = alias_str if alias_str else strip_alias(name_str)
        guessed = guess_pay_type(str(department)) if department else None
        if eid in book:
            if department and not book[eid].get('department'):
                book[eid]['department'] = str(department).strip()
        else:
            book[eid] = {
                'name': disp,
                'department': str(department).strip() if department else '',
                'phone': str(phone).strip() if phone else '',
                'guessed_type': guessed,
            }

    wb.close()
    return book


def parse_basic_info(filepath):
    """
    解析员工基础信息表（可选）→ {
        employee_id: { day_rate: int, monthly_salary: int }
    }
    """
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception:
        return {}

    result = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header = {}
        for col in range(1, ws.max_column + 1):
            v = ws.cell(1, col).value
            if v:
                header[col] = str(v).strip().lower()

        name_col = None
        day_col = None
        month_col = None

        for col, h in header.items():
            if '姓名' in h or 'name' in h:
                name_col = col
            if '日薪' in h or ('day' in h and 'rate' in h):
                day_col = col
            if '月薪' in h or 'month' in h or 'salary' in h:
                month_col = col

        if not name_col:
            continue

        for row in range(2, ws.max_row + 1):
            name_raw = ws.cell(row, name_col).value
            if not name_raw:
                continue
            eid = make_employee_id(str(name_raw))
            if not eid:
                continue
            info = {}
            if day_col:
                v = ws.cell(row, day_col).value
                if v:
                    info['day_rate'] = int(float(v))
            if month_col:
                v = ws.cell(row, month_col).value
                if v:
                    info['monthly_salary'] = int(float(v))
            if info:
                result[eid] = info

    wb.close()
    return result
