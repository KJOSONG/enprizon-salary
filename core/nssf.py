"""
NSSF（社保）管理模块 — SQLite 版
"""
from .database import load_nssf_enrollment as _db_load, save_nssf_enrollment as _db_save
from .namematch import make_employee_id

def parse_sdl_list(filepath):
    """解析 SDL LIST 文件 → { employee_id: { name, tin, basic_salary } }"""
    import openpyxl
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet_name = 'Tanzania Mainland' if 'Tanzania Mainland' in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]
    members = {}
    for row in range(2, ws.max_row + 1):
        name_raw = ws.cell(row, 3).value
        tin = ws.cell(row, 2).value
        basic_sal = ws.cell(row, 7).value
        if not name_raw:
            continue
        eid = make_employee_id(name_raw)
        if not eid:
            continue
        members[eid] = {
            'name': str(name_raw).strip(),
            'tin': str(tin).strip() if tin else '',
            'basic_salary': float(basic_sal) if basic_sal else 0,
        }
    wb.close()
    return members

def load_nssf_enrollment(data_folder):
    return _db_load(data_folder)

def save_nssf_enrollment(data_folder, employee_id, enrolled):
    _db_save(data_folder, employee_id, enrolled)

def calc_nssf_deduction(employee, enrolled, gross):
    eid = employee['id'] if isinstance(employee, dict) else None
    if eid and enrolled.get(eid, {}).get('enrolled', False):
        return round(gross * 0.10)
    return 0
