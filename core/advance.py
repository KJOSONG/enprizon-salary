"""
预支汇总解析
从 预支汇总数据.xlsx → Sheet 预支汇总 读取
"""
import openpyxl
from .namematch import make_employee_id

def parse_advance(filepath, month=None):
    """
    解析预支汇总表 → { employee_id: { name, count, total, dates } }
    按姓名标准化匹配

    month: "2026-06" → 只统计该月的预支（按 E 列日期前缀过滤）
           None → 全部（兼容旧行为）
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)

    # 尝试多个可能的sheet名
    sheet_name = None
    for name in ['预支汇总', 'Sheet1']:
        if name in wb.sheetnames:
            sheet_name = name
            break
    if not sheet_name:
        wb.close()
        return {}

    ws = wb[sheet_name]
    advance_map = {}

    for row in range(2, ws.max_row + 1):
        name_raw = ws.cell(row, 2).value  # B: 姓名
        count = ws.cell(row, 3).value or 0      # C: 笔数
        total = ws.cell(row, 4).value or 0      # D: 总额
        dates_raw = ws.cell(row, 5).value       # E: 日期范围

        if not name_raw:
            continue

        # 跳过合计行
        name_str = str(name_raw).strip()
        if name_str in ('合计', '合计:', 'Total'):
            continue

        eid = make_employee_id(name_raw)
        if not eid:
            continue

        dates = str(dates_raw).split(', ') if dates_raw else []

        # ── 月份过滤 ──
        if month:
            # 只保留属于目标月份的预支日期
            month_dates = [d for d in dates if d.startswith(month)]
            if not month_dates:
                continue  # 该行不属于目标月，跳过
            # 按比例折算当月预支金额（当月笔数 / 该行总笔数 × 该行总额）
            ratio = len(month_dates) / len(dates)
            total = float(total) * ratio if total else 0
            count = len(month_dates)
            dates = month_dates

        # 如果同一个人有多行，合并
        if eid in advance_map:
            advance_map[eid]['count'] += int(count) if count else 0
            advance_map[eid]['total'] += float(total) if total else 0
            advance_map[eid]['dates'].extend(dates)
        else:
            advance_map[eid] = {
                'name': name_raw,
                'count': int(count) if count else 0,
                'total': float(total) if total else 0,
                'dates': dates,
            }

    wb.close()
    return advance_map
