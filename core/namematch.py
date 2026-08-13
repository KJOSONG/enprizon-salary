"""
姓名标准化管线（通讯录驱动）
原始名（含企业微信别名括号）→ 去别名 → 通讯录索引查找 → (账号, 显示名)
"""
import re

# ── 通讯录驱动的员工索引（key: 变体去空格大写 → (账号, 显示名)）──
_AB_INDEX = {}

# ── 通讯录外人员补充索引（离职但有未结工资，需保留计薪；8月导入新增）──
# key: 姓名变体去空格大写 → (新ID, 显示名)
_EXTRA_AB_ENTRIES = {
    'EZRAIBRAHIM':  ('129', 'EZRA IBRAHIM'),
    'PAULOLAIZA':   ('130', 'PAULO LAIZA'),
    'PAULOKISENA':  ('131', 'paulo kisena'),
}

# ── 遗留 CANONICAL 映射（通讯录外人员的手动回退）──
_LEGACY_CANONICAL = {
    'SHEDRACK':                  'SHEDRACK PINIEL LAIZER',
    'SHEDRACKPINIELLAIZER':      'SHEDRACK PINIEL LAIZER',
    'JOHN':                      'JOHN BOAY BURA',
    'JOHNBOAYBURA':              'JOHN BOAY BURA',
    'BARAKALAIZER':              'BARAKA LAIZER',
    'JOSEPH':                    'JOSEPH DONALD',
    'JOSEPHDONALD':              'JOSEPH DONALD',
    'JULIASISAYA':              'JULIAS ISAYA',
    'JOSHUATAJIRI':             'JOSHUA TAJIRI',
    'SHAFIRIYAHAYA':            'SHAFIRI YAHAYA',
    'HERIMAULIDI':              'HERI MAULIDI',
    'HOSEALAIZER':              'HOSEA LAIZER',
    'BONIKIVUYO':               'BONI KIVUYO',
    'RAMAZANISAIDINAMAWALA':    'RAMAZANI SAIDI NAMAWALA',
}

# ── 已知钻工队长（显示名列表）──
DRILLER_LEADER_NAMES = [
    'SHEDRACK PINIEL LAIZER',
    'JOHN BOAY BURA',
    'BARAKA LAIZER',
    'JOSEPH DONALD',
]

# ── 钻工队长账号（load_address_book_index 后自动计算）──
DRILLER_LEADERS = []


def _norm_header(val):
    """规范化表头文字用于匹配（去空格/括号/下划线，大写）"""
    if not val:
        return ''
    return re.sub(r'[\s_（）()]', '', str(val).strip()).upper()


def _find_header_row(ws):
    """扫描第 1 列找表头行（含『姓名』）"""
    for row in range(1, (ws.max_row or 0) + 1):
        v = ws.cell(row, 1).value
        if v and '姓名' in str(v):
            return row
    return None


def _build_header_col_map(ws, header_row):
    """扫描表头行建立 { 规范化表头: 列号 } 映射"""
    col_map = {}
    for col in range(1, (ws.max_column or 0) + 1):
        val = ws.cell(header_row, col).value
        if val:
            col_map[_norm_header(val)] = col
    return col_map


def load_address_book_index(filepath):
    """从通讯录 Excel 加载员工索引，填充 _AB_INDEX。

    列位置通过表头文字识别（兼容新旧格式）：
    - 旧格式: 姓名|账号|别名|职务|部门|性别|手机...
    - 新格式: 姓名|账号|职务|部门|性别|手机...（无别名列）
    """
    global _AB_INDEX, DRILLER_LEADERS
    _AB_INDEX = {}
    import openpyxl
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet = None
    for name in ['成员列表', 'Sheet1']:
        if name in wb.sheetnames:
            sheet = name
            break
    if not sheet:
        wb.close()
        return
    ws = wb[sheet]
    header_row = _find_header_row(ws)
    if not header_row:
        wb.close()
        return
    cm = _build_header_col_map(ws, header_row)
    # 列定位：表头优先，回退到旧硬编码（姓名1/账号2/别名3）
    name_col = cm.get('姓名') or 1
    acct_col = cm.get('账号') or 2
    alias_col = cm.get('别名')
    for row in range(header_row + 1, (ws.max_row or 0) + 1):
        name_raw = ws.cell(row, name_col).value
        acct = ws.cell(row, acct_col).value
        alias = ws.cell(row, alias_col).value if alias_col else None
        if not name_raw:
            continue
        name_str = str(name_raw).strip()
        acct_str = str(acct).strip() if acct else ''
        alias_str = str(alias).strip() if alias else ''
        if not acct_str:
            continue
        display = alias_str if alias_str else strip_alias(name_str)
        # 姓名变体
        key_name = norm_key_static(name_str)
        if key_name:
            _AB_INDEX[key_name] = (acct_str, display)
        # 别名变体
        if alias_str:
            key_alias = norm_key_static(alias_str)
            if key_alias:
                _AB_INDEX[key_alias] = (acct_str, display)
        # 带括号的原始姓名去括号后
        sa = strip_alias(name_str)
        if sa:
            key_sa = re.sub(r'\s+', '', sa).upper()
            if key_sa and key_sa != key_name:
                _AB_INDEX[key_sa] = (acct_str, display)
        # 去最后一个词的键（用于匹配短名如 ADHIRUDIN SIJAE RASHID -> ADHIRUDINSIJAE）
        sa_words = sa.split()
        if len(sa_words) > 1:
            short_name = ' '.join(sa_words[:-1])
            short_key = re.sub(r'\s+', '', short_name).upper()
            if short_key and short_key not in _AB_INDEX:
                _AB_INDEX[short_key] = (acct_str, display)
    wb.close()
    # 合并通讯录外补充索引（离职人员等）
    for _k, _v in _EXTRA_AB_ENTRIES.items():
        _AB_INDEX.setdefault(_k, _v)
    # 自动计算钻工队长账号（修改列表内容而非重新赋值，确保外部 import 可见）
    DRILLER_LEADERS.clear()
    for leader_name in DRILLER_LEADER_NAMES:
        for key, (acct, _) in _AB_INDEX.items():
            leader_key = re.sub(r'\s+', '', leader_name).upper()
            if key == leader_key:
                DRILLER_LEADERS.append(acct)
                break


def _is_na(val):
    if val is None: return True
    if isinstance(val, float) and (val != val): return True
    return False


def strip_alias(name):
    if not name or _is_na(name):
        return ''
    return re.sub(r'\s*\([^)]*\)\s*', '', str(name)).strip()


def norm_key_static(name):
    """去空格+大写（静态版本，不依赖 CANONICAL）"""
    sa = strip_alias(name)
    if not sa:
        return ''
    return re.sub(r'\s+', '', sa).upper()


def _extract_paren_names(name):
    """提取括号内的全名候选（按长度降序，优先最长/最全）"""
    if not name:
        return []
    cands = re.findall(r'\(([^)]+)\)', str(name))
    # 去空白+大写后的 key 列表（去重，保持长度降序）
    seen = set()
    result = []
    for c in sorted(cands, key=len, reverse=True):
        k = re.sub(r'\s+', '', c).upper()
        if k and k not in seen:
            seen.add(k)
            result.append((k, c.strip()))
    return result


def _ab_lookup(key):
    """通讯录索引查找：返回 (账号, 显示名) 或 None"""
    if _AB_INDEX and key in _AB_INDEX:
        return _AB_INDEX[key]
    return None


def canonical(name):
    """标准化姓名：通讯录查找 → 显示名；未匹配时回退到遗留 CANONICAL + strip_alias

    优先尝试去括号短名，未命中再尝试括号内全名（源数据格式『短名(全名)』，
    全名才是新通讯录的标准姓名）。
    """
    if not name or _is_na(name):
        return None
    # 去括号短名
    hit = _ab_lookup(norm_key_static(name))
    if hit:
        return hit[1]
    # 括号内全名（如 ALLY VENANCE(Ally Venasi Matias) → Ally Venasi Matias）
    for pk, raw in _extract_paren_names(name):
        hit = _ab_lookup(pk)
        if hit:
            return hit[1]
    # 遗留 CANONICAL 回退
    key = norm_key_static(name)
    if key in _LEGACY_CANONICAL:
        return _LEGACY_CANONICAL[key]
    # 最终回退：去括号结果
    result = strip_alias(name)
    return result if result else None


def make_employee_id(name):
    """生成唯一员工ID：通讯录 → 账号；未匹配时回退到姓名去空格大写

    匹配顺序：去括号短名 → 括号内全名 → 遗留 CANONICAL → 姓名回退。
    括号内全名处理源数据『ALLY VENANCE(Ally Venasi Matias)』这类变体，
    避免回退成旧格式姓名 ID（破坏『账号=ID』原则）。
    """
    if not name or _is_na(name):
        return None
    key = norm_key_static(name)
    # 遗留 CANONICAL：短名先展开为全名，再查通讯录
    full_name = None
    if key in _LEGACY_CANONICAL:
        full_name = _LEGACY_CANONICAL[key]
        full_key = re.sub(r'\s+', '', full_name).upper()
        if _AB_INDEX and full_key in _AB_INDEX:
            return _AB_INDEX[full_key][0]
    # 通讯录索引直接查找（去括号短名）
    if _AB_INDEX and key in _AB_INDEX:
        return _AB_INDEX[key][0]
    # 括号内全名匹配（源数据『短名(全名)』格式）
    for pk, raw in _extract_paren_names(name):
        if _AB_INDEX and pk in _AB_INDEX:
            return _AB_INDEX[pk][0]
    # 遗留 CANONICAL → 转换为旧格式 ID
    if full_name:
        return re.sub(r'\s+', '', full_name).upper()
    # 最终回退
    c = strip_alias(name)
    return re.sub(r'\s+', '', c).upper() if c else None


def display_name(name):
    """返回该员工的显示名（别名优先，否则去括号的姓名）"""
    if not name or _is_na(name):
        return ''
    key = norm_key_static(name)
    if _AB_INDEX and key in _AB_INDEX:
        return _AB_INDEX[key][1]
    return strip_alias(name) or str(name)


def is_driller_leader(name_or_id):
    """判断是否为已知的钻工队长（支持传入账号或姓名）"""
    eid = make_employee_id(name_or_id) if name_or_id else None
    return eid in DRILLER_LEADERS if eid else False


def split_names(raw_str):
    """拆分名称字符串（逗号/分号分隔）→ 标准化姓名列表（显示名）"""
    if not raw_str or _is_na(raw_str):
        return []
    parts = re.split(r'\s*[;,、\n]\s*', str(raw_str))
    result = []
    for p in parts:
        c = canonical(p.strip())
        if c:
            result.append(c)
    return result


def build_master_list(main_data):
    """
    构建员工主列表
    根据人员来源自动分类
    """
    piece_rate_people = main_data.get('piece_rate_people', {})
    daily_salary_people = main_data.get('daily_salary_people', {})

    driller_people = piece_rate_people.get('driller', set())
    underground_people = piece_rate_people.get('underground', set())

    all_ids = set()
    for name in driller_people: all_ids.add(make_employee_id(name))
    for name in underground_people: all_ids.add(make_employee_id(name))
    for name in daily_salary_people: all_ids.add(make_employee_id(name))

    employees = []
    for eid in sorted(eid for eid in all_ids if eid):
        in_driller = any(make_employee_id(n) == eid for n in driller_people)
        in_underground = any(make_employee_id(n) == eid for n in underground_people)
        in_daily = any(make_employee_id(n) == eid for n in daily_salary_people)

        in_piece = in_driller or in_underground

        if in_piece and not in_daily:
            source = 'piece_rate_sheet'
            default_type = 'piece_driller' if in_driller else 'piece_underground'
        elif in_daily and not in_piece:
            source = 'daily_salary_sheet'
            default_type = 'day_rate'
        else:
            source = 'both'
            default_type = 'both'

        # 取第一个出现的显示名
        name = ''
        if in_driller:
            name = next((display_name(n) for n in driller_people if make_employee_id(n) == eid), '')
        if not name and in_underground:
            name = next((display_name(n) for n in underground_people if make_employee_id(n) == eid), '')
        if not name:
            name = next((display_name(n) for n in daily_salary_people if make_employee_id(n) == eid), eid)

        employees.append({
            'id': eid,
            'name': name,
            'default_type': default_type,
            'source': source,
            'override_type': None,
            'overrides': [],
            'day_rate': 0,
            'monthly_salary': 0,
            'advance_total': 0,
        })

    return employees
