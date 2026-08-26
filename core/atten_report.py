"""出勤分部门横版报表 builder V2（纯 openpyxl，无 Flask 依赖）。

V2 布局：
- 按部门分组（原始 department 字符串排序，组内保持输入顺序），每部门一 sheet。
- ROW1 标题行：f"{department}　{month}"（U+3000 全角空格）或仅 department；合并 A1..max_col，Font(bold,size14) 居中 行高24，无填充无边框。
- ROW2 表头：A2='Name'；B2.. 每日 datetime(y,m,d) number_format='d'，均 bold+居中+thin 边框。
- ROWS 3+ 数据：A 列姓名 plain（thin 边框，默认左对齐）；日单元格 str(days.get(ds,'')) verbatim，bold+居中+thin 边框。Department 列已移除。
- Freeze B3；列宽按内容自适应（标题行不计入）；打印设置 landscape/paperSize9/fitToWidth1/fitToHeight0/fitToPageTrue/print_title_rows='1:2'/margins 0.4/0.5。
"""

import datetime as _dt

import openpyxl
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties


def sanitize_sheet_name(name, used):
    s = name or ""
    s = s.replace("/", "-")
    for ch in ":\\?*[]":
        s = s.replace(ch, "")
    s = s.strip()
    s = s[:31]
    if not s:
        s = "No Department"
    base = s
    candidate = base
    n = 2
    while candidate in used:
        suffix = f" -{n}"
        candidate = base[: 31 - len(suffix)] + suffix
        n += 1
    used.add(candidate)
    return candidate


def build_attendance_report(dates, rows, month=""):
    # Never mutate inputs
    # Shallow copy of dates list; rows are not modified, only read
    thin = Side(style="thin")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
    font_title = Font(bold=True, size=14)
    font_bold = Font(bold=True)
    align_center = Alignment(horizontal="center", vertical="center")
    align_center_no_vert = Alignment(horizontal="center", vertical="center")
    # Header/data centered alignment - vertical center for consistency
    align_header = Alignment(horizontal="center", vertical="center")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # group rows by department preserving input order within each dept
    groups: dict[str, list] = {}
    for r in rows:
        dept = r.get("department", "")
        if dept not in groups:
            groups[dept] = []
        groups[dept].append(r)

    sorted_depts = sorted(groups.keys())

    used_names: set[str] = set()

    for dept in sorted_depts:
        dept_rows = groups[dept]
        sheet_title = sanitize_sheet_name(dept, used_names)
        ws = wb.create_sheet(sheet_title)

        max_col = 1 + len(dates)  # A + dates
        if max_col < 1:
            max_col = 1

        # ROW 1 TITLE
        if month:
            # U+3000 full-width space between dept and month
            title_value = f"{dept}\u3000{month}"
        else:
            title_value = dept
        ws.cell(row=1, column=1, value=title_value)
        ws["A1"].font = font_title
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 24
        # Merge A1 across all used columns
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
        # No fill, no border on title - leave defaults

        # ROW 2 HEADER
        c_a2 = ws.cell(row=2, column=1, value="Name")
        c_a2.font = font_bold
        c_a2.alignment = align_header
        c_a2.border = border_thin

        for i, ds in enumerate(dates):
            col = 2 + i
            cell = ws.cell(row=2, column=col)
            cell.value = _dt.datetime(int(ds[0:4]), int(ds[5:7]), int(ds[8:10]))
            cell.number_format = "d"
            cell.font = font_bold
            cell.alignment = align_header
            cell.border = border_thin

        # ROWS 3+ DATA
        for r_idx, r in enumerate(dept_rows):
            row_num = 3 + r_idx
            c_a = ws.cell(row=row_num, column=1, value=r.get("name", ""))
            c_a.border = border_thin
            # A column: default left align (no explicit center)
            # Leave alignment as None/default (left)
            days = r.get("days", {})
            for i, ds in enumerate(dates):
                col = 2 + i
                val = str(days.get(ds, ""))
                cell = ws.cell(row=row_num, column=col, value=val)
                cell.font = font_bold
                cell.alignment = align_header
                cell.border = border_thin

        # AUTO-FIT WIDTHS (content-based, computed per sheet, title excluded)
        # Column A width = max(len(str(name)) for sheet rows, len('Name')) +2
        max_name_len = len("Name")
        for r in dept_rows:
            nl = len(str(r.get("name", "")))
            if nl > max_name_len:
                max_name_len = nl
        ws.column_dimensions["A"].width = max_name_len + 2

        # Each date column width = max(display_len_of_day_number, max(len(cellvalue)))+2
        for i, ds in enumerate(dates):
            col = 2 + i
            col_letter = get_column_letter(col)
            # display len of day number = len(str(int(day))) where day is dd part
            # ds format YYYY-MM-DD, day = int(ds[8:10])
            day_num = int(ds[8:10])
            day_len = len(str(day_num))
            max_status_len = day_len
            for r in dept_rows:
                days = r.get("days", {})
                v = str(days.get(ds, ""))
                vl = len(v)
                if vl > max_status_len:
                    max_status_len = vl
            # need at least day_len, but also if max_status_len is 0 (empty) we still keep day_len
            ws.column_dimensions[col_letter].width = max_status_len + 2

        # FREEZE
        ws.freeze_panes = "B3"

        # PRINT SETUP
        ws.page_setup.orientation = "landscape"
        ws.page_setup.paperSize = 9
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
        ws.print_title_rows = "1:2"
        ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)

    return wb
