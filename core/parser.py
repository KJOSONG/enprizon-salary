"""
Excel 解析引擎
解析 Attendance+data+daily+and+piece+rate.xlsx 的 3 个 sheet
全部列索引通过扫描表头获取，不依赖硬编码列号
"""
import re
import openpyxl
from collections import defaultdict
from .namematch import canonical, split_names, make_employee_id

# ── 产量解析 ─────────────────────────────────────────────
def parse_prod_string(s):
    """解析产量字符串 'MAWE:43 | NICKEL（H）:0 | ...' → dict"""
    r = {'MAWE': 0, 'NICKEL（H）': 0, 'NICKEL（L）': 0,
         'FUTA': 0, 'WAYA': 0, 'KIBIRITI': 0}
    if not s or not isinstance(s, str):
        return r
    for item in re.split(r'\s*\|\s*', s):
        m = re.match(r'(.+?)\s*[：:]\s*([\d.]+)', item.strip())
        if m:
            k, v = m.group(1).strip(), float(m.group(2))
            if 'MAWE' in k and 'NICKEL' not in k:
                r['MAWE'] += v
            elif 'NICKEL' in k and ('H' in k or 'H）' in k or 'H)' in k):
                r['NICKEL（H）'] += v
            elif 'NICKEL' in k and ('L' in k or 'L）' in k or 'L)' in k):
                r['NICKEL（L）'] += v
            elif 'FUTA' in k:
                r['FUTA'] = v
            elif 'WAYA' in k:
                r['WAYA'] = v
            elif 'KIBIRITI' in k:
                r['KIBIRITI'] = v
    return r

# ── 表头扫描 ──────────────────────────────────────────────

def _norm_hdr(val):
    """规范化表头文字用于匹配"""
    if not val:
        return ''
    s = re.sub(r'\s+', '', str(val).strip().upper())
    return s

def _build_col_map(ws):
    """
    扫描第 1 行建立 { 规范化表头: 列号 } 映射
    """
    col_map = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(1, col).value
        if val:
            col_map[_norm_hdr(val)] = col
    return col_map

def _get_col(col_map, *keys, fallback=None):
    """从 col_map 中查找第一个匹配的规范化 key，支持多个候选"""
    for k in keys:
        nk = _norm_hdr(k)
        if nk in col_map:
            return col_map[nk]
    return fallback

def _find_driller_teams(ws, col_map):
    """
    扫描表头，通过 'DrillerTeamNameXX！！！' 和 'DrillerTeamXXMember' 匹配
    """
    team_names = {}
    team_members = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(1, col).value
        if not val:
            continue
        nk = _norm_hdr(val)
        m = re.match(r'DRILLERTEAMNAME(\d+)', nk)
        if m:
            team_names[int(m.group(1))] = col
        m2 = re.match(r'DRILLERTEAM(\d+)MEMBER', nk)
        if m2:
            team_members[int(m2.group(1))] = col
    result = []
    for num in sorted(team_names):
        cap_col = team_names[num]
        result.append({
            'cap_col': cap_col,
            'prod_col': cap_col + 1,
            'mem_col': team_members.get(num, cap_col + 12),
        })
    return result

# ── 默认回退列（当扫描不到时使用） ────────────────────
_DR_COLS = [(20, 21, 32), (22, 23, 33), (24, 25, 34), (26, 27, 35)]

# ═══════════════════════════════════════════════════════════
#  Sheet 1: Piece Rate salary attendance EV
# ═══════════════════════════════════════════════════════════

def parse_piece_rate_sheet(ws):
    """
    解析计件表 → { shift_production, driller_production, piece_rate_people, dates }
    列索引通过表头扫描获取
    """
    cm = _build_col_map(ws)
    date_col = _get_col(cm, 'DATE', '日期', fallback=4)
    day_prod_col = _get_col(cm, 'DAYSHIFT', 'DAY', fallback=10)
    night_prod_col = _get_col(cm, 'NIGHTSHIFT', 'NIGHT', fallback=15)
    day_emps_col = _get_col(cm, 'ATTENDANCEPERSONNEL', fallback=14)
    night_emps_col = _get_col(cm, 'ATTENDANCEPERSONNEL1', fallback=19)

    # 钻工队伍
    driller_teams = _find_driller_teams(ws, cm)
    if not driller_teams:
        driller_teams = [{'cap_col': c, 'prod_col': p, 'mem_col': m}
                         for c, p, m in _DR_COLS]

    shift_production = []
    driller_production = []
    driller_people = set()
    underground_people = set()

    for row in range(2, ws.max_row + 1):
        date_val = ws.cell(row, date_col).value
        if not date_val:
            continue
        date_str = str(date_val)[:10] if not isinstance(date_val, str) else date_val

        # ── Part A: 班次产量+出勤 ──
        day_prod_str = ws.cell(row, day_prod_col).value
        night_prod_str = ws.cell(row, night_prod_col).value
        day_emps_str = ws.cell(row, day_emps_col).value
        night_emps_str = ws.cell(row, night_emps_col).value

        day_prod = parse_prod_string(str(day_prod_str)) if day_prod_str else None
        night_prod = parse_prod_string(str(night_prod_str)) if night_prod_str else None
        day_emps = split_names(day_emps_str) if day_emps_str else []
        night_emps = split_names(night_emps_str) if night_emps_str else []

        for n in day_emps + night_emps:
            underground_people.add(n)

        shift_production.append({
            'date': date_str,
            'day_prod': day_prod,
            'night_prod': night_prod,
            'day_emps': day_emps,
            'night_emps': night_emps,
        })

        # ── Part B: 钻工队伍 ──
        for team in driller_teams:
            cap_raw = ws.cell(row, team['cap_col']).value
            if not cap_raw:
                continue
            cap_cn = canonical(cap_raw)
            if not cap_cn:
                continue

            prod_str = ws.cell(row, team['prod_col']).value
            prod = parse_prod_string(str(prod_str)) if prod_str else None

            mem_str = ws.cell(row, team['mem_col']).value
            members = split_names(mem_str) if mem_str else []

            driller_people.add(cap_cn)
            for m in members:
                driller_people.add(m)

            driller_production.append({
                'date': date_str,
                'captain': cap_cn,
                'slot': team['cap_col'],
                'nh': prod['NICKEL（H）'] if prod else 0,
                'nl': prod['NICKEL（L）'] if prod else 0,
                'mw': prod['MAWE'] if prod else 0,
                'futa': prod['FUTA'] if prod else 0,
                'waya': prod['WAYA'] if prod else 0,
                'kibiriti': prod['KIBIRITI'] if prod else 0,
                'members': members,
                'has_members': bool(members),
            })

    dates = sorted(set(d['date'] for d in shift_production))

    return {
        'shift_production': shift_production,
        'driller_production': driller_production,
        'piece_rate_people': {
            'driller': driller_people,
            'underground': underground_people,
        },
        'dates': dates,
    }

# ═══════════════════════════════════════════════════════════
#  Sheet 2: Daily salary attendance EVERY D
# ═══════════════════════════════════════════════════════════

def parse_daily_salary_sheet(ws):
    """解析日薪考勤表，列索引通过表头扫描获取"""
    cm = _build_col_map(ws)
    date_col = _get_col(cm, 'DATE', '日期', fallback=4)
    normal_col = _get_col(cm, 'NORMALATTENDANCE', 'NORMAL', fallback=6)
    leave_col = _get_col(cm, 'ABNORMALATTENDANCE', 'LEAVE', '请假', fallback=7)
    absent_col = _get_col(cm, 'ABSENTEEISM', 'ABSENT', '旷工', fallback=8)

    people = set()
    attendance = []

    for row in range(2, ws.max_row + 1):
        date_val = ws.cell(row, date_col).value
        if not date_val:
            continue
        date_str = str(date_val)[:10]

        normal_str = ws.cell(row, normal_col).value
        leave_str = ws.cell(row, leave_col).value
        absent_str = ws.cell(row, absent_col).value

        normal_emps = split_names(normal_str) if normal_str else []
        leave_emps = split_names(leave_str) if leave_str else []
        absent_emps = split_names(absent_str) if absent_str else []

        for n in normal_emps + leave_emps + absent_emps:
            people.add(n)

        attendance.append({
            'date': date_str,
            'normal': normal_emps,
            'leave': leave_emps,
            'absent': absent_emps,
        })

    return {'daily_salary_people': people, 'attendance': attendance}

# ═══════════════════════════════════════════════════════════
#  破碎计件文件解析 (CRUSH TEAM Production Data)
# ═══════════════════════════════════════════════════════════

def parse_crush_sheet(filepath):
    """
    解析破碎计件文件 (CRUSH TEAM Production Data_精简.xlsx)
    Sheet: CRUSH TEAM Production Data
    列: Date | How many Bgas | Attendance personnel
    单价: 300 TZS/bag (由 calculator 端统一处理)
    返回: [{date, bags, personnel}, ...]
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet_names = wb.sheetnames
    ws = None
    if sheet_names and sheet_names[0]:
        ws = wb[sheet_names[0]]
    if ws is None:
        wb.close()
        return []

    cm = _build_col_map(ws)
    date_col = _get_col(cm, 'DATE', fallback=1)
    bags_col = _get_col(cm, 'HOWMANYBGAS', 'BAGAS', 'BAGS', fallback=2)
    person_col = _get_col(cm, 'ATTENDANCEPERSONNEL', fallback=3)

    result = []
    for row in range(2, ws.max_row + 1):
        date_val = ws.cell(row, date_col).value
        if not date_val:
            continue
        date_str = str(date_val)[:10] if not isinstance(date_val, str) else date_val

        bags = ws.cell(row, bags_col).value
        try:
            bags = int(bags) if bags else 0
        except (ValueError, TypeError):
            bags = 0

        person_str = ws.cell(row, person_col).value
        personnel = split_names(person_str) if person_str else []

        result.append({
            'date': date_str,
            'bags': bags,
            'personnel': personnel,
        })

    wb.close()
    return result

# ═══════════════════════════════════════════════════════════
#  主解析入口
# ═══════════════════════════════════════════════════════════

def parse_all(filepath):
    """解析完整主文件 → dict"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    result = {}

    if 'Piece Rate salary attendance EV' in wb.sheetnames:
        result.update(parse_piece_rate_sheet(wb['Piece Rate salary attendance EV']))

    if 'Daily salary attendance EVERY D' in wb.sheetnames:
        result.update(parse_daily_salary_sheet(wb['Daily salary attendance EVERY D']))

    wb.close()
    return result
